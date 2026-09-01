from __future__ import annotations

from typing import Optional
import re
from io import BytesIO
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.deps import get_current_user, has_permission, require_permissions
from app.models import (
    CarryOverDecision,
    Comment,
    CommentAttachment,
    CommentStatus,
    CipherTemplate,
    CipherTemplateField,
    Document,
    DocumentAttachment,
    MDRRecord,
    Notification,
    Project,
    ProjectReference,
    Revision,
    ReviewEvent,
    RevisionReviewerState,
    User,
)
from app.services import matrix_gap
from app.schemas import (
    CipherTemplateFieldRead,
    CipherTemplateRead,
    CipherTemplateUpsert,
    MDRChildCreate,
    MDRCreate,
    MDRRead,
    MDRUpdate,
)

router = APIRouter()


class ComposeCipherPayload(BaseModel):
    project_code: str
    category: str | None = None
    values: dict[str, str] = Field(default_factory=dict)
    originator_code: str | None = None
    title_object: str | None = None
    discipline_code: str | None = None
    doc_type: str | None = None
    serial_number: str | None = None


class ComposeCipherResponse(BaseModel):
    cipher: str
    rule: str


class CheckCipherResponse(BaseModel):
    exists: bool


class MdrImportError(BaseModel):
    row: int
    message: str


def _compose_legacy_cipher(payload: ComposeCipherPayload) -> tuple[str, str]:
    if not payload.category:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="category is required")
    if not payload.originator_code or not payload.title_object or not payload.discipline_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Not enough fields for legacy cipher")
    category = payload.category.upper()
    title_number = payload.title_object.strip().upper()
    section_code = payload.discipline_code.strip().upper()
    part_code = (payload.doc_type or "").strip()
    book_code = (payload.serial_number or "").strip()

    if category == "PD":
        # PD: IMP-CTR-PD-0001-AR(.1)(.1.2) — раздел из справочника pd_section.
        if not re.fullmatch(r"\d{4}", title_number):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"title_object must be 4 digits for {category}")
        if not re.fullmatch(r"[A-Z0-9]{2,8}", section_code):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="discipline_code must be 2-8 alnum chars (section/report code from reference)",
            )
        if part_code and not re.fullmatch(r"\d{1,2}", part_code):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="doc_type must be 1-2 digits (part)")
        if book_code and not re.fullmatch(r"[0-9.]{1,5}", book_code):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="serial_number must be 1-5 chars [0-9.] (book)")

        section_part = f"{section_code}{part_code}" if part_code else section_code
        cipher = f"{payload.project_code.upper()}-{payload.originator_code.upper()}-{category}-{title_number}-{section_part}"
        if book_code:
            cipher = f"{cipher}.{book_code}"
        return cipher, f"{category}_SIMPLE"

    # Все категории кроме PD (SE и далее — рабочая документация получит
    # свою маску отдельно): IMP-CTR-SE-1001-SE-REP-001 — код Титульного
    # объекта, Дисциплина (для SE — вид отчёта из справочника
    # se_reporting_type, для остальных — из общего справочника discipline),
    # тип документов (справочник document_type), порядковый номер —
    # уникальный для комбинации project+category+title_object+discipline+doc_type.
    doc_type_code = (payload.doc_type or "").strip().upper()
    serial_number = (payload.serial_number or "").strip()
    if not title_number or not re.fullmatch(r"[A-Z0-9]{1,10}", title_number):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="title_object must be 1-10 alnum chars")
    if not re.fullmatch(r"[A-Z0-9]{2,8}", section_code):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="discipline_code must be 2-8 alnum chars")
    if not doc_type_code or not re.fullmatch(r"[A-Z0-9]{2,8}", doc_type_code):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="doc_type must be 2-8 alnum chars (from document_type reference)")
    if not serial_number or not re.fullmatch(r"\d{1,6}", serial_number):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="serial_number must be 1-6 digits")

    parts = [
        payload.project_code.upper(),
        payload.originator_code.upper(),
        category,
        title_number,
        section_code,
        doc_type_code,
        serial_number,
    ]
    return "-".join(parts), f"{category}_GENERIC"


def _template_read(template: CipherTemplate, project_code: str, db: Session) -> CipherTemplateRead:
    fields = (
        db.query(CipherTemplateField)
        .filter(CipherTemplateField.template_id == template.id)
        .order_by(CipherTemplateField.order_index.asc(), CipherTemplateField.id.asc())
        .all()
    )
    return CipherTemplateRead(
        id=template.id,
        project_id=template.project_id,
        project_code=project_code,
        category=template.category,
        fields=[CipherTemplateFieldRead.model_validate(item, from_attributes=True) for item in fields],
        created_at=template.created_at,
        updated_at=template.updated_at,
    )


def _next_serial_for_project_category(db: Session, project_code: str, category: str, length: int | None) -> str:
    rows = (
        db.query(MDRRecord.serial_number)
        .filter(MDRRecord.project_code == project_code, MDRRecord.category == category)
        .all()
    )
    max_idx = 0
    for row in rows:
        raw = (row[0] or "").strip()
        if raw.isdigit():
            max_idx = max(max_idx, int(raw))
    value = str(max_idx + 1)
    if length:
        return value.zfill(length)[-length:]
    return value


def _template_fields_for_project_category(db: Session, project: Project, category: str) -> list[CipherTemplateField]:
    template = (
        db.query(CipherTemplate)
        .filter(CipherTemplate.project_id == project.id, CipherTemplate.category == category.upper())
        .first()
    )
    if template is None:
        return []
    return (
        db.query(CipherTemplateField)
        .filter(CipherTemplateField.template_id == template.id)
        .order_by(CipherTemplateField.order_index.asc(), CipherTemplateField.id.asc())
        .all()
    )


def _compose_by_template(db: Session, payload: ComposeCipherPayload, template: CipherTemplate) -> tuple[str, str]:
    project = db.query(Project).filter(Project.id == template.project_id).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Template project not found")
    category = template.category
    fields = (
        db.query(CipherTemplateField)
        .filter(CipherTemplateField.template_id == template.id)
        .order_by(CipherTemplateField.order_index.asc(), CipherTemplateField.id.asc())
        .all()
    )
    if not fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cipher template has no fields")

    refs_by_type: dict[str, set[str]] = {}
    for ref in db.query(ProjectReference).filter(ProjectReference.project_id == project.id, ProjectReference.is_active.is_(True)).all():
        refs_by_type.setdefault(ref.ref_type, set()).add(ref.code.upper())

    parts: list[str] = []
    for item in fields:
        value = ""
        if item.source_type == "STATIC":
            value = item.static_value or ""
        elif item.source_type == "AUTO_SERIAL":
            value = _next_serial_for_project_category(db, project.code, category, item.length)
        else:
            value = (payload.values.get(item.field_key) or "").strip()
            if item.required and not value:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Field is required: {item.label}")
            if item.source_type == "REFERENCE" and item.source_ref_type:
                allowed = refs_by_type.get(item.source_ref_type, set())
                if value.upper() not in allowed:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Field must be from reference {item.source_ref_type}: {item.label}",
                    )
        if item.uppercase:
            value = value.upper()
        if item.length:
            value = value[: item.length]
        parts.append(value)
        if item.separator and item.separator != "NONE":
            parts.append(item.separator)

    cipher = "".join(parts).rstrip("-_/ ")
    return cipher, f"TEMPLATE:{category}"


def _validate_weight_limit(
    db: Session,
    *,
    project_code: str,
    category: str,
    new_weight: float,
    exclude_id: int | None = None,
) -> None:
    # Вес 1000 считается отдельно на каждую категорию документации
    # (1000 у PD, свои 1000 у SE и т.д.), не на весь проект разом.
    query = db.query(func.coalesce(func.sum(MDRRecord.doc_weight), 0.0)).filter(
        MDRRecord.project_code == project_code,
        MDRRecord.category == category,
    )
    if exclude_id is not None:
        query = query.filter(MDRRecord.id != exclude_id)
    current_total = float(query.scalar() or 0.0)
    if current_total + new_weight > 1000:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Weight limit exceeded for {project_code}/{category}: {current_total + new_weight:.2f} > 1000",
        )


def _latest_effective_review_code(db: Session, mdr_id: int) -> str | None:
    """Код замечаний по последней ревизии документа этого MDR: явный
    review_code ревизии, иначе «худший» из опубликованных родительских
    замечаний (RJ>CO>AN>AP). None — если ревизий/кодов нет."""
    doc = db.query(Document).filter(Document.mdr_id == mdr_id).first()
    if doc is None:
        return None
    latest = (
        db.query(Revision)
        .filter(Revision.document_id == doc.id)
        .order_by(Revision.id.desc())
        .first()
    )
    if latest is None:
        return None
    if latest.review_code is not None:
        return latest.review_code.value if hasattr(latest.review_code, "value") else str(latest.review_code)
    rows = (
        db.query(Comment.review_code)
        .filter(
            Comment.revision_id == latest.id,
            Comment.parent_id.is_(None),
            Comment.is_published_to_contractor.is_(True),
            Comment.status != CommentStatus.REJECTED,
            Comment.review_code.isnot(None),
        )
        .all()
    )
    vals = {(c[0].value if hasattr(c[0], "value") else str(c[0])) for c in rows if c[0] is not None}
    for code in ("RJ", "CO", "AN", "AP"):
        if code in vals:
            return code
    return None


@router.get("", response_model=list[MDRRead])
def list_mdr(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    project_code: str | None = Query(default=None),
):
    query = db.query(MDRRecord)
    if project_code:
        query = query.filter(MDRRecord.project_code == project_code)
    rows = query.order_by(MDRRecord.id.desc()).all()
    result: list[MDRRead] = []
    for item in rows:
        data = MDRRead.model_validate(item, from_attributes=True)
        data.latest_effective_review_code = _latest_effective_review_code(db, item.id)
        result.append(data)
    return result


@router.get("/{mdr_id}", response_model=MDRRead)
def get_mdr(
    mdr_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    mdr = db.query(MDRRecord).filter(MDRRecord.id == mdr_id).first()
    if not mdr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    return mdr


@router.post("/compose-cipher", response_model=ComposeCipherResponse)
def compose_cipher(
    payload: ComposeCipherPayload,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.code == payload.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown project_code")
    category = (payload.category or project.document_category or "").upper()
    payload.category = category
    cipher, rule = _compose_legacy_cipher(payload)
    return ComposeCipherResponse(cipher=cipher, rule=rule)


@router.get("/cipher-template", response_model=Optional[CipherTemplateRead])
def get_cipher_template(
    project_code: str = Query(...),
    category: str = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.code == project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    template = (
        db.query(CipherTemplate)
        .filter(CipherTemplate.project_id == project.id, CipherTemplate.category == category.upper())
        .first()
    )
    if template is None:
        return None
    return _template_read(template, project.code, db)


@router.put("/cipher-template", response_model=CipherTemplateRead)
def upsert_cipher_template(
    payload: CipherTemplateUpsert,
    project_code: str = Query(...),
    category: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin" and not has_permission(current_user, "can_edit_project_references"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to edit cipher template")
    project = db.query(Project).filter(Project.code == project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    selected_category = category.upper()
    template = (
        db.query(CipherTemplate)
        .filter(CipherTemplate.project_id == project.id, CipherTemplate.category == selected_category)
        .first()
    )
    if template is None:
        template = CipherTemplate(project_id=project.id, category=selected_category)
        db.add(template)
        db.flush()

    db.query(CipherTemplateField).filter(CipherTemplateField.template_id == template.id).delete(synchronize_session=False)
    for idx, item in enumerate(payload.fields):
        db.add(
            CipherTemplateField(
                template_id=template.id,
                order_index=item.order_index if item.order_index is not None else idx,
                field_key=item.field_key,
                label=item.label,
                source_type=item.source_type,
                source_ref_type=item.source_ref_type,
                static_value=item.static_value,
                length=item.length,
                required=item.required,
                uppercase=item.uppercase,
                separator=item.separator,
            )
        )
    db.add(template)
    db.commit()
    db.refresh(template)
    return _template_read(template, project.code, db)


@router.get("/check-cipher", response_model=CheckCipherResponse)
def check_cipher(
    project_code: str = Query(...),
    value: str = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    exists = (
        db.query(MDRRecord.id)
        .filter(MDRRecord.project_code == project_code, MDRRecord.doc_number == value)
        .first()
        is not None
    )
    return CheckCipherResponse(exists=exists)


def _ensure_reviewers_assigned(
    db: Session,
    *,
    project: Project,
    current_user: User,
    doc_number: str,
    category: str | None,
    discipline_code: str | None,
    doc_type: str | None,
) -> None:
    """Не даём завести документ, по которому некому работать.

    Без LR в матрице ревизия уходит «в никуда»: рассылка сваливается в fallback
    «всем сотрудникам заказчика», а согласовать документ и отправить CRS никто
    не может. Админа не блокируем — он и заполняет матрицу.
    """
    if current_user.role.value == "admin":
        return
    discipline = matrix_gap.matrix_discipline(category=category, discipline_code=discipline_code)
    if matrix_gap.has_lead_reviewer(db, project_id=project.id, discipline=discipline, category=category):
        return
    matrix_gap.notify_gap(
        db,
        project=project,
        requester=current_user,
        doc_number=doc_number,
        discipline=discipline,
        category=category,
    )
    raise HTTPException(
        status_code=status.HTTP_409_CONFLICT,
        detail=matrix_gap.gap_detail(discipline=discipline, category=category),
    )


@router.post("", response_model=MDRRead, status_code=status.HTTP_201_CREATED)
def create_mdr(
    payload: MDRCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    project = db.query(Project).filter(Project.code == payload.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown project_code")
    if project.document_category and payload.category.upper() != project.document_category.upper():
        # Категория проекта — основная, но допускаются и другие АКТИВНЫЕ
        # категории из справочника document_category (например SE —
        # инженерные изыскания живут рядом с ПД в одном проекте).
        allowed = (
            db.query(ProjectReference.id)
            .filter(
                ProjectReference.project_id == project.id,
                ProjectReference.ref_type == "document_category",
                ProjectReference.code == payload.category.upper(),
                ProjectReference.is_active.is_(True),
            )
            .first()
        )
        if allowed is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Category must match project category ({project.document_category}) or an active document_category reference",
            )

    exists = db.query(MDRRecord).filter(MDRRecord.document_key == payload.document_key).first()
    if exists:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="document_key already exists")

    exists_cipher = (
        db.query(MDRRecord.id)
        .filter(MDRRecord.project_code == payload.project_code, MDRRecord.doc_number == payload.doc_number)
        .first()
    )
    if exists_cipher:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="doc_number already exists in project")

    _validate_weight_limit(
        db,
        project_code=payload.project_code,
        category=payload.category.upper(),
        new_weight=payload.doc_weight,
    )

    _ensure_reviewers_assigned(
        db,
        project=project,
        current_user=current_user,
        doc_number=payload.doc_number,
        category=payload.category,
        discipline_code=payload.discipline_code,
        doc_type=payload.doc_type,
    )

    mdr = MDRRecord(**payload.model_dump())
    db.add(mdr)
    db.commit()
    db.refresh(mdr)
    return mdr


def _next_serial(value: str) -> str:
    """Следующий порядковый номер шифра с сохранением ширины: 001 → 002, 09 → 10.

    Нужен, когда несколько вложенных идут под одним титулом и дисциплиной:
    шифры совпадают, и различает их только последний сегмент маски.
    Нечисловой номер наращиваем суффиксом-цифрой (ABC → ABC1)."""
    raw = (value or "").strip()
    if raw.isdigit():
        return str(int(raw) + 1).zfill(len(raw))
    return f"{raw}1" if raw else "1"


@router.post("/{parent_id}/child", response_model=MDRRead, status_code=status.HTTP_201_CREATED)
def create_child_mdr(
    parent_id: int,
    payload: MDRChildCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    """Вложенный документ под родительским (напр. программа изысканий под
    отчётом). Шифровочные поля берём у родителя, шифр = шифр родителя + "-" +
    порядковый. Создаёт и MDRRecord, и Document — чтобы карточка и цикл
    рассмотрения работали так же, как у обычного документа."""
    parent = db.query(MDRRecord).filter(MDRRecord.id == parent_id).first()
    if parent is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Parent MDR not found")
    if parent.parent_id is not None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вложенность только один уровень (нельзя под вложенным)")

    # Порядковый номер ребёнка: задан или авто = следующий среди детей.
    serial = (payload.serial or "").strip()
    if serial:
        if not re.fullmatch(r"[A-Za-z0-9]{1,4}", serial):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="serial: 1-4 буквы/цифры")
    else:
        siblings = db.query(MDRRecord.serial_number).filter(MDRRecord.parent_id == parent.id).all()
        max_idx = 0
        for row in siblings:
            raw = (row[0] or "").strip()
            if raw.isdigit():
                max_idx = max(max_idx, int(raw))
        serial = str(max_idx + 1).zfill(2)

    # Свои шифровочные поля: вложенный получает не «шифр родителя + -NN», а
    # полный шифр по маске категории (IMP-SHP-PF-9505-SE-ACT-001). Так уже
    # заведены документы под «Концепцией АБЗ» — титул там номер здания.
    # Состав ревьюверов при этом всё равно берётся от родителя
    # (см. _matrix_source в routers/documents.py).
    child_title = (payload.title_object or "").strip().upper() or (parent.title_object or "").strip().upper()
    child_discipline = (payload.discipline_code or "").strip().upper() or (parent.discipline_code or "").strip().upper()
    child_doc_type = (payload.doc_type or "").strip().upper() or (parent.doc_type or "").strip().upper()
    own_cipher_fields = (
        child_title != (parent.title_object or "").strip().upper()
        or child_discipline != (parent.discipline_code or "").strip().upper()
        or child_doc_type != (parent.doc_type or "").strip().upper()
    )

    def _cipher_taken(value: str) -> bool:
        return (
            db.query(MDRRecord.id)
            .filter(MDRRecord.project_code == parent.project_code, MDRRecord.doc_number == value)
            .first()
            is not None
        )

    if own_cipher_fields:
        # Порядковый номер: свой, если задан, иначе родительский. Если такой
        # шифр уже занят (несколько частей с одним титулом и дисциплиной) —
        # инкрементим последний сегмент маски, пока не найдём свободный.
        # Постфикс не добавляем: шифр должен оставаться разбираемым по маске.
        base_serial = serial if payload.serial else (parent.serial_number or "").strip()
        child_serial = base_serial
        child_doc_number = ""
        for _ in range(1000):
            child_doc_number, _kind = _compose_legacy_cipher(
                ComposeCipherPayload(
                    project_code=parent.project_code,
                    originator_code=parent.originator_code,
                    category=parent.category,
                    title_object=child_title,
                    discipline_code=child_discipline,
                    doc_type=child_doc_type,
                    serial_number=child_serial,
                )
            )
            if not _cipher_taken(child_doc_number):
                break
            child_serial = _next_serial(child_serial)
        else:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Не удалось подобрать свободный порядковый номер — укажите его вручную",
            )
    else:
        child_serial = serial
        child_doc_number = f"{parent.doc_number}-{serial}"
        if _cipher_taken(child_doc_number):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Вложенный документ {child_doc_number} уже существует",
            )

    # Уникальный document_key (DOC-XXXX).
    max_key = 0
    for (key,) in db.query(MDRRecord.document_key).all():
        m = re.fullmatch(r"DOC-(\d+)", (key or "").upper())
        if m:
            max_key = max(max_key, int(m.group(1)))
    document_key = f"DOC-{str(max_key + 1).zfill(4)}"

    _validate_weight_limit(
        db,
        project_code=parent.project_code,
        category=parent.category.upper(),
        new_weight=payload.doc_weight or 0,
    )

    # Матрица у вложенного — родительская (см. _matrix_source), поэтому и
    # проверяем раздел родителя, а не тот, что задан в шифре части.
    parent_project = db.query(Project).filter(Project.code == parent.project_code).first()
    if parent_project is not None:
        _ensure_reviewers_assigned(
            db,
            project=parent_project,
            current_user=current_user,
            doc_number=child_doc_number,
            category=parent.category,
            discipline_code=parent.discipline_code,
            doc_type=parent.doc_type,
        )

    child = MDRRecord(
        document_key=document_key,
        project_code=parent.project_code,
        originator_code=parent.originator_code,
        category=parent.category,
        title_object=child_title or parent.title_object,
        discipline_code=child_discipline or parent.discipline_code,
        doc_type=child_doc_type or parent.doc_type,
        serial_number=child_serial,
        doc_number=child_doc_number,
        parent_id=parent.id,
        doc_name=payload.doc_name,
        planned_dev_start=payload.planned_dev_start,
        progress_percent=0,
        doc_weight=payload.doc_weight or 0,
        dates={},
        status="DRAFT",
        is_confidential=parent.is_confidential,
    )
    db.add(child)
    db.flush()

    db.add(
        Document(
            mdr_id=child.id,
            document_num=child_doc_number,
            title=payload.doc_name,
            discipline=parent.discipline_code,
            weight=payload.doc_weight or 0,
            created_by_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(child)
    return child


@router.put("/{mdr_id}", response_model=MDRRead)
def update_mdr(
    mdr_id: int,
    payload: MDRUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    mdr = db.query(MDRRecord).filter(MDRRecord.id == mdr_id).first()
    if not mdr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")

    changes = payload.model_dump(exclude_unset=True)
    if "doc_number" in changes and changes["doc_number"] != mdr.doc_number:
        exists_cipher = (
            db.query(MDRRecord.id)
            .filter(
                MDRRecord.project_code == (changes.get("project_code") or mdr.project_code),
                MDRRecord.doc_number == changes["doc_number"],
                MDRRecord.id != mdr.id,
            )
            .first()
        )
        if exists_cipher:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="doc_number already exists in project")

    if "document_key" in changes and changes["document_key"] != mdr.document_key:
        exists_key = db.query(MDRRecord.id).filter(MDRRecord.document_key == changes["document_key"], MDRRecord.id != mdr.id).first()
        if exists_key:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="document_key already exists")

    if "category" in changes and changes["category"] and changes["category"] != mdr.category:
        # Та же проверка, что и при создании: категория проекта либо
        # активная категория из справочника document_category.
        project = db.query(Project).filter(Project.code == mdr.project_code).first()
        if project and project.document_category and changes["category"].upper() != project.document_category.upper():
            allowed = (
                db.query(ProjectReference.id)
                .filter(
                    ProjectReference.project_id == project.id,
                    ProjectReference.ref_type == "document_category",
                    ProjectReference.code == changes["category"].upper(),
                    ProjectReference.is_active.is_(True),
                )
                .first()
            )
            if allowed is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Category must match project category ({project.document_category}) or an active document_category reference",
                )

    def _jsonable(value: object) -> object:
        """История правок пишется в JSON-колонку mdr.dates, поэтому значения
        нужно приводить к сериализуемым: date/datetime и Enum (напр.
        planned_dev_start, review_code) иначе роняли PUT с 500 —
        «Object of type date is not JSON serializable»."""
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if hasattr(value, "value"):  # Enum (ReviewCode и т.п.)
            return value.value
        return value

    changed_fields: dict[str, dict[str, object]] = {}
    for field, value in changes.items():
        old_value = getattr(mdr, field)
        if old_value != value:
            changed_fields[field] = {"from": _jsonable(old_value), "to": _jsonable(value)}

    for field, value in changes.items():
        setattr(mdr, field, value)

    if "doc_weight" in changes or "category" in changes:
        # Смена категории переносит вес в другой бюджет 1000 — перепроверяем.
        _validate_weight_limit(
            db,
            project_code=mdr.project_code,
            category=mdr.category,
            new_weight=mdr.doc_weight,
            exclude_id=mdr.id,
        )

    # documents.document_num/discipline/title/weight дублируют поля MDR —
    # без синхронизации карточка ревизий продолжала бы жить под старым шифром.
    if changed_fields.keys() & {"doc_number", "discipline_code", "doc_name", "doc_weight"}:
        for doc in db.query(Document).filter(Document.mdr_id == mdr.id).all():
            doc.document_num = mdr.doc_number
            doc.discipline = mdr.discipline_code
            doc.title = mdr.doc_name
            doc.weight = mdr.doc_weight

    if changed_fields:
        dates_payload = dict(mdr.dates or {})
        history = list(dates_payload.get("update_history") or [])
        history.append(
            {
                "updated_at": datetime.utcnow().isoformat(),
                "updated_by": current_user.email,
                "changed_fields": changed_fields,
            }
        )
        dates_payload["update_history"] = history
        mdr.dates = dates_payload

    db.add(mdr)
    db.commit()
    db.refresh(mdr)
    return mdr


def _purge_mdr_tree(db: Session, root: MDRRecord) -> list[str]:
    """Удаляет MDR вместе с вложенными и всем, что на них ссылается.

    Порядок обхода FK обязателен: Postgres отбивает удаление IntegrityError'ом,
    а пользователь видит только «ничего не происходит». Уже ловили это на
    document_attachments/carry_over_decisions/notifications; позже так же
    выстрелили review_events, revision_reviewer_states и comment_attachments —
    таблицы, добавленные после первого фикса. Добавляя новую таблицу со ссылкой
    на revisions/comments/documents, дописывай зачистку сюда.

    Возвращает пути файлов, которые можно стереть после успешного коммита.
    """
    file_paths: list[str] = []

    # Вложенные документы (один уровень, см. create_child_mdr) сносим вместе с
    # родителем: иначе у них остаётся висячий parent_id и удаление падает.
    mdr_ids = [root.id] + [
        row[0] for row in db.query(MDRRecord.id).filter(MDRRecord.parent_id == root.id).all()
    ]
    doc_numbers = [
        row[0] for row in db.query(MDRRecord.doc_number).filter(MDRRecord.id.in_(mdr_ids)).all()
    ]

    document_ids = [row[0] for row in db.query(Document.id).filter(Document.mdr_id.in_(mdr_ids)).all()]
    if document_ids:
        revision_ids = [row[0] for row in db.query(Revision.id).filter(Revision.document_id.in_(document_ids)).all()]
        if revision_ids:
            comment_ids = [row[0] for row in db.query(Comment.id).filter(Comment.revision_id.in_(revision_ids)).all()]
            if comment_ids:
                db.query(CommentAttachment).filter(
                    CommentAttachment.comment_id.in_(comment_ids)
                ).delete(synchronize_session=False)
                db.query(CarryOverDecision).filter(
                    CarryOverDecision.source_comment_id.in_(comment_ids)
                ).delete(synchronize_session=False)
            db.query(CarryOverDecision).filter(
                CarryOverDecision.target_revision_id.in_(revision_ids)
            ).delete(synchronize_session=False)
            db.query(ReviewEvent).filter(ReviewEvent.revision_id.in_(revision_ids)).delete(synchronize_session=False)
            db.query(RevisionReviewerState).filter(
                RevisionReviewerState.revision_id.in_(revision_ids)
            ).delete(synchronize_session=False)
            db.query(Notification).filter(Notification.revision_id.in_(revision_ids)).delete(synchronize_session=False)
            db.query(Comment).filter(Comment.revision_id.in_(revision_ids)).delete(synchronize_session=False)
            file_paths += [
                row[0]
                for row in db.query(Revision.file_path).filter(Revision.id.in_(revision_ids)).all()
                if row[0]
            ]
        file_paths += [
            row[0]
            for row in db.query(DocumentAttachment.file_path)
            .filter(DocumentAttachment.document_id.in_(document_ids))
            .all()
            if row[0]
        ]
        db.query(DocumentAttachment).filter(
            DocumentAttachment.document_id.in_(document_ids)
        ).delete(synchronize_session=False)
        db.query(Revision).filter(Revision.document_id.in_(document_ids)).delete(synchronize_session=False)
        db.query(Document).filter(Document.id.in_(document_ids)).delete(synchronize_session=False)

    # Контекстные уведомления по шифрам (без revision_id).
    if doc_numbers:
        db.query(Notification).filter(Notification.document_num.in_(doc_numbers)).delete(synchronize_session=False)

    # Сначала вложенные, потом родитель — parent_id ссылается на mdr_records.
    db.query(MDRRecord).filter(MDRRecord.parent_id == root.id).delete(synchronize_session=False)
    db.delete(root)
    return file_paths


@router.delete("/{mdr_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_mdr(
    mdr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can delete MDR records")

    mdr = db.query(MDRRecord).filter(MDRRecord.id == mdr_id).first()
    if not mdr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")

    file_paths = _purge_mdr_tree(db, mdr)
    db.commit()

    # Файлы чистим только после успешного коммита БД — сбой ФС не ломает удаление.
    import os

    for path in file_paths:
        try:
            os.unlink(path)
        except OSError:
            pass


@router.get("/template")
def download_mdr_template(
    project_code: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    project = db.query(Project).filter(Project.code == project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    category = (project.document_category or "").upper()
    template_fields = _template_fields_for_project_category(db, project, category) if category else []
    dynamic_columns = [item.field_key for item in template_fields if item.source_type != "AUTO_SERIAL"]
    base_columns = [
        "originator_code",
        "title_object",
        "discipline_code",
        "doc_type",
        "serial_number",
        "doc_name",
        "doc_weight",
        "planned_dev_start",
    ]
    columns = [*dynamic_columns]
    for col in base_columns:
        if col not in columns:
            columns.append(col)
    wb = Workbook()
    ws = wb.active
    ws.title = "mdr"
    ws.append(columns)
    sample = {
        "originator_code": "CTR",
        "title_object": "0001",
        "discipline_code": "AR",
        "doc_type": "DWG",
        "serial_number": "0001",
        "doc_name": "Пример документа",
        "doc_weight": 10,
        "planned_dev_start": "2026-01-15",
    }
    ws.append([sample.get(col, "") for col in columns])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="mdr_template_{project_code}.xlsx"'},
    )


@router.get("/export")
def export_mdr(
    project_code: str = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.code == project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    rows = db.query(MDRRecord).filter(MDRRecord.project_code == project_code).order_by(MDRRecord.id.asc()).all()
    category = (project.document_category or "").upper()
    template_fields = _template_fields_for_project_category(db, project, category) if category else []
    dynamic_columns = [item.field_key for item in template_fields if item.source_type != "AUTO_SERIAL"]
    base_columns = [
        "originator_code",
        "title_object",
        "discipline_code",
        "doc_type",
        "serial_number",
        "doc_name",
        "doc_weight",
        "planned_dev_start",
    ]
    columns = [*dynamic_columns]
    for col in base_columns:
        if col not in columns:
            columns.append(col)
    trailing_columns = ["doc_number", "document_key", "status"]
    for col in trailing_columns:
        if col not in columns:
            columns.append(col)

    wb = Workbook()
    ws = wb.active
    ws.title = "mdr"
    ws.append(columns)

    for item in rows:
        row_map = {
            "document_key": item.document_key,
            "project_code": item.project_code,
            "category": item.category,
            "originator_code": item.originator_code,
            "title_object": item.title_object,
            "discipline_code": item.discipline_code,
            "doc_type": item.doc_type,
            "serial_number": item.serial_number,
            "doc_number": item.doc_number,
            "doc_name": item.doc_name,
            "doc_weight": item.doc_weight,
            "planned_dev_start": item.planned_dev_start.isoformat() if item.planned_dev_start else "",
            "status": item.status,
        }
        ws.append([row_map.get(col, "") for col in columns])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="mdr_export_{project_code}.xlsx"'},
    )


@router.post("/import")
def import_mdr(
    project_code: str = Query(...),
    dry_run: bool = Query(False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    project = db.query(Project).filter(Project.code == project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    raw = file.file.read()
    wb = load_workbook(BytesIO(raw))
    ws = wb.active

    category = (project.document_category or "").upper()
    template = (
        db.query(CipherTemplate)
        .filter(CipherTemplate.project_id == project.id, CipherTemplate.category == category)
        .first()
        if category
        else None
    )
    template_fields = _template_fields_for_project_category(db, project, category) if template else []

    refs_by_type: dict[str, set[str]] = {}
    for ref in db.query(ProjectReference).filter(ProjectReference.project_id == project.id, ProjectReference.is_active.is_(True)).all():
        refs_by_type.setdefault(ref.ref_type, set()).add(ref.code.upper())

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Template header is missing")
    headers = [str(value or "").strip() for value in header_cells]
    normalized_headers = [header.lower() for header in headers]
    header_idx = {name: idx for idx, name in enumerate(normalized_headers) if name}

    existing_doc_numbers = {
        value
        for (value,) in db.query(MDRRecord.doc_number).filter(MDRRecord.project_code == project_code).all()
        if value
    }
    reserved_doc_numbers = set(existing_doc_numbers)

    existing_keys = {
        value
        for (value,) in db.query(MDRRecord.document_key).filter(MDRRecord.project_code == project_code).all()
        if value
    }
    next_doc_idx = 1
    for key in existing_keys:
        if key.upper().startswith("DOC-"):
            suffix = key[4:]
            if suffix.isdigit():
                next_doc_idx = max(next_doc_idx, int(suffix) + 1)

    def next_document_key() -> str:
        nonlocal next_doc_idx
        while True:
            candidate = f"DOC-{str(next_doc_idx).zfill(4)}"
            next_doc_idx += 1
            if candidate not in existing_keys:
                existing_keys.add(candidate)
                return candidate

    imported = 0
    skipped = 0
    errors: list[MdrImportError] = []
    gap_disciplines: set[str] = set()
    auto_serial_next_by_field: dict[int, int] = {}

    # Вес 1000 — на категорию (весь импорт идёт под одной category), а не на doc_type.
    weight_totals = {
        row_category: float(total_weight or 0.0)
        for row_category, total_weight in db.query(MDRRecord.category, func.coalesce(func.sum(MDRRecord.doc_weight), 0.0))
        .filter(MDRRecord.project_code == project_code)
        .group_by(MDRRecord.category)
        .all()
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = list(row or [])
        if all(str(cell or "").strip() == "" for cell in row_data):
            skipped += 1
            continue

        def get_value(key: str) -> str:
            idx = header_idx.get(key.lower())
            if idx is None or idx >= len(row_data):
                return ""
            return str(row_data[idx] or "").strip()

        line_errors: list[str] = []
        doc_name = get_value("doc_name")
        if not doc_name:
            line_errors.append("Field is required: doc_name")

        raw_weight = get_value("doc_weight")
        try:
            doc_weight = float(raw_weight) if raw_weight else 0.0
        except ValueError:
            line_errors.append("Field doc_weight must be a number")
            doc_weight = 0.0

        values = {name: get_value(name) for name in normalized_headers if name}

        serial_from_template = ""
        cipher = ""
        originator_code = get_value("originator_code").upper()
        title_object = get_value("title_object").upper()
        discipline_code = get_value("discipline_code").upper()
        doc_type = get_value("doc_type").upper()
        serial_number = get_value("serial_number")

        if template and template_fields:
            parts: list[str] = []
            for field in template_fields:
                value = ""
                if field.source_type == "STATIC":
                    value = field.static_value or ""
                elif field.source_type == "AUTO_SERIAL":
                    if field.id not in auto_serial_next_by_field:
                        seed = _next_serial_for_project_category(db, project.code, category, field.length)
                        auto_serial_next_by_field[field.id] = int(seed) if seed.isdigit() else 1
                    next_serial = auto_serial_next_by_field[field.id]
                    auto_serial_next_by_field[field.id] = next_serial + 1
                    value = str(next_serial)
                    if field.length:
                        value = value.zfill(field.length)[-field.length:]
                    serial_from_template = value
                else:
                    value = (values.get(field.field_key.lower()) or "").strip()
                    if field.required and not value:
                        line_errors.append(f"Field is required: {field.label}")
                    if field.source_type == "REFERENCE" and field.source_ref_type and value:
                        allowed = refs_by_type.get(field.source_ref_type, set())
                        if value.upper() not in allowed:
                            line_errors.append(f"Field must be from reference {field.source_ref_type}: {field.label}")
                if field.uppercase:
                    value = value.upper()
                if field.length:
                    value = value[: field.length]
                parts.append(value)
                if field.separator and field.separator != "NONE":
                    parts.append(field.separator)
            cipher = "".join(parts).rstrip("-_/ ")
            serial_number = serial_number or serial_from_template
        else:
            if not (originator_code and title_object and discipline_code and doc_type and serial_number):
                line_errors.append("Legacy fields are required: originator_code, title_object, discipline_code, doc_type, serial_number")
            if not line_errors:
                payload = ComposeCipherPayload(
                    project_code=project_code,
                    category=category,
                    originator_code=originator_code,
                    title_object=title_object,
                    discipline_code=discipline_code,
                    doc_type=doc_type,
                    serial_number=serial_number,
                    values={
                        "originator_code": originator_code,
                        "title_object": title_object,
                        "discipline_code": discipline_code,
                        "doc_type": doc_type,
                        "serial_number": serial_number,
                    },
                )
                try:
                    cipher, _rule = _compose_legacy_cipher(payload)
                except HTTPException as exc:
                    line_errors.append(str(exc.detail))

        if cipher and cipher in reserved_doc_numbers:
            line_errors.append("doc_number already exists in project")

        planned_dev_start = None
        raw_planned_start = get_value("planned_dev_start")
        if raw_planned_start:
            try:
                planned_dev_start = datetime.fromisoformat(raw_planned_start).date()
            except ValueError:
                line_errors.append("Field planned_dev_start must be ISO date (YYYY-MM-DD)")

        if not line_errors and category:
            projected = weight_totals.get(category, 0.0) + doc_weight
            if projected > 1000:
                line_errors.append(f"Weight limit exceeded for {project_code}/{category}: {projected:.2f} > 1000")

        # Строки без LR в матрице пропускаем с понятной причиной: иначе
        # импорт заводит документы, которые некому рассматривать.
        if not line_errors and current_user.role.value != "admin":
            row_discipline = matrix_gap.matrix_discipline(category=category, discipline_code=discipline_code)
            if not matrix_gap.has_lead_reviewer(
                db, project_id=project.id, discipline=row_discipline, category=category
            ):
                line_errors.append(
                    f"В матрице назначений нет LR по разделу «{row_discipline or '—'}» — "
                    f"обратитесь к администратору системы заказчика"
                )
                gap_disciplines.add(row_discipline or "—")

        if line_errors:
            errors.append(MdrImportError(row=row_idx, message="; ".join(line_errors)))
            skipped += 1
            continue

        document_key = next_document_key()
        reserved_doc_numbers.add(cipher)

        mdr = MDRRecord(
            document_key=document_key,
            project_code=project_code,
            originator_code=originator_code,
            category=category,
            title_object=title_object,
            discipline_code=discipline_code,
            doc_type=doc_type,
            serial_number=serial_number,
            doc_number=cipher,
            doc_name=doc_name,
            planned_dev_start=planned_dev_start,
            progress_percent=0,
            doc_weight=doc_weight,
            dates={},
            status="DRAFT",
            is_confidential=False,
        )
        if not dry_run:
            db.add(mdr)
        weight_totals[category] = weight_totals.get(category, 0.0) + doc_weight
        imported += 1

    if not dry_run:
        db.commit()

    # Одно уведомление на импорт, а не на каждую строку.
    if gap_disciplines and not dry_run:
        for discipline in sorted(gap_disciplines):
            matrix_gap.notify_gap(
                db,
                project=project,
                requester=current_user,
                doc_number=f"импорт реестра ({project_code})",
                discipline=discipline,
                category=None,
            )

    return {
        "dry_run": dry_run,
        "imported": imported,
        "skipped": skipped,
        "errors": [item.model_dump() for item in errors],
    }
