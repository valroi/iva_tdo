from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi import File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, has_permission, is_main_admin
from app.models import (
    Comment,
    CompanyType,
    Document,
    MDRRecord,
    Project,
    ProjectMember,
    ProjectMemberRole,
    ProjectReference,
    Revision,
    ReviewMatrixMember,
    User,
)
from app.schemas import (
    ProjectCreate,
    ProjectMemberCreate,
    ProjectMemberRead,
    ProjectRead,
    ProjectReferenceCreate,
    ProjectReferenceRead,
    ProjectReferenceUpdate,
    ProjectUpdate,
    ReviewMatrixMemberCreate,
    ReviewMatrixMemberRead,
    ReviewMatrixMemberUpdate,
)

router = APIRouter()

DOCUMENT_CATEGORIES: list[tuple[str, str]] = [
    ("PF", "Pre-FEED, предпроектные исследования"),
    ("BEP", "Базовый проект / Basic Engineering Package"),
    ("SE", "Инженерные изыскания / Engineering Survey"),
    ("FD", "FEED / Front End Engineering Design"),
    ("PD", "Проектная документация / Design Documentation"),
    ("DD", "Рабочая документация / Detailed Design Documentation"),
    ("PM", "Документы управления проектом / Project Management Documents"),
]

NUMBERING_ATTRIBUTES: list[tuple[str, str]] = [
    ("AAA", "Буквенный код проекта / Alphabetic project code"),
    ("BBB", "Код компании разработчика / Originator code"),
    ("CC(C)", "Категория документа (стадия) / Document category"),
    ("DDDDDDD", "Титул (номер объекта) / Facility number"),
    ("EE", "Код дисциплины / Discipline code"),
    ("FFFFFF", "Раздел ПД / PD section"),
    ("HH", "Код марки / Marka code"),
    ("JJJ", "Тип документа / Document type"),
    ("TT(T)", "Тип документа для РД / DD document type"),
    ("KKKKKKKKKKKK", "Номер заказа на закупку / Purchase order number"),
    ("LLL", "Код документа поставщика / Vendor document code"),
    ("MMM", "Тип пакета для поставщиков / Supplier package type"),
    ("O(OO)", "Код типа оборудования / Equipment type code"),
    ("PPP", "Код подразделения / Subdivision code"),
    ("RR", "Тип запроса / Request type"),
    ("NNNNN", "Порядковый номер / Sequence number"),
]

DOCUMENT_TYPES: list[tuple[str, str]] = [
    ("ACL", "Перечень поручений / Action List"),
    ("BED", "Основа проектирования / Basic Engineering Design Data"),
    ("BLD", "Блок-схемы / Block Diagrams"),
    ("BOD", "Исходные данные / Basis of Design"),
    ("BQT", "Тендерное предложение / Bid or Quotation"),
    ("BRV", "Оценка тендерных предложений / Bid Reviews"),
    ("CBA", "Оценка коммерческой части / Commercial Bid Evaluation"),
    ("CBL", "Кабельный журнал / Cable list"),
    ("CHK", "Чек-лист / Checklist"),
    ("CLC", "Расчеты / Calculations"),
    ("COD", "Кодекс / Code"),
    ("CON", "Заключение / Conclusion"),
    ("COO", "Запрос на изменение / Change Order"),
    ("DEB", "Основа проектирования / Design Engineering Basis"),
    ("DSH", "Опросные листы / Data Sheets"),
    ("DWG", "Чертеж / Drawings"),
    ("EML", "Электронное письмо / E-Mails"),
    ("EST", "Сметный расчет / Estimate Sheet"),
    ("HMB", "Материальный и тепловой баланс / Heat & Material Balances"),
    ("INV", "Счет-фактура / Invoice"),
    ("IRQ", "Заявка / Inquiry requisition"),
    ("JSD", "ТЗ на проектирование / Job specification for design"),
    ("JSS", "ТЗ на поставку / Job specification for supply"),
    ("LEL", "Извлеченные уроки / Lesson Learned"),
    ("LET", "Письмо / Letter"),
    ("LST", "Перечень / Lists"),
    ("MAN", "Руководство / Manual"),
    ("MDR", "Главный реестр документации / Master Document Register"),
    ("MOD", "Инжиниринговая модель / Engineering Model"),
    ("MOM", "Протокол совещания / Minutes of Meeting"),
    ("MRQ", "Заявка на материалы / Material Requisitions"),
    ("MTO", "Ведомость материалов / Material Take Off"),
    ("MTS", "Основные технические решения / Main technical solutions"),
    ("MTX", "Матрица и таблицы / Matrix"),
    ("NTS", "Пояснительная записка / Notes"),
    ("PFD", "Принципиальная схема / Process flow diagram"),
    ("PHL", "Философия / Philosophy"),
    ("PID", "Схема KИП / Piping and instrumentation diagram"),
    ("PLN", "План / Plan"),
    ("POL", "Политика / Policy"),
    ("POR", "Заявка на закупку / Purchase order"),
    ("PRE", "Презентация / Presentation"),
    ("PRG", "Программа / Program"),
    ("PRV", "Положение / Provision"),
    ("PRO", "Процедура / Procedure"),
    ("REG", "Реестр / Register"),
    ("REP", "Отчет / Reports"),
    ("SCH", "График проекта / Schedules"),
    ("SDO", "Документы поставщиков / Supplier's Documents"),
    ("SHM", "Схемы / Schems"),
    ("SKT", "Мемо / Sketch"),
    ("SOW", "Объем работ / Scope of Work"),
    ("SPE", "Спецификация / Specification"),
    ("STD", "Стандарт / Standard"),
    ("TBE", "Оценка технического предложения / Technical Bid Evaluation"),
    ("TEA", "Техническое задание / Technical Assignment"),
    ("TEQ", "Технический запрос / Technical Query"),
    ("TRM", "Трансмиттал / Transmittal"),
    ("TSC", "Технические условия на подключение / Technical Specification for Connection"),
    ("TSP", "Технические требования / Technical specifications"),
    ("WDM", "Журнал сварочных работ / Welding Map"),
    ("WIT", "Рабочая инструкция / Work Instruction"),
]

DISCIPLINES: list[tuple[str, str]] = [
    ("3D", "3D Модель / 3D Model"),
    ("AD", "Управление контрактом / Contract Management"),
    ("AF", "Администрация и финансы / Administration and Finance"),
    ("AR", "Архитектура / Architecture"),
    ("BL", "Здания / Buildings"),
    ("CI", "Общестроительные дисциплины / Civil"),
    ("CM", "ПНР (вкл. подготовку) / Commissioning"),
    ("CS", "Проектирование конструкций / Structural Engineering"),
    ("CT", "СМР / Construction"),
    ("DC", "Документационный контроль / Document control"),
    ("DR", "Бурение и КРС / Drilling and workover"),
    ("EL", "Электрическая часть / Electrical"),
    ("EP", "Электрохимическая защита / Electrochemical Protection"),
    ("EQ", "Электрооборудование / Electrical Equipment"),
    ("ES", "Сметы / Estimating"),
    ("EW", "Ранние работы / Early works"),
    ("EX", "Экспедирование / Expediting"),
    ("FF", "Пожаротушение / Fire Fighting"),
    ("FI", "Инспектирование изготовления / Fabrication inspection"),
    ("FP", "Пожарозащита / Fire Proofing"),
    ("FS", "Система пожарной автоматики / Fire automation system"),
    ("GE", "Общее / General"),
    ("GO", "Геология / Geology"),
    ("GT", "Геотехнический отчет / Geotechnical report"),
    ("HQ", "Оборудование HVAC / HVAC Equipment"),
    ("HR", "Управление персоналом / Human Resources"),
    ("HS", "ОТ, ПБ и экология / HSE"),
    ("HV", "ОВиК / HVAC"),
    ("IF", "Управление интерфейсами / Interface Management"),
    ("IN", "КИПиА системы / Instrumentation and Control Systems"),
    ("IQ", "Оборудование КИПиА / Instrumentation Equipment"),
    ("IS", "Изоляция / Insulation"),
    ("IT", "Информационные технологии / Information Technology"),
    ("LG", "Юридическое сопровождение / Legal"),
    ("ME", "Механика / Mechanical"),
    ("MM", "Управление материалами / Materials Management"),
    ("MT", "Материалы / Material"),
    ("OP", "Эксплуатация / Operation"),
    ("PA", "Окраска и антикоррозия / Painting"),
    ("PB", "Промышленная и пожарная безопасность / Industrial and fire safety"),
    ("PC", "Технология / Process"),
    ("PE", "Инжиниринг проекта / Project Engineering"),
    ("PI", "Трубопроводы / Piping"),
    ("PM", "Управление проектом / Project Management"),
    ("PN", "Контроль реализации / Project Control"),
    ("PP", "Генплан / Plot Plan"),
    ("PR", "Закупки / Procurement"),
    ("QA", "Обеспечение качества / Quality Assurance"),
    ("QC", "Контроль качества / Quality Control"),
    ("RM", "Управление рисками / Risk Management"),
    ("RQ", "Динамическое оборудование / Rotating Equipment"),
    ("SB", "Субподряды / Subcontracting"),
    ("SC", "Безопасность / Security"),
    ("SE", "Инженерные изыскания / Survey Engineering"),
    ("SI", "Информационная безопасность / Information Security"),
    ("SQ", "Статическое оборудование / Static Equipment"),
    ("SS", "Металлоконструкции / Steel Structures"),
    ("ST", "Конструкции и сооружения / Structural"),
    ("TE", "Телекоммуникации / Telecommunications"),
    ("TL", "Транспорт и логистика / Transport and Logistics"),
    ("WL", "НК и сварка / Welding Technical Support and Testing"),
    ("WS", "Водоснабжение и водоотведение / Water Supply, Water Drainage"),
]

SE_REPORTING_TYPES: list[tuple[str, str]] = [
    ("IGD", "Техотчет по инженерно-геодезическим изысканиям / Engineering-geodesic report"),
    ("IGL", "Техотчет по инженерно-геологическим изысканиям / Engineering-geological report"),
    ("IGM", "Техотчет по инженерно-гидрометеорологическим изысканиям / Hydro-meteorological report"),
    ("IEL", "Техотчет по инженерно-экологическим изысканиям / Engineering-ecological report"),
    ("IKO", "Отчет по историко-культурному обследованию / Archaeology report"),
    ("IGT", "Техотчет по инженерно-геотехническим изысканиям / Geotechnical report"),
    ("IGF", "Техотчет по инженерно-геофизическим изысканиям / Geophysical report"),
]

PROCUREMENT_REQUEST_TYPES: list[tuple[str, str]] = [
    ("MA", "Конъюнктурный анализ / Economic analysis"),
]

EQUIPMENT_TYPE_CODES: list[tuple[str, str]] = [
    ("A", "Мешалки / Mixers"),
    ("B", "Паровой риформинг, горелка / Steam reforming, burner"),
    ("BL", "Воздуходувки и вентиляторы / Blowers and fans"),
    ("C", "Компрессорное оборудование / Compressing equipment"),
    ("CT", "Турбины компрессоров / Compressor turbines"),
    ("D", "Колонны и абсорберы / Towers and absorbers"),
    ("E", "Теплообменники / Heat exchangers"),
    ("EA", "Воздухоохладители / Air coolers"),
    ("EH", "Электронагреватели / Electric heaters"),
    ("F", "Фильтры / Filters"),
    ("J", "Форсунки / Nozzles"),
    ("M", "Электродвигатели / Electric motors"),
    ("P", "Насосы / Pumps"),
    ("PU", "Комплектное оборудование / Package units"),
    ("R", "Реакторы / Reactors"),
    ("S", "Факельные и выхлопные системы / Flare and exhaust"),
    ("SM", "Статические смесители / Static mixers"),
    ("TK", "Резервуары / Tanks"),
    ("U", "Дизельный двигатель / Diesel engine"),
    ("V", "Емкости и сепараторы / Vessels and separators"),
    ("X", "Электрогенератор / Generator"),
    ("Y", "Дозирование химреагентов / Chemical dosing"),
]

IDENTIFIER_PATTERNS: list[tuple[str, str]] = [
    ("SE_BASE", "SE: AAA-BBB-CC(C)-DDDDDDD-EE-JJJ-NNNNN"),
    ("SE_TEXT", "SE текстовая часть: базовый шифр + -Т"),
    ("SE_GRAPH", "SE графическая часть: базовый шифр + -Г.N"),
    ("PD_BASE", "PD: AAA-BBB-CC(C)-DDDDDDD-FFFFFF"),
    ("PD_GRAPHIC", "PD графика: базовый шифр + -00001"),
    ("SM_VOLUME", "Сметы ПД: AAA-BBB-CC(C)-DDDDDDD-SMx"),
    ("EST_LOCAL", "Локальная смета: AAA-BBB-CC(C)-DDDDDDD-FFFFFF-HH-EST-NNNNN"),
    ("SOW", "Ведомость объемов: AAA-BBB-CC(C)-DDDDDDD-FFFFFF-HH-SOW-NNNNN"),
    ("TSC_TEA", "TSC/TEA: AAA-BBB-CC(C)-DDDDDDD-EE-JJJ-NNNNN"),
]

REVIEW_SLA_DAYS: list[tuple[str, str]] = [
    ("*:*:INITIAL", "14"),
    ("*:*:NEXT", "7"),
]


def _default_project_references(selected_document_category: str) -> list[tuple[str, str, str]]:
    refs: list[tuple[str, str, str]] = []
    refs.extend(
        ("document_category", code, value) for code, value in DOCUMENT_CATEGORIES if code == selected_document_category
    )
    refs.extend(("numbering_attribute", code, value) for code, value in NUMBERING_ATTRIBUTES)
    refs.extend(("document_type", code, value) for code, value in DOCUMENT_TYPES)
    refs.extend(("discipline", code, value) for code, value in DISCIPLINES)
    refs.extend(("se_reporting_type", code, value) for code, value in SE_REPORTING_TYPES)
    refs.extend(("procurement_request_type", code, value) for code, value in PROCUREMENT_REQUEST_TYPES)
    refs.extend(("equipment_type", code, value) for code, value in EQUIPMENT_TYPE_CODES)
    refs.extend(("identifier_pattern", code, value) for code, value in IDENTIFIER_PATTERNS)
    refs.extend(("review_sla_days", code, value) for code, value in REVIEW_SLA_DAYS)
    refs.extend(
        [
            ("other", "TRM_RECEIVER_COMPANY_CODE", "IVA"),
        ]
    )
    return refs


def _get_project_or_404(db: Session, project_id: int) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    return project


def _ensure_project_access(db: Session, project_id: int, user: User) -> None:
    if user.role.value == "admin":
        return

    member = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id, ProjectMember.user_id == user.id)
        .first()
    )
    if member is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No project access")


def _is_contractor_tdo_lead(db: Session, project_id: int, user_id: int) -> bool:
    member = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id, ProjectMember.user_id == user_id)
        .first()
    )
    if member is None:
        return False
    return member.member_role == ProjectMemberRole.contractor_tdo_lead or member.can_manage_contractor_users


def _can_manage_project_matrix(db: Session, project_id: int, user: User) -> bool:
    if is_main_admin(user):
        return True
    return _is_contractor_tdo_lead(db, project_id, user.id)


def _project_member_read(member: ProjectMember, db: Session) -> ProjectMemberRead:
    linked = db.query(User).filter(User.id == member.user_id).first()
    return ProjectMemberRead.model_validate(member, from_attributes=True).model_copy(
        update={
            "user_email": linked.email if linked else None,
            "user_full_name": linked.full_name if linked else None,
            "user_company_type": linked.company_type if linked else None,
        }
    )


def _review_matrix_read(item: ReviewMatrixMember, db: Session) -> ReviewMatrixMemberRead:
    linked = db.query(User).filter(User.id == item.user_id).first()
    return ReviewMatrixMemberRead.model_validate(item, from_attributes=True).model_copy(
        update={
            "user_email": linked.email if linked else None,
            "user_full_name": linked.full_name if linked else None,
        }
    )


def _validate_mark_discipline_mapping(
    db: Session,
    *,
    project_id: int,
    discipline_code: str,
    mark_code: str,
) -> None:
    mapping = (
        db.query(ProjectReference)
        .filter(
            ProjectReference.project_id == project_id,
            ProjectReference.ref_type == "mark_discipline",
            ProjectReference.code == mark_code,
            ProjectReference.is_active.is_(True),
        )
        .first()
    )
    if mapping is None:
        return
    if (mapping.value or "").strip().upper() != (discipline_code or "").strip().upper():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Mark {mark_code} is mapped to discipline {mapping.value}. "
                f"Use matching discipline for LR/R assignment."
            ),
        )


@router.get("", response_model=list[ProjectRead])
def list_projects(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value == "admin":
        return db.query(Project).order_by(Project.created_at.desc()).all()

    return (
        db.query(Project)
        .join(ProjectMember, ProjectMember.project_id == Project.id)
        .filter(ProjectMember.user_id == current_user.id)
        .order_by(Project.created_at.desc())
        .all()
    )


@router.post("", response_model=ProjectRead, status_code=status.HTTP_201_CREATED)
def create_project(
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can create projects")

    existing = db.query(Project).filter(Project.code == payload.code).first()
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project code already exists")

    selected_category = payload.document_category.upper().strip()
    category_exists = any(code == selected_category for code, _ in DOCUMENT_CATEGORIES)
    if not category_exists:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown document_category")

    project = Project(
        code=payload.code,
        name=payload.name,
        document_category=selected_category,
        description=payload.description,
        created_by_id=current_user.id,
    )
    db.add(project)
    db.flush()

    db.add(
        ProjectMember(
            project_id=project.id,
            user_id=current_user.id,
            member_role=ProjectMemberRole.main_admin,
            can_manage_contractor_users=True,
        )
    )

    for ref_type, code, value in _default_project_references(selected_category):
        db.add(
            ProjectReference(
                project_id=project.id,
                ref_type=ref_type,
                code=code,
                value=value,
                is_active=True,
            )
        )

    db.commit()
    db.refresh(project)
    return project


@router.put("/{project_id}", response_model=ProjectRead)
def update_project(
    project_id: int,
    payload: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can edit projects")

    project = _get_project_or_404(db, project_id)
    changes = payload.model_dump(exclude_unset=True)

    if "document_category" in changes and changes["document_category"] is not None:
        selected_category = str(changes["document_category"]).upper().strip()
        category_exists = any(code == selected_category for code, _ in DOCUMENT_CATEGORIES)
        if not category_exists:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown document_category")

        mismatched_mdr = (
            db.query(MDRRecord.id)
            .filter(MDRRecord.project_code == project.code, MDRRecord.category != selected_category)
            .first()
        )
        if mismatched_mdr is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot change category: project already has MDR records with another category",
            )
        changes["document_category"] = selected_category

        db.query(ProjectReference).filter(
            ProjectReference.project_id == project.id, ProjectReference.ref_type == "document_category"
        ).delete(synchronize_session=False)
        selected_value = next((value for code, value in DOCUMENT_CATEGORIES if code == selected_category), selected_category)
        db.add(
            ProjectReference(
                project_id=project.id,
                ref_type="document_category",
                code=selected_category,
                value=selected_value,
                is_active=True,
            )
        )

    for field, value in changes.items():
        setattr(project, field, value)

    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project(
    project_id: int,
    purge: bool = Query(default=False),
    confirm_code: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not has_permission(current_user, "can_manage_projects"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Project management permission required")

    project = _get_project_or_404(db, project_id)

    mdr_rows = db.query(MDRRecord.id).filter(MDRRecord.project_code == project.code).all()
    mdr_ids = [row[0] for row in mdr_rows]
    if mdr_ids and not purge:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete project with existing MDR records",
        )
    if purge and (confirm_code or "").strip().upper() != project.code.upper():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"confirm_code must match project code: {project.code}",
        )

    if mdr_ids:
        document_rows = db.query(Document.id).filter(Document.mdr_id.in_(mdr_ids)).all()
        document_ids = [row[0] for row in document_rows]
        if document_ids:
            revision_rows = db.query(Revision.id).filter(Revision.document_id.in_(document_ids)).all()
            revision_ids = [row[0] for row in revision_rows]
            if revision_ids:
                db.query(Comment).filter(Comment.revision_id.in_(revision_ids)).delete(synchronize_session=False)
                db.query(Revision).filter(Revision.id.in_(revision_ids)).delete(synchronize_session=False)
            db.query(Document).filter(Document.id.in_(document_ids)).delete(synchronize_session=False)
        db.query(MDRRecord).filter(MDRRecord.id.in_(mdr_ids)).delete(synchronize_session=False)

    db.query(ProjectReference).filter(ProjectReference.project_id == project_id).delete()
    db.query(ReviewMatrixMember).filter(ReviewMatrixMember.project_id == project_id).delete()
    db.query(ProjectMember).filter(ProjectMember.project_id == project_id).delete()
    db.delete(project)
    db.commit()


@router.get("/{project_id}/members", response_model=list[ProjectMemberRead])
def list_project_members(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id, current_user)
    _get_project_or_404(db, project_id)
    items = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id)
        .order_by(ProjectMember.id.asc())
        .all()
    )
    return [_project_member_read(item, db) for item in items]


@router.post("/{project_id}/members", response_model=ProjectMemberRead, status_code=status.HTTP_201_CREATED)
def add_project_member(
    project_id: int,
    payload: ProjectMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_project_or_404(db, project_id)

    target_user = db.query(User).filter(User.id == payload.user_id).first()
    if target_user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    existing = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id, ProjectMember.user_id == payload.user_id)
        .first()
    )
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="User already in project")

    can_manage = False
    if is_main_admin(current_user):
        can_manage = payload.member_role == ProjectMemberRole.contractor_tdo_lead
    elif _is_contractor_tdo_lead(db, project_id, current_user.id):
        if target_user.company_type != CompanyType.contractor:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="TDO lead can add only contractor users",
            )
        if (target_user.company_code or "").upper() != (current_user.company_code or "").upper():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="TDO lead can add only contractor users with same company code",
            )
        if payload.member_role not in {
            ProjectMemberRole.contractor_member,
            ProjectMemberRole.contractor_tdo_lead,
        }:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="TDO lead can assign only contractor roles",
            )
        can_manage = payload.member_role == ProjectMemberRole.contractor_tdo_lead
    elif has_permission(current_user, "can_manage_users"):
        _ensure_project_access(db, project_id, current_user)
        if target_user.company_type != CompanyType.contractor:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Can add only contractor users",
            )
        if current_user.company_type == CompanyType.contractor and (
            (target_user.company_code or "").upper() != (current_user.company_code or "").upper()
        ):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Can add only contractor users with same company code",
            )
        if payload.member_role not in {
            ProjectMemberRole.contractor_member,
            ProjectMemberRole.contractor_tdo_lead,
        }:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Can assign only contractor roles",
            )
        can_manage = payload.member_role == ProjectMemberRole.contractor_tdo_lead
    else:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to add project members")

    member = ProjectMember(
        project_id=project_id,
        user_id=payload.user_id,
        member_role=payload.member_role,
        can_manage_contractor_users=can_manage,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return _project_member_read(member, db)


@router.delete("/{project_id}/members/{member_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project_member(
    project_id: int,
    member_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    member = (
        db.query(ProjectMember)
        .filter(ProjectMember.id == member_id, ProjectMember.project_id == project_id)
        .first()
    )
    if member is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project member not found")

    target_user = db.query(User).filter(User.id == member.user_id).first()
    if target_user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    if is_main_admin(current_user):
        pass
    elif _is_contractor_tdo_lead(db, project_id, current_user.id):
        if target_user.company_type != CompanyType.contractor:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="TDO lead can remove only contractor users",
            )
        if (target_user.company_code or "").upper() != (current_user.company_code or "").upper():
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="TDO lead can remove only contractor users with same company code",
            )
    elif has_permission(current_user, "can_manage_users"):
        _ensure_project_access(db, project_id, current_user)
        if target_user.company_type != CompanyType.contractor:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Can remove only contractor users",
            )
    else:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to remove project members")

    if member.member_role == ProjectMemberRole.main_admin:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot remove main admin from project")

    db.delete(member)
    db.commit()


@router.get("/{project_id}/references", response_model=list[ProjectReferenceRead])
def list_project_references(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    ref_type: str | None = Query(default=None),
):
    _ensure_project_access(db, project_id, current_user)
    _get_project_or_404(db, project_id)

    query = db.query(ProjectReference).filter(ProjectReference.project_id == project_id)
    if ref_type:
        query = query.filter(ProjectReference.ref_type == ref_type)

    return query.order_by(ProjectReference.ref_type.asc(), ProjectReference.code.asc()).all()


@router.post("/{project_id}/references", response_model=ProjectReferenceRead, status_code=status.HTTP_201_CREATED)
def create_project_reference(
    project_id: int,
    payload: ProjectReferenceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not has_permission(current_user, "can_edit_project_references"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Reference edit permission required")

    _get_project_or_404(db, project_id)

    exists = (
        db.query(ProjectReference)
        .filter(
            ProjectReference.project_id == project_id,
            ProjectReference.ref_type == payload.ref_type,
            ProjectReference.code == payload.code,
        )
        .first()
    )
    if exists is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Reference with this code already exists")

    item = ProjectReference(project_id=project_id, **payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/{project_id}/references/template")
def download_project_references_template(
    project_id: int,
    ref_type: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id, current_user)
    _get_project_or_404(db, project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "references"
    ws.append(["ref_type", "code", "value", "is_active"])
    ws.append([ref_type, "EXAMPLE_CODE", "Example value", True])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="project_references_template_{ref_type}.xlsx"'},
    )


@router.get("/{project_id}/references/export")
def export_project_references(
    project_id: int,
    ref_type: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id, current_user)
    _get_project_or_404(db, project_id)
    refs = (
        db.query(ProjectReference)
        .filter(ProjectReference.project_id == project_id, ProjectReference.ref_type == ref_type)
        .order_by(ProjectReference.code.asc())
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "references"
    ws.append(["ref_type", "code", "value", "is_active"])
    for item in refs:
        ws.append([item.ref_type, item.code, item.value, item.is_active])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="project_references_{ref_type}.xlsx"'},
    )


@router.post("/{project_id}/references/import")
def import_project_references(
    project_id: int,
    ref_type: str = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not has_permission(current_user, "can_edit_project_references"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Reference edit permission required")
    _get_project_or_404(db, project_id)
    raw = file.file.read()
    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    imported = 0
    updated = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_ref_type = str(row[0] or "").strip()
        code = str(row[1] or "").strip()
        value = str(row[2] or "").strip()
        is_active = bool(row[3]) if row[3] is not None else True
        if not code:
            continue
        effective_ref_type = row_ref_type or ref_type
        if effective_ref_type != ref_type:
            continue
        exists = (
            db.query(ProjectReference)
            .filter(
                ProjectReference.project_id == project_id,
                ProjectReference.ref_type == effective_ref_type,
                ProjectReference.code == code,
            )
            .first()
        )
        if exists is None:
            db.add(
                ProjectReference(
                    project_id=project_id,
                    ref_type=effective_ref_type,
                    code=code,
                    value=value or code,
                    is_active=is_active,
                )
            )
            imported += 1
        else:
            exists.value = value or exists.value
            exists.is_active = is_active
            db.add(exists)
            updated += 1
    db.commit()
    return {"imported": imported, "updated": updated}


@router.put("/references/{reference_id}", response_model=ProjectReferenceRead)
def update_project_reference(
    reference_id: int,
    payload: ProjectReferenceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not has_permission(current_user, "can_edit_project_references"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Reference edit permission required")

    item = db.query(ProjectReference).filter(ProjectReference.id == reference_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reference not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, field, value)

    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/{project_id}/review-matrix", response_model=list[ReviewMatrixMemberRead])
def list_review_matrix(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_project_access(db, project_id, current_user)
    _get_project_or_404(db, project_id)
    items = (
        db.query(ReviewMatrixMember)
        .filter(ReviewMatrixMember.project_id == project_id)
        .order_by(
            ReviewMatrixMember.discipline_code.asc(),
            ReviewMatrixMember.doc_type.asc(),
            ReviewMatrixMember.level.asc(),
            ReviewMatrixMember.id.asc(),
        )
        .all()
    )
    return [_review_matrix_read(item, db) for item in items]


@router.post("/{project_id}/review-matrix", response_model=ReviewMatrixMemberRead, status_code=status.HTTP_201_CREATED)
def create_review_matrix_item(
    project_id: int,
    payload: ReviewMatrixMemberCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not has_permission(current_user, "can_manage_review_matrix"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Review matrix permission required")
    if not _can_manage_project_matrix(db, project_id, current_user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to manage review matrix")
    _get_project_or_404(db, project_id)

    user = db.query(User).filter(User.id == payload.user_id).first()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.company_type != CompanyType.owner:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only owner users can be assigned to review matrix",
        )
    member = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id, ProjectMember.user_id == payload.user_id)
        .first()
    )
    if member is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User must be project member",
        )
    if member.member_role not in {ProjectMemberRole.owner_member, ProjectMemberRole.observer}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only owner-side project members can be assigned to matrix",
        )

    _validate_mark_discipline_mapping(
        db,
        project_id=project_id,
        discipline_code=payload.discipline_code,
        mark_code=payload.doc_type,
    )

    exists = (
        db.query(ReviewMatrixMember)
        .filter(
            ReviewMatrixMember.project_id == project_id,
            ReviewMatrixMember.discipline_code == payload.discipline_code,
            ReviewMatrixMember.doc_type == payload.doc_type,
            ReviewMatrixMember.user_id == payload.user_id,
            ReviewMatrixMember.level == payload.level,
        )
        .first()
    )
    if exists is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Matrix row already exists")

    item = ReviewMatrixMember(project_id=project_id, **payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return _review_matrix_read(item, db)


@router.put("/review-matrix/{item_id}", response_model=ReviewMatrixMemberRead)
def update_review_matrix_item(
    item_id: int,
    payload: ReviewMatrixMemberUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    item = db.query(ReviewMatrixMember).filter(ReviewMatrixMember.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Matrix row not found")
    if not has_permission(current_user, "can_manage_review_matrix"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Review matrix permission required")
    if not _can_manage_project_matrix(db, item.project_id, current_user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to manage review matrix")

    next_discipline = payload.discipline_code if payload.discipline_code is not None else item.discipline_code
    next_mark = payload.doc_type if payload.doc_type is not None else item.doc_type
    _validate_mark_discipline_mapping(
        db,
        project_id=item.project_id,
        discipline_code=next_discipline,
        mark_code=next_mark,
    )

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, field, value)

    db.add(item)
    db.commit()
    db.refresh(item)
    return _review_matrix_read(item, db)


@router.delete("/review-matrix/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_review_matrix_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    item = db.query(ReviewMatrixMember).filter(ReviewMatrixMember.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Matrix row not found")
    if not has_permission(current_user, "can_manage_review_matrix"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Review matrix permission required")
    if not _can_manage_project_matrix(db, item.project_id, current_user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No rights to manage review matrix")

    db.delete(item)
    db.commit()
