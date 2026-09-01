from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
import tempfile
import re
from uuid import uuid4
import zipfile

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, get_effective_permissions, require_permissions, users_by_company_types
from app.services import matrix_gap
from app.models import (
    Comment,
    CommentAttachment,
    CarryOverDecision,
    CommentStatus,
    ContractorCommentStatus,
    Document,
    DocumentAttachment,
    MDRRecord,
    Notification,
    CompanyType,
    Project,
    ProjectMember,
    ProjectMemberRole,
    ReviewMatrixMember,
    ReviewCode,
    ReviewEvent,
    ProjectReference,
    Revision,
    RevisionReviewerState,
    SystemSetting,
    User,
)
from app.services import review_events as review_events_service
from app.schemas import (
    CommentCreate,
    CarryDecisionRead,
    CarryDecisionUpdate,
    CommentAttachmentRead,
    CommentOwnerDecision,
    RemarkCardRead,
    RemarkHistoryEvent,
    CommentRead,
    CsrQueueItem,
    CsrSendPayload,
    CommentResponse,
    DocumentCreate,
    DocumentRead,
    DocumentAttachmentRead,
    FileUploadResponse,
    PublishCommentsResult,
    RevisionCreate,
    RevisionCardRead,
    RevisionCommentThreadRead,
    RevisionRead,
    RevisionReviewCodeUpdate,
    RevisionTdoBulkDecision,
    RevisionTdoDecision,
    RevisionOverviewRead,
    RevisionRegistryCommentRead,
    RevisionRegistryRead,
    DocumentRegistryRead,
    ReviewEventRead,
    ReviewerStateRead,
    ReviewReportRow,
    RevisionReviewerSummary,
    TdoQueueItem,
    TrmListItem,
    TrmRevisionItem,
)

router = APIRouter()
from app.config import get_settings  # noqa: E402

UPLOAD_ROOT = Path(get_settings().tdo_uploads_root)
# Легаси-корень: файлы, загруженные до выноса пути в env, лежат тут. Отдаём их
# тоже, чтобы при смене TDO_UPLOADS_ROOT старые PDF не превратились в 404.
LEGACY_UPLOAD_ROOTS = {Path("/tmp/tdo_uploads").resolve()}
OWNER_VISIBLE_REVISION_STATUSES = {
    "UNDER_REVIEW",
    "OWNER_COMMENTS_SENT",
    "CONTRACTOR_REPLY_I",
    "CONTRACTOR_REPLY_A",
    "SUBMITTED",
}
REVISION_STATUS_TRANSITIONS: dict[str, set[str]] = {
    "UPLOADED_WAITING_TDO": {"REVISION_CREATED", "SUBMITTED", "CANCELLED_BY_TDO", "UPLOADED_WAITING_TDO"},
    "UNDER_REVIEW": {"UPLOADED_WAITING_TDO", "UNDER_REVIEW", "SUBMITTED"},
    "CANCELLED_BY_TDO": {"UPLOADED_WAITING_TDO", "CANCELLED_BY_TDO"},
    "OWNER_COMMENTS_SENT": {"UNDER_REVIEW", "SUBMITTED", "CONTRACTOR_REPLY_I", "CONTRACTOR_REPLY_A", "OWNER_COMMENTS_SENT"},
    "CONTRACTOR_REPLY_I": {"SUBMITTED", "OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I"},
    "CONTRACTOR_REPLY_A": {"SUBMITTED", "UNDER_REVIEW", "OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I", "CONTRACTOR_REPLY_A"},
}


def _observer_project_codes(db: Session, current_user: User) -> set[str]:
    """Коды проектов, где пользователь — наблюдатель.

    Наблюдатель смотрит весь реестр проекта (read-only), поэтому фильтр
    заказчика «только ревизии, переданные на рассмотрение» к нему не
    применяется — иначе до первой передачи ревизии он видел бы пустой список.
    """
    rows = (
        db.query(Project.code)
        .join(ProjectMember, ProjectMember.project_id == Project.id)
        .filter(
            ProjectMember.user_id == current_user.id,
            ProjectMember.member_role == ProjectMemberRole.observer,
        )
        .all()
    )
    return {row[0] for row in rows}


def _owner_can_access_revision(db: Session, current_user: User, revision: Revision) -> bool:
    if current_user.role.value == "admin":
        return True
    if current_user.company_type != CompanyType.owner:
        return True
    if revision.status in OWNER_VISIBLE_REVISION_STATUSES:
        return True
    # Наблюдатель (member_role=observer) видит весь реестр проекта read-only —
    # ревизии в ЛЮБЫХ статусах (ещё не переданные на рассмотрение, отклонённые
    # ТДО, завершённые). Действий у него нет: замечания/решения гейтятся
    # правами и матрицей рассмотрения, где наблюдателя нет.
    observer_codes = _observer_project_codes(db, current_user)
    if observer_codes:
        doc = db.query(Document).filter(Document.id == revision.document_id).first()
        mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first() if doc else None
        if mdr and mdr.project_code in observer_codes:
            return True
    return False


def _is_completed_document(db: Session, document_id: int) -> bool:
    latest = (
        db.query(Revision)
        .filter(Revision.document_id == document_id)
        .order_by(Revision.id.desc())
        .first()
    )
    if latest is None:
        return False
    return (latest.issue_purpose or "").upper() == "AFD" and latest.review_code == ReviewCode.AP


def _ensure_revision_not_locked_by_ap(revision: Revision) -> None:
    """После проставления AP по ревизии её цикл считается закрытым: ни LR,
    ни R, ни подрядчик не могут править замечания на этой ревизии
    (добавлять в CRS, отклонять, удалять, возвращать в работу). Это
    защита от случайного «оживления» уже согласованной ревизии."""
    if revision.review_code == ReviewCode.AP:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Revision is closed by AP — no further changes allowed",
        )


def _set_revision_status(revision: Revision, next_status: str) -> bool:
    current = revision.status
    if current == next_status:
        return False
    allowed_from = REVISION_STATUS_TRANSITIONS.get(next_status)
    if allowed_from and current not in allowed_from:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Revision status transition is not allowed: {current} -> {next_status}",
        )
    revision.status = next_status
    return True


def _matrix_fallback_receivers(db: Session) -> set[int]:
    """Кому слать, когда по разделу никто не назначен.

    Раньше письмо уходило ВСЕМ сотрудникам заказчика — спам людям, которые к
    документу отношения не имеют. Теперь это администраторы системы и адресаты
    из настройки matrix_gap_notify_emails: пробел в матрице чинят они.
    Новые документы без LR вообще не создаются (services/matrix_gap), но у
    старых записей матрица могла быть удалена задним числом.
    """
    return matrix_gap.gap_recipient_ids(db)


def _matrix_source(db: Session, mdr: MDRRecord | None) -> MDRRecord | None:
    """Запись, по которой определяется состав ревьюверов.

    Для вложенного документа это ВСЕГДА головной: вложенные части живут в цикле
    рассмотрения родителя, даже если у них своя дисциплина/тип/титул в шифре
    (напр. под концепцией АБЗ идут документы по разным зданиям). Иначе смена
    дисциплины у части уводила бы её к другим LR/R ([[Вложенные документы]])."""
    if mdr is None or mdr.parent_id is None:
        return mdr
    parent = db.query(MDRRecord).filter(MDRRecord.id == mdr.parent_id).first()
    return parent or mdr


def _matrix_discipline(db: Session, mdr: MDRRecord | None) -> str | None:
    """Код раздела для матрицы назначений.

    Категория SE (инженерные изыскания) в матрице представлена ОДНИМ условным
    разделом "SE": назначения на него распространяются на все SE-документы,
    независимо от конкретного вида отчёта в шифре. Для остальных категорий —
    фактический раздел документа (раздел ПД). У вложенного берётся раздел
    головного документа."""
    source = _matrix_source(db, mdr)
    if source is None:
        return None
    if (source.category or "").upper() == "SE":
        return "SE"
    return source.discipline_code


def _matrix_doc_type(db: Session, mdr: MDRRecord | None) -> str | None:
    """Тип документа для матрицы — у вложенного тип головного."""
    source = _matrix_source(db, mdr)
    return source.doc_type if source is not None else None


def _can_manage_owner_remark(
    db: Session,
    *,
    current_user: User,
    project_id: int,
    discipline_code: str | None,
    comment_author_id: int,
) -> bool:
    if current_user.role.value == "admin":
        return True
    if current_user.company_type != CompanyType.owner:
        return True
    if comment_author_id == current_user.id:
        return True
    if discipline_code:
        matrix_row = (
            db.query(ReviewMatrixMember.id)
            .filter(
                ReviewMatrixMember.project_id == project_id,
                ReviewMatrixMember.user_id == current_user.id,
                ReviewMatrixMember.discipline_code == discipline_code,
                ReviewMatrixMember.level == 1,
                ReviewMatrixMember.state == "LR",
            )
            .first()
        )
        if matrix_row is not None:
            return True
    # Owner reviewer (R) may manage only own remarks.
    # Managing others is allowed only for LR/admin.
    return False


def _is_lr_for_document(
    db: Session,
    *,
    current_user: User,
    project_id: int,
    discipline_code: str | None,
    doc_type: str | None,
) -> bool:
    if current_user.role.value == "admin":
        return True
    if current_user.company_type != CompanyType.owner:
        return False
    if not discipline_code:
        return False
    query = db.query(ReviewMatrixMember.id).filter(
        ReviewMatrixMember.project_id == project_id,
        ReviewMatrixMember.user_id == current_user.id,
        ReviewMatrixMember.discipline_code == discipline_code,
        ReviewMatrixMember.level == 1,
        ReviewMatrixMember.state == "LR",
    )
    if doc_type:
        exact = query.filter(ReviewMatrixMember.doc_type == doc_type).first()
        if exact is not None:
            return True
    return query.first() is not None


def _owner_matrix_role_for_document(
    db: Session,
    *,
    current_user: User,
    project_id: int,
    discipline_code: str | None,
    doc_type: str | None,
) -> str | None:
    if current_user.role.value == "admin":
        return "LR"
    if current_user.company_type != CompanyType.owner or not discipline_code:
        return None
    base_filter = [
        ReviewMatrixMember.project_id == project_id,
        ReviewMatrixMember.user_id == current_user.id,
        ReviewMatrixMember.discipline_code == discipline_code,
    ]
    # Сначала ищем строгое совпадение по типу документа.
    rows = []
    if doc_type:
        rows = db.query(ReviewMatrixMember).filter(*base_filter, ReviewMatrixMember.doc_type == doc_type).all()
    # Fallback: если по доку нет, пробуем без типа (назначение «по
    # дисциплине целиком» — типичный случай в матрице, когда строка
    # покрывает все типы документов в этой дисциплине).
    if not rows:
        rows = db.query(ReviewMatrixMember).filter(*base_filter).all()
    if not rows:
        return None
    lr_row = next((item for item in rows if item.level == 1 and item.state == "LR"), None)
    if lr_row is not None:
        return "LR"
    row = sorted(rows, key=lambda item: (item.level, 0 if item.state == "LR" else 1))[0]
    return "R"


def _ensure_lr_can_publish_for_revision(db: Session, *, current_user: User, revision: Revision) -> tuple[Document, MDRRecord, Project]:
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if not _is_lr_for_document(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        doc_type=_matrix_doc_type(db, mdr),
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only LR can add remarks to CRS and send CRS to contractor",
        )
    return document, mdr, project


def _is_contractor_developer_for_project(
    db: Session,
    *,
    current_user: User,
    project_id: int,
) -> bool:
    if current_user.role.value == "admin":
        return True
    if current_user.company_type != CompanyType.contractor:
        return False
    member = (
        db.query(ProjectMember)
        .filter(
            ProjectMember.project_id == project_id,
            ProjectMember.user_id == current_user.id,
        )
        .first()
    )
    # Доп. файлы (редактируемые исходники) грузит подрядчик: разработчик
    # (contractor_member) и рук. ТДО подрядчика (contractor_tdo_lead).
    return member is not None and member.member_role in (
        ProjectMemberRole.contractor_member,
        ProjectMemberRole.contractor_tdo_lead,
    )


def _comment_read(
    comment: Comment,
    author: User | None,
    *,
    contractor_response_text: str | None = None,
    contractor_response_at: datetime | None = None,
    attachment_count: int = 0,
) -> CommentRead:
    return CommentRead(
        id=comment.id,
        revision_id=comment.revision_id,
        parent_id=comment.parent_id,
        author_id=comment.author_id,
        author_name=author.full_name if author else None,
        author_email=author.email if author else None,
        text=comment.text,
        status=comment.status,
        review_code=comment.review_code,
        is_published_to_contractor=comment.is_published_to_contractor,
        backlog_status=comment.backlog_status,
        contractor_status=comment.contractor_status,
        contractor_response_text=contractor_response_text,
        contractor_response_at=contractor_response_at,
        in_crs=comment.in_crs,
        crs_sent_at=comment.crs_sent_at,
        crs_number=comment.crs_number,
        remark_number=comment.remark_number,
        carry_finalized=comment.carry_finalized,
        page=comment.page,
        area_x=comment.area_x,
        area_y=comment.area_y,
        area_w=comment.area_w,
        area_h=comment.area_h,
        created_at=comment.created_at,
        resolved_at=comment.resolved_at,
        attachment_count=attachment_count,
    )


def _comment_attachment_counts(db: Session, comments: list[Comment]) -> dict[int, int]:
    """Число вложений по каждому замечанию (bulk, чтобы не делать N запросов)."""
    ids = [c.id for c in comments if c.id is not None]
    if not ids:
        return {}
    rows = (
        db.query(CommentAttachment.comment_id, func.count(CommentAttachment.id))
        .filter(CommentAttachment.comment_id.in_(ids))
        .group_by(CommentAttachment.comment_id)
        .all()
    )
    return {comment_id: int(count) for comment_id, count in rows}


def _attachment_read(item: DocumentAttachment, uploader: User | None) -> DocumentAttachmentRead:
    return DocumentAttachmentRead(
        id=item.id,
        document_id=item.document_id,
        revision_id=item.revision_id,
        uploaded_by_id=item.uploaded_by_id,
        uploaded_by_name=uploader.full_name if uploader else None,
        uploaded_by_email=uploader.email if uploader else None,
        file_name=item.file_name,
        created_at=item.created_at,
    )


def _carry_decision_read(item: CarryOverDecision, decider: User | None) -> CarryDecisionRead:
    return CarryDecisionRead(
        id=item.id,
        target_revision_id=item.target_revision_id,
        source_comment_id=item.source_comment_id,
        status=item.status,
        decided_by_id=item.decided_by_id,
        decided_by_name=decider.full_name if decider else None,
        decided_by_email=decider.email if decider else None,
        decided_at=item.decided_at,
    )


def _latest_contractor_responses(
    db: Session,
    parent_comments: list[Comment],
) -> dict[int, tuple[str | None, datetime | None]]:
    parent_ids = [item.id for item in parent_comments if item.id is not None]
    if not parent_ids:
        return {}
    rows = (
        db.query(Comment, User)
        .join(User, User.id == Comment.author_id)
        .filter(
            Comment.parent_id.in_(parent_ids),
            User.company_type == CompanyType.contractor,
        )
        .order_by(Comment.created_at.desc(), Comment.id.desc())
        .all()
    )
    result: dict[int, tuple[str | None, datetime | None]] = {}
    for response, _author in rows:
        if response.parent_id is None:
            continue
        if response.parent_id in result:
            continue
        result[response.parent_id] = (response.text, response.created_at)
    return result


def _recompute_revision_contractor_status(db: Session, revision: Revision) -> None:
    published_comments = (
        db.query(Comment)
        .filter(
            Comment.revision_id == revision.id,
            Comment.parent_id.is_(None),
            Comment.is_published_to_contractor.is_(True),
            Comment.status != CommentStatus.REJECTED,
        )
        .all()
    )
    if not published_comments:
        # Do not advance contractor workflow before CRS is actually sent.
        # When revision is still on owner side (e.g. UNDER_REVIEW), keep status untouched.
        if revision.status in {"OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I", "CONTRACTOR_REPLY_A"}:
            _set_revision_status(revision, "CONTRACTOR_REPLY_A")
        return
    if any(item.contractor_status is None for item in published_comments):
        _set_revision_status(revision, "OWNER_COMMENTS_SENT")
        return
    if any(
        item.contractor_status == ContractorCommentStatus.I and item.status not in {CommentStatus.REJECTED, CommentStatus.RESOLVED}
        for item in published_comments
    ):
        _set_revision_status(revision, "CONTRACTOR_REPLY_I")
        return
    _set_revision_status(revision, "CONTRACTOR_REPLY_A")



def _revision_remarks_digest(db: Session, revision: Revision) -> tuple[int, str]:
    """(сколько замечаний передано подрядчику, итоговый код) по ревизии.

    Итоговый код — «худший» из опубликованных: RJ > CO > AN > AP."""
    rows = (
        db.query(Comment.review_code)
        .filter(
            Comment.revision_id == revision.id,
            Comment.parent_id.is_(None),
            Comment.is_published_to_contractor.is_(True),
            Comment.status != CommentStatus.REJECTED,
        )
        .all()
    )
    codes = {
        (value.value if hasattr(value, "value") else str(value))
        for (value,) in rows
        if value is not None
    }
    worst = next((code for code in ("RJ", "CO", "AN", "AP") if code in codes), "—")
    return len(rows), worst


def _upsert_crs_notification(
    db: Session,
    *,
    user_id: int,
    revision: Revision,
    document: Document,
    mdr: MDRRecord,
    crs_number: str | None = None,
) -> None:
    """Одно уведомление подрядчику на РЕВИЗИЮ вместо письма на каждое замечание.

    Подрядчик жаловался на спам: при поштучной публикации летело уведомление и
    письмо на каждое замечание. Теперь по ревизии держим одно непрочитанное
    уведомление-сводку и обновляем его текст (кол-во замечаний + итоговый код).
    Обновление существующего письма не порождает — рассылка идёт только на
    новые Notification (см. services/notification_email)."""
    db.flush()  # чтобы digest увидел только что опубликованные замечания
    count, worst = _revision_remarks_digest(db, revision)
    crs_part = f"CRS {crs_number}: " if crs_number else "Замечания заказчика: "
    message = (
        f"{crs_part}{document.document_num}, ревизия {revision.revision_code} — "
        f"замечаний: {count}, итоговый код: {worst}."
    )
    existing = (
        db.query(Notification)
        .filter(
            Notification.user_id == user_id,
            Notification.revision_id == revision.id,
            Notification.event_type.in_(["OWNER_COMMENTS_PUBLISHED", "OWNER_COMMENT_PUBLISHED"]),
            Notification.is_read.is_(False),
        )
        .order_by(Notification.id.desc())
        .first()
    )
    if existing is not None:
        existing.message = message
        existing.event_type = "OWNER_COMMENTS_PUBLISHED"
        db.add(existing)
        return
    db.add(
        Notification(
            user_id=user_id,
            event_type="OWNER_COMMENTS_PUBLISHED",
            message=message,
            project_code=mdr.project_code,
            document_num=document.document_num,
            revision_id=revision.id,
        )
    )


def _mark_notifications_read(
    db: Session,
    *,
    user_id: int,
    revision_id: int | None,
    event_types: list[str] | None = None,
) -> None:
    if revision_id is None:
        return
    query = db.query(Notification).filter(
        Notification.user_id == user_id,
        Notification.revision_id == revision_id,
        Notification.is_read.is_(False),
    )
    if event_types:
        query = query.filter(Notification.event_type.in_(event_types))
    query.update({Notification.is_read: True}, synchronize_session=False)


def _archive_document_notifications(
    db: Session,
    *,
    project_code: str | None,
    document_num: str | None,
    revision_id: int | None = None,
    event_types: list[str] | None = None,
) -> None:
    if not project_code or not document_num:
        return
    query = db.query(Notification).filter(
        Notification.project_code == project_code,
        Notification.document_num == document_num,
        Notification.is_read.is_(False),
    )
    if revision_id is not None:
        query = query.filter(or_(Notification.revision_id == revision_id, Notification.revision_id.is_(None)))
    if event_types:
        query = query.filter(Notification.event_type.in_(event_types))
    query.update({Notification.is_read: True}, synchronize_session=False)


def _is_project_tdo_lead(db: Session, project_id: int, user_id: int) -> bool:
    member = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_id == project_id, ProjectMember.user_id == user_id)
        .first()
    )
    if member is None:
        return False
    return member.member_role == ProjectMemberRole.contractor_tdo_lead or member.can_manage_contractor_users


@router.get("/revisions/tdo-queue", response_model=list[TdoQueueItem])
def list_tdo_queue(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    memberships = (
        db.query(ProjectMember)
        .filter(
            ProjectMember.user_id == current_user.id,
            (ProjectMember.member_role == ProjectMemberRole.contractor_tdo_lead)
            | (ProjectMember.can_manage_contractor_users.is_(True)),
        )
        .all()
    )
    project_ids = [item.project_id for item in memberships]
    if not project_ids:
        return []
    projects = db.query(Project).filter(Project.id.in_(project_ids)).all()
    project_code_to_id = {item.code: item.id for item in projects}
    allowed_codes = set(project_code_to_id.keys())
    if not allowed_codes:
        return []
    rows = (
        db.query(Revision, Document, MDRRecord, User)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .outerjoin(User, User.id == Revision.author_id)
        .filter(MDRRecord.project_code.in_(allowed_codes), Revision.status == "UPLOADED_WAITING_TDO")
        .order_by(Revision.id.asc())
        .all()
    )
    result: list[TdoQueueItem] = []
    for revision, document, mdr, author in rows:
        result.append(
            TdoQueueItem(
                revision_id=revision.id,
                project_code=mdr.project_code,
                document_num=document.document_num,
                document_title=document.title,
                revision_code=revision.revision_code,
                issue_purpose=revision.issue_purpose,
                status=revision.status,
                created_at=revision.created_at,
                review_deadline=revision.review_deadline,
                file_path=revision.file_path,
                author_id=revision.author_id,
                author_name=author.full_name if author else None,
                author_email=author.email if author else None,
            )
        )
    return result


@router.get("/revisions/overview", response_model=list[RevisionOverviewRead])
def list_revisions_overview(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = (
        db.query(Revision, Document, MDRRecord, User)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .outerjoin(User, User.id == Revision.author_id)
    )
    if current_user.role.value != "admin":
        allowed_codes = (
            db.query(Project.code)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .all()
        )
        allowed_project_codes = [row[0] for row in allowed_codes]
        if not allowed_project_codes:
            return []
        query = query.filter(MDRRecord.project_code.in_(allowed_project_codes))
    if current_user.role.value != "admin" and current_user.company_type == CompanyType.owner:
        query = query.filter(Revision.status.in_(OWNER_VISIBLE_REVISION_STATUSES))
    rows = query.order_by(Revision.id.desc()).all()
    return [
        RevisionOverviewRead(
            revision_id=revision.id,
            project_code=mdr.project_code,
            document_num=document.document_num,
            document_title=document.title,
            revision_code=revision.revision_code,
            issue_purpose=revision.issue_purpose,
            status=revision.status,
            trm_number=revision.trm_number,
            review_deadline=revision.review_deadline,
            file_path=revision.file_path,
            author_id=revision.author_id,
            author_name=author.full_name if author else None,
            author_email=author.email if author else None,
            created_at=revision.created_at,
        )
        for revision, document, mdr, author in rows
    ]


@router.get("/revisions/owner-review-queue", response_model=list[TdoQueueItem])
def list_owner_review_queue(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Подстраховка демона: при заходе в очередь досоздаём напоминания о
    # дедлайне (идемпотентно) — на случай перезапуска backend-процесса.
    try:
        from app.services.deadline_reminders import scan_and_notify

        scan_and_notify(db)
    except Exception:  # noqa: BLE001 — фоновая мелочь не должна ронять очередь
        pass
    # Очередь рассмотрения — задачи ревьюверов/подрядчика, НЕ наблюдателя.
    # Наблюдатель (member_role=observer) видит ТРМ через отдельную страницу «ТРМ»,
    # а «задач» у него быть не должно (иначе на дашборде висел чужой список).
    membership_project_ids = {
        item.project_id
        for item in db.query(ProjectMember)
        .filter(
            ProjectMember.user_id == current_user.id,
            ProjectMember.member_role != ProjectMemberRole.observer,
        )
        .all()
    }
    matrix_project_ids = {
        item.project_id
        for item in db.query(ReviewMatrixMember)
        .filter(ReviewMatrixMember.user_id == current_user.id, ReviewMatrixMember.level == 1)
        .all()
    }
    project_ids = list(membership_project_ids | matrix_project_ids)
    if not project_ids:
        return []
    projects = db.query(Project).filter(Project.id.in_(project_ids)).all()
    allowed_codes = {item.code for item in projects if item.code}
    project_id_by_code = {item.code: item.id for item in projects if item.code}
    queue_query = (
        db.query(Revision, Document, MDRRecord, User)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .outerjoin(User, User.id == Revision.author_id)
        .filter(or_(Revision.status == "UNDER_REVIEW", Revision.trm_number.isnot(None)))
    )
    notified_revision_ids = {
        item.revision_id
        for item in db.query(Notification)
        .filter(
            Notification.user_id == current_user.id,
            Notification.event_type == "TDO_SENT_TO_OWNER",
            Notification.revision_id.isnot(None),
        )
        .all()
        if item.revision_id
    }
    if allowed_codes:
        if notified_revision_ids:
            queue_query = queue_query.filter(
                or_(
                    MDRRecord.project_code.in_(allowed_codes),
                    Revision.id.in_(notified_revision_ids),
                )
            )
        else:
            queue_query = queue_query.filter(MDRRecord.project_code.in_(allowed_codes))
    elif notified_revision_ids:
        queue_query = queue_query.filter(Revision.id.in_(notified_revision_ids))
    else:
        return []
    rows = queue_query.order_by(Revision.created_at.asc()).all()
    # Очередь LR/R должна показывать только «живые» задачи. Если по
    # ревизии уже выставлен review_code (AN/CO/RJ/AP) — цикл закрыт,
    # задача висеть не должна. Аналогично исключаем не-последние ревизии
    # документа: ответ на A после выпуска B уже не нужен, мяч ушёл вперёд.
    latest_revision_by_doc: dict[int, int] = {}
    for revision, document, _mdr, _author in rows:
        existing = latest_revision_by_doc.get(document.id)
        if existing is None or revision.id > existing:
            latest_revision_by_doc[document.id] = revision.id

    result: list[TdoQueueItem] = []
    for revision, document, mdr, author in rows:
        # Только последняя ревизия документа
        if latest_revision_by_doc.get(document.id) != revision.id:
            continue
        # Закрытые review_code-ом ревизии больше не задачи
        if revision.review_code is not None:
            continue
        # Завершённые документы (AFD+AP уже стоит на последней) — исключить
        if _is_completed_document(db, document.id):
            continue
        if revision.status == "UNDER_REVIEW" and not revision.trm_number:
            sender_code = (author.company_code if author else None) or (mdr.originator_code or "").strip().upper() or "CTR"
            receiver_code = _project_reference_value(
                db,
                project_id=project_id_by_code.get(mdr.project_code) or 0,
                code="TRM_RECEIVER_COMPANY_CODE",
            ) or "IVA"
            revision.trm_number = _next_trm_number(
                db,
                project_code=mdr.project_code,
                sender_company_code=sender_code,
                receiver_company_code=receiver_code,
            )
            db.add(revision)
        project_id = project_id_by_code.get(mdr.project_code)
        can_publish_to_contractor = bool(
            project_id
            and _is_lr_for_document(
                db,
                current_user=current_user,
                project_id=project_id,
                discipline_code=_matrix_discipline(db, mdr),
                doc_type=_matrix_doc_type(db, mdr),
            )
        )
        result.append(
            TdoQueueItem(
                revision_id=revision.id,
                project_code=mdr.project_code,
                document_num=document.document_num,
                document_title=document.title,
                revision_code=revision.revision_code,
                issue_purpose=revision.issue_purpose,
                status=revision.status,
                created_at=revision.created_at,
                review_deadline=revision.review_deadline,
                trm_number=revision.trm_number,
                file_path=revision.file_path,
                can_publish_to_contractor=can_publish_to_contractor,
                author_id=revision.author_id,
                author_name=author.full_name if author else None,
                author_email=author.email if author else None,
            )
        )
    db.commit()
    return result


def _trm_allowed_project_codes(db: Session, current_user: User) -> set[str] | None:
    """Коды проектов, где пользователь — участник (любая роль, вкл. наблюдателя).
    None = админ (все проекты)."""
    if current_user.role.value == "admin":
        return None
    rows = (
        db.query(Project.code)
        .join(ProjectMember, ProjectMember.project_id == Project.id)
        .filter(ProjectMember.user_id == current_user.id)
        .all()
    )
    return {r[0] for r in rows if r[0]}


@router.get("/trm", response_model=list[TrmListItem])
def list_trm(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Реестр ТРМ по проектам пользователя (read-only, доступен всем участникам,
    включая наблюдателя). Группировка по trm_number."""
    allowed = _trm_allowed_project_codes(db, current_user)
    q = (
        db.query(Revision, Document, MDRRecord)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .filter(Revision.trm_number.isnot(None))
    )
    if allowed is not None:
        q = q.filter(MDRRecord.project_code.in_(list(allowed) or ["__none__"]))
    rows = q.order_by(Revision.id.asc()).all()
    groups: dict[str, dict] = {}
    for rev, doc, mdr in rows:
        g = groups.setdefault(
            rev.trm_number,
            {"project_code": mdr.project_code, "docs": set(), "revisions": [], "last_status": None, "deadline": None, "last_id": -1},
        )
        g["docs"].add(doc.id)
        g["revisions"].append(
            TrmRevisionItem(
                revision_id=rev.id,
                document_num=doc.document_num,
                document_title=doc.title,
                revision_code=rev.revision_code,
                issue_purpose=rev.issue_purpose,
                status=rev.status,
                review_code=rev.review_code,
                created_at=rev.created_at,
            )
        )
        if rev.id > g["last_id"]:
            g["last_id"] = rev.id
            g["last_status"] = rev.status
            g["deadline"] = rev.review_deadline
    result = [
        TrmListItem(
            trm_number=trm,
            project_code=g["project_code"],
            document_count=len(g["docs"]),
            last_status=g["last_status"],
            review_deadline=g["deadline"],
            revisions=g["revisions"],
        )
        for trm, g in groups.items()
    ]
    # Новые ТРМ (больший серийник) сверху.
    result.sort(key=lambda x: x.trm_number, reverse=True)
    return result


@router.get("/trm/{trm_number}/transmittal.xlsx")
def export_trm_transmittal_xlsx(
    trm_number: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Сопроводительный лист ТРМ (Excel): шапка + список документов/ревизий."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from fastapi.responses import StreamingResponse

    allowed = _trm_allowed_project_codes(db, current_user)
    q = (
        db.query(Revision, Document, MDRRecord)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .filter(Revision.trm_number == trm_number)
    )
    if allowed is not None:
        q = q.filter(MDRRecord.project_code.in_(list(allowed) or ["__none__"]))
    rows = q.order_by(Document.document_num, Revision.id).all()
    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ТРМ не найден или нет доступа")

    project_code = rows[0][2].project_code
    wb = Workbook()
    ws = wb.active
    ws.title = "Transmittal"
    header_pairs = [
        ("Сопроводительный лист (ТРМ)", ""),
        ("Номер ТРМ", trm_number),
        ("Проект", project_code),
        ("Дата формирования", datetime.utcnow().strftime("%Y-%m-%d")),
        ("Документов в ТРМ", str(len({d.id for _r, d, _m in rows}))),
    ]
    for label, value in header_pairs:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])

    header_row = ws.max_row + 1
    headers = ["№", "Шифр документа", "Название", "Ревизия", "Цель выпуска", "Статус", "Код замечаний"]
    ws.append(headers)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    status_label = {
        "REVISION_CREATED": "Ревизия создана",
        "UPLOADED_WAITING_TDO": "Загружен PDF (ожидает ТДО)",
        "UNDER_REVIEW": "На рассмотрении заказчиком",
        "OWNER_COMMENTS_SENT": "Замечания отправлены подрядчику",
        "CONTRACTOR_REPLY_I": "Ответ подрядчика (обсуждение)",
        "CONTRACTOR_REPLY_A": "Ответ подрядчика (принято)",
        "CANCELLED_BY_TDO": "Отклонено ТДО",
        "SUBMITTED": "Выпущено",
    }
    for idx, (rev, doc, _mdr) in enumerate(rows, start=1):
        ws.append([
            idx,
            doc.document_num,
            doc.title or "",
            rev.revision_code,
            rev.issue_purpose,
            status_label.get(rev.status, rev.status),
            (rev.review_code.value if rev.review_code else ""),
        ])
    widths = [5, 30, 46, 10, 14, 30, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w

    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", trm_number)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe}_transmittal.xlsx"},
    )


def _project_reference_value(db: Session, *, project_id: int, code: str) -> str | None:
    ref = (
        db.query(ProjectReference)
        .filter(
            ProjectReference.project_id == project_id,
            ProjectReference.ref_type == "other",
            ProjectReference.code == code,
            ProjectReference.is_active.is_(True),
        )
        .first()
    )
    if ref is None:
        return None
    raw = (ref.value or "").strip().upper()
    return raw or None


def _next_trm_number(db: Session, *, project_code: str, sender_company_code: str, receiver_company_code: str = "IVA") -> str:
    prefix = f"{project_code.upper()}-{sender_company_code.upper()}-{receiver_company_code.upper()}-TRM-"
    matches = (
        db.query(Revision.trm_number)
        .filter(Revision.trm_number.isnot(None), Revision.trm_number.like(f"{prefix}%"))
        .all()
    )
    max_seq = 0
    for (value,) in matches:
        if not value:
            continue
        match = re.match(rf"^{re.escape(prefix)}(\d{{5}})$", value)
        if match:
            max_seq = max(max_seq, int(match.group(1)))
    next_seq = max_seq + 1
    return f"{prefix}{next_seq:05d}"


def _next_remark_number(db: Session, *, project_code: str) -> str:
    """Сквозной по проекту номер замечания: IMP-RMK-000123.

    Присваивается один раз при создании и дальше неизменен — по нему ищется
    вся история замечания. Нумерация независима от документа и ревизии,
    поэтому carry-over на следующую ревизию номер не меняет."""
    prefix = f"{project_code.upper()}-RMK-"
    rows = (
        db.query(Comment.remark_number)
        .filter(Comment.remark_number.isnot(None), Comment.remark_number.like(f"{prefix}%"))
        .all()
    )
    max_seq = 0
    for (value,) in rows:
        match = re.match(rf"^{re.escape(prefix)}(\d{{6}})$", value or "")
        if match:
            max_seq = max(max_seq, int(match.group(1)))
    return f"{prefix}{max_seq + 1:06d}"


def _next_crs_number(
    db: Session,
    *,
    document_id: int,
    project_code: str,
    sender_company_code: str,
    receiver_company_code: str,
) -> str:
    prefix = f"{project_code.upper()}-{sender_company_code.upper()}-{receiver_company_code.upper()}-CRS-"
    matches = (
        db.query(Comment.crs_number)
        .join(Revision, Revision.id == Comment.revision_id)
        .filter(
            Revision.document_id == document_id,
            Comment.crs_number.isnot(None),
            Comment.crs_number.like(f"{prefix}%"),
        )
        .all()
    )
    max_seq = 0
    for (value,) in matches:
        if not value:
            continue
        match = re.match(rf"^{re.escape(prefix)}(\d{{5}})$", value)
        if match:
            max_seq = max(max_seq, int(match.group(1)))
    return f"{prefix}{max_seq + 1:05d}"


def _sla_days_for_revision(
    db: Session,
    *,
    project_id: int,
    category: str,
    issue_purpose: str,
    is_initial_revision: bool,
) -> int:
    revision_kind = "INITIAL" if is_initial_revision else "NEXT"
    keys = [
        f"{category.upper()}:{issue_purpose.upper()}:{revision_kind}",
        f"{category.upper()}:*:{revision_kind}",
        f"*:{issue_purpose.upper()}:{revision_kind}",
        f"*:*:{revision_kind}",
    ]
    refs = (
        db.query(ProjectReference)
        .filter(
            ProjectReference.project_id == project_id,
            ProjectReference.ref_type == "review_sla_days",
            ProjectReference.is_active.is_(True),
        )
        .all()
    )
    by_code = {item.code.upper(): item.value for item in refs}
    for key in keys:
        raw = by_code.get(key)
        if raw is None:
            continue
        try:
            days = int(str(raw).strip())
            if days > 0:
                return days
        except ValueError:
            continue
    fallback_key = "review_sla_default_initial_days" if is_initial_revision else "review_sla_default_next_days"
    setting = db.query(SystemSetting).filter(SystemSetting.key == fallback_key).first()
    if setting is not None:
        try:
            parsed = int(str(setting.value).strip())
            if parsed > 0:
                return parsed
        except ValueError:
            pass
    return 14 if is_initial_revision else 7


def _setting_days(db: Session, key: str, default: float) -> float:
    item = db.query(SystemSetting).filter(SystemSetting.key == key).first()
    if item is None:
        return default
    try:
        value = float(str(item.value).strip())
        return value if value > 0 else default
    except ValueError:
        return default


@router.get("/documents", response_model=list[DocumentRead])
def list_documents(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    docs_query = db.query(Document).join(MDRRecord, MDRRecord.id == Document.mdr_id)
    if current_user.role.value != "admin":
        allowed_codes = (
            db.query(Project.code)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .all()
        )
        allowed_project_codes = [row[0] for row in allowed_codes]
        if not allowed_project_codes:
            return []
        docs_query = docs_query.filter(MDRRecord.project_code.in_(allowed_project_codes))
    docs = docs_query.order_by(Document.id.desc()).all()
    observer_codes = _observer_project_codes(db, current_user) if current_user.role.value != "admin" else set()
    result: list[DocumentRead] = []
    for doc in docs:
        latest = (
            db.query(Revision)
            .filter(Revision.document_id == doc.id)
            .order_by(Revision.id.desc())
            .first()
        )
        if (
            current_user.role.value != "admin"
            and current_user.company_type == CompanyType.owner
            and doc.mdr.project_code not in observer_codes
        ):
            if latest is None or latest.status not in OWNER_VISIBLE_REVISION_STATUSES:
                continue
        result.append(
            DocumentRead.model_validate(doc, from_attributes=True).model_copy(
                update={
                    "latest_revision_code": latest.revision_code if latest else None,
                    "latest_revision_status": latest.status if latest else None,
                    "latest_review_code": latest.review_code if latest else None,
                    "latest_issue_purpose": latest.issue_purpose if latest else None,
                }
            )
        )
    return result


@router.get("/documents/registry", response_model=list[DocumentRegistryRead])
def list_documents_registry(
    project_code: str | None = Query(default=None),
    category: str | None = Query(default=None),
    discipline_code: str | None = Query(default=None),
    document_title: str | None = Query(default=None),
    release_status: str | None = Query(default=None),
    revision_status: str | None = Query(default=None),
    comments_scope: str | None = Query(default=None, pattern="^(ANY|OPEN|NONE)$"),
    overdue_only: bool = Query(default=False),
    for_reporting: bool = Query(default=False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Для отчётности заказчику показываем все ревизии проекта (включая
    # pre-owner статусы), чтобы корректно строить план/факт по всему циклу.
    bypass_owner_filter = for_reporting and bool(get_effective_permissions(current_user).get("can_view_reporting"))
    docs_query = db.query(Document).join(MDRRecord, MDRRecord.id == Document.mdr_id)
    if current_user.role.value != "admin":
        allowed_codes = (
            db.query(Project.code)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .all()
        )
        allowed_project_codes = [row[0] for row in allowed_codes]
        if not allowed_project_codes:
            return []
        docs_query = docs_query.filter(MDRRecord.project_code.in_(allowed_project_codes))
    docs = docs_query.order_by(Document.id.desc()).all()
    observer_codes = _observer_project_codes(db, current_user) if current_user.role.value != "admin" else set()

    normalized_scope = (comments_scope or "").upper()
    result: list[DocumentRegistryRead] = []
    overdue_notifications_to_create: list[Notification] = []
    for doc in docs:
        mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
        if mdr is None:
            continue
        skip_owner_filter = bypass_owner_filter or mdr.project_code in observer_codes
        revisions = (
            db.query(Revision)
            .filter(Revision.document_id == doc.id)
            .order_by(Revision.id.desc())
            .all()
        )
        if current_user.role.value != "admin" and current_user.company_type == CompanyType.owner and not skip_owner_filter:
            revisions = [item for item in revisions if item.status in OWNER_VISIBLE_REVISION_STATUSES]
        latest = revisions[0] if revisions else None
        if current_user.role.value != "admin" and current_user.company_type == CompanyType.owner and not skip_owner_filter and latest is None:
            continue

        if project_code and mdr.project_code != project_code:
            continue
        if category and mdr.category != category:
            continue
        if discipline_code and mdr.discipline_code != discipline_code:
            continue
        if document_title and document_title.strip().lower() not in (doc.title or "").lower():
            continue
        if release_status and (latest.review_code.value if latest and latest.review_code else None) != release_status:
            continue
        if revision_status and (latest.status if latest else None) != revision_status:
            continue

        revision_rows: list[RevisionRegistryRead] = []
        total_comments = 0
        open_comments = 0
        first_upload_date = None
        for rev in revisions:
            if rev.file_path and (first_upload_date is None or rev.created_at < first_upload_date):
                first_upload_date = rev.created_at
            comments = (
                db.query(Comment)
                .filter(Comment.revision_id == rev.id, Comment.parent_id.is_(None))
                .order_by(Comment.id.asc())
                .all()
            )
            comment_rows = [
                RevisionRegistryCommentRead(
                    id=item.id,
                    parent_id=item.parent_id,
                    remark_number=item.remark_number,
                    text=item.text,
                    status=item.status,
                    review_code=item.review_code,
                    in_crs=item.in_crs,
                    contractor_status=item.contractor_status,
                    is_published_to_contractor=item.is_published_to_contractor,
                    author_id=item.author_id,
                    created_at=item.created_at,
                    carry_finalized=item.carry_finalized,
                )
                for item in comments
            ]
            comments_count = len(comment_rows)
            open_count = sum(1 for item in comment_rows if item.status in {"OPEN", "IN_PROGRESS"})
            total_comments += comments_count
            open_comments += open_count
            author_name = None
            if rev.author_id:
                author = db.query(User).filter(User.id == rev.author_id).first()
                author_name = author.full_name if author else None
            revision_rows.append(
                RevisionRegistryRead(
                    id=rev.id,
                    revision_code=rev.revision_code,
                    issue_purpose=rev.issue_purpose,
                    status=rev.status,
                    review_code=rev.review_code,
                    trm_number=rev.trm_number,
                    trm_flag=bool(rev.trm_number),
                    author_id=rev.author_id,
                    author_name=author_name,
                    created_at=rev.created_at,
                    comments_count=comments_count,
                    open_comments_count=open_count,
                    comments=comment_rows,
                )
            )

        if normalized_scope == "OPEN" and open_comments == 0:
            continue
        if normalized_scope == "NONE" and total_comments > 0:
            continue
        is_overdue = bool(
            mdr.planned_dev_start
            and mdr.planned_dev_start < date.today()
            and first_upload_date is None
        )
        if overdue_only and not is_overdue:
            continue

        latest_author_name = revision_rows[0].author_name if revision_rows else None
        result.append(
            DocumentRegistryRead(
                document_id=doc.id,
                project_code=mdr.project_code,
                category=mdr.category,
                discipline_code=mdr.discipline_code,
                document_num=doc.document_num,
                document_title=doc.title,
                latest_revision_code=latest.revision_code if latest else None,
                latest_revision_status=latest.status if latest else None,
                latest_issue_purpose=latest.issue_purpose if latest else None,
                latest_review_code=latest.review_code if latest else None,
                latest_author_name=latest_author_name,
                planned_dev_start=mdr.planned_dev_start,
                first_upload_date=first_upload_date,
                is_overdue=is_overdue,
                total_comments_count=total_comments,
                open_comments_count=open_comments,
                revisions=revision_rows,
            )
        )
        # Просрочки — задача рук. ТДО подрядчика, не админа-наблюдателя
        # (иначе админ засыпан уведомлениями по всем документам).
        if current_user.permissions.get("can_process_tdo_queue") and current_user.role.value != "admin":
            if is_overdue:
                exists_overdue_notification = (
                    db.query(Notification.id)
                    .filter(
                        Notification.user_id == current_user.id,
                        Notification.project_code == mdr.project_code,
                        Notification.document_num == doc.document_num,
                        Notification.event_type == "DOC_OVERDUE_PLAN_START",
                    )
                    .first()
                )
                if exists_overdue_notification is None:
                    overdue_notifications_to_create.append(
                        Notification(
                            user_id=current_user.id,
                            event_type="DOC_OVERDUE_PLAN_START",
                            message=f"Просрочка старта разработки: {doc.document_num}",
                            project_code=mdr.project_code,
                            document_num=doc.document_num,
                        )
                    )
            else:
                # Документ больше не просрочен (PDF загружен / старт был) —
                # гасим ранее выданное напоминание, чтобы не висело вечно (item 14).
                db.query(Notification).filter(
                    Notification.user_id == current_user.id,
                    Notification.project_code == mdr.project_code,
                    Notification.document_num == doc.document_num,
                    Notification.event_type == "DOC_OVERDUE_PLAN_START",
                    Notification.is_read.is_(False),
                ).update({Notification.is_read: True}, synchronize_session=False)
    if overdue_notifications_to_create:
        for item in overdue_notifications_to_create:
            db.add(item)
        db.commit()
    return result


@router.get("/documents/{document_id}", response_model=DocumentRead)
def get_document(document_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    latest = (
        db.query(Revision)
        .filter(Revision.document_id == doc.id)
        .order_by(Revision.id.desc())
        .first()
    )
    if current_user.role.value != "admin" and current_user.company_type == CompanyType.owner:
        if latest is None or latest.status not in OWNER_VISIBLE_REVISION_STATUSES:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    return DocumentRead.model_validate(doc, from_attributes=True).model_copy(
        update={
            "latest_revision_code": latest.revision_code if latest else None,
            "latest_revision_status": latest.status if latest else None,
            "latest_review_code": latest.review_code if latest else None,
            "latest_issue_purpose": latest.issue_purpose if latest else None,
        }
    )


@router.post("/documents", response_model=DocumentRead, status_code=status.HTTP_201_CREATED)
def create_document(
    payload: DocumentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_create_mdr")),
):
    mdr = db.query(MDRRecord).filter(MDRRecord.id == payload.mdr_id).first()
    if not mdr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")

    doc = Document(
        **payload.model_dump(),
        created_by_id=current_user.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return DocumentRead.model_validate(doc, from_attributes=True).model_copy(
        update={
            "latest_revision_code": None,
            "latest_revision_status": None,
            "latest_review_code": None,
            "latest_issue_purpose": None,
        }
    )


@router.post("/documents/upload", response_model=FileUploadResponse, status_code=status.HTTP_201_CREATED)
def upload_document_file(
    file: UploadFile = File(...),
    revision_id: int | None = Form(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_upload_files")),
):
    revision: Revision | None = None
    if revision_id is not None:
        revision = db.query(Revision).filter(Revision.id == revision_id).first()
        if revision is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
        if _is_completed_document(db, revision.document_id):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); uploads are locked")

    original_name = file.filename or "document.pdf"
    extension = Path(original_name).suffix.lower()
    if extension != ".pdf":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only PDF files are allowed")

    day_folder = datetime.utcnow().strftime("%Y%m%d")
    destination_dir = UPLOAD_ROOT / day_folder
    destination_dir.mkdir(parents=True, exist_ok=True)

    safe_name = f"{uuid4().hex}_{Path(original_name).name.replace(' ', '_')}"
    destination = destination_dir / safe_name

    file_bytes = file.file.read()
    destination.write_bytes(file_bytes)

    storage_path = str(destination)

    if revision is not None:
        revision.file_path = storage_path
        _set_revision_status(revision, "UPLOADED_WAITING_TDO")

        doc = db.query(Document).filter(Document.id == revision.document_id).first()
        if doc is not None:
            mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
            if mdr is not None:
                # Close cancellation tasks after developer re-upload.
                db.query(Notification).filter(
                    Notification.user_id == current_user.id,
                    Notification.revision_id == revision.id,
                    Notification.event_type == "TDO_CANCELLED_REVISION",
                    Notification.is_read.is_(False),
                ).update({Notification.is_read: True}, synchronize_session=False)
                db.query(Notification).filter(
                    Notification.user_id == current_user.id,
                    Notification.project_code == mdr.project_code,
                    Notification.document_num == doc.document_num,
                    Notification.event_type == "TDO_CANCELLED_REVISION",
                    Notification.is_read.is_(False),
                ).update({Notification.is_read: True}, synchronize_session=False)
                project = db.query(Project).filter(Project.code == mdr.project_code).first()
                if project is not None:
                    lead_members = (
                        db.query(ProjectMember)
                        .filter(
                            ProjectMember.project_id == project.id,
                            ProjectMember.member_role == ProjectMemberRole.contractor_tdo_lead,
                        )
                        .all()
                    )
                    for lead in lead_members:
                        # Дедупликация: если у рук. ТДО уже есть непрочитанное
                        # уведомление по этой же ревизии — не плодим новые.
                        # Подрядчик может перезаливать PDF несколько раз —
                        # ТДО всё равно нужна одна задача «проверить и в TRM».
                        already_exists = (
                            db.query(Notification.id)
                            .filter(
                                Notification.user_id == lead.user_id,
                                Notification.event_type == "REVISION_UPLOADED_FOR_TDO",
                                Notification.revision_id == revision.id,
                                Notification.is_read.is_(False),
                            )
                            .first()
                            is not None
                        )
                        if already_exists:
                            continue
                        db.add(
                            Notification(
                                user_id=lead.user_id,
                                event_type="REVISION_UPLOADED_FOR_TDO",
                                message=(
                                    f"Загружен PDF для документа {doc.document_num}, ревизия {revision.revision_code}. "
                                    "Требуется решение ТДО (проверить комплектность и отправить в TRM)."
                                ),
                                project_code=mdr.project_code,
                                document_num=doc.document_num,
                                revision_id=revision.id,
                            )
                        )
        db.add(revision)
        db.commit()

    return FileUploadResponse(
        file_name=original_name,
        file_path=storage_path,
        content_type=file.content_type or "application/pdf",
        file_size=len(file_bytes),
    )


@router.get("/documents/{document_id}/attachments", response_model=list[DocumentAttachmentRead])
def list_document_attachments(
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    document = db.query(Document).filter(Document.id == document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    latest_revision = (
        db.query(Revision)
        .filter(Revision.document_id == document.id)
        .order_by(Revision.id.desc())
        .first()
    )
    if latest_revision is not None and not _owner_can_access_revision(db, current_user, latest_revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    items = (
        db.query(DocumentAttachment)
        .filter(DocumentAttachment.document_id == document.id)
        .order_by(DocumentAttachment.created_at.desc(), DocumentAttachment.id.desc())
        .all()
    )
    user_ids = {item.uploaded_by_id for item in items}
    users = (
        db.query(User)
        .filter(User.id.in_(user_ids))
        .all()
        if user_ids
        else []
    )
    user_map = {user.id: user for user in users}
    return [_attachment_read(item, user_map.get(item.uploaded_by_id)) for item in items]


@router.get("/revisions/{revision_id}/attachments", response_model=list[DocumentAttachmentRead])
def list_revision_attachments(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    items = (
        db.query(DocumentAttachment)
        .filter(DocumentAttachment.revision_id == revision.id)
        .order_by(DocumentAttachment.created_at.desc(), DocumentAttachment.id.desc())
        .all()
    )
    user_ids = {item.uploaded_by_id for item in items}
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {user.id: user for user in users}
    return [_attachment_read(item, user_map.get(item.uploaded_by_id)) for item in items]


@router.post("/documents/{document_id}/attachments", response_model=DocumentAttachmentRead, status_code=status.HTTP_201_CREATED)
def upload_document_attachment(
    document_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_upload_files")),
):
    document = db.query(Document).filter(Document.id == document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    project = db.query(Project).filter(Project.code == mdr.project_code).first() if mdr else None
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if not _is_contractor_developer_for_project(db, current_user=current_user, project_id=project.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only contractor (developer or TDO lead) can upload extra files")
    original_name = file.filename or "attachment.bin"
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).name)
    destination_dir = UPLOAD_ROOT / "attachments" / str(document.id)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / f"{uuid4().hex}_{safe_name}"
    destination.write_bytes(file.file.read())
    item = DocumentAttachment(
        document_id=document.id,
        uploaded_by_id=current_user.id,
        file_name=safe_name,
        file_path=str(destination),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _attachment_read(item, current_user)


@router.post("/revisions/{revision_id}/attachments", response_model=DocumentAttachmentRead, status_code=status.HTTP_201_CREATED)
def upload_revision_attachment(
    revision_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_upload_files")),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    if _is_completed_document(db, document.id):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); uploads are locked")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    project = db.query(Project).filter(Project.code == mdr.project_code).first() if mdr else None
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if not _is_contractor_developer_for_project(db, current_user=current_user, project_id=project.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only contractor (developer or TDO lead) can upload extra files")
    original_name = file.filename or "attachment.bin"
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).name)
    destination_dir = UPLOAD_ROOT / "attachments" / str(document.id) / str(revision.id)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / f"{uuid4().hex}_{safe_name}"
    destination.write_bytes(file.file.read())
    item = DocumentAttachment(
        document_id=document.id,
        revision_id=revision.id,
        uploaded_by_id=current_user.id,
        file_name=safe_name,
        file_path=str(destination),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _attachment_read(item, current_user)


@router.get("/documents/{document_id}/attachments/archive")
def download_document_attachments_archive(
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    document = db.query(Document).filter(Document.id == document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    latest_revision = (
        db.query(Revision)
        .filter(Revision.document_id == document.id)
        .order_by(Revision.id.desc())
        .first()
    )
    if latest_revision is not None and not _owner_can_access_revision(db, current_user, latest_revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    attachments = db.query(DocumentAttachment).filter(DocumentAttachment.document_id == document.id).all()
    if not attachments:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="К этой ревизии не приложены файлы")
    temp_file = tempfile.NamedTemporaryFile(prefix=f"doc_{document.id}_", suffix=".zip", delete=False)
    temp_path = Path(temp_file.name)
    temp_file.close()
    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for item in attachments:
            file_path = Path(item.file_path)
            if file_path.exists():
                archive.write(file_path, arcname=item.file_name)
    return FileResponse(
        path=str(temp_path),
        media_type="application/zip",
        filename=f"{document.document_num}_files.zip",
    )


@router.get("/revisions/{revision_id}/attachments/archive")
def download_revision_attachments_archive(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    attachments = db.query(DocumentAttachment).filter(DocumentAttachment.revision_id == revision.id).all()
    # В архив кладём и основной PDF ревизии (если загружен), и редактируемые
    # доп.файлы — одним zip. Санитайзим имена для arcname.
    safe_doc = re.sub(r"[^A-Za-z0-9._-]+", "_", document.document_num or "document")
    safe_rev = re.sub(r"[^A-Za-z0-9._-]+", "_", revision.revision_code or "")
    files_to_zip: list[tuple[Path, str]] = []
    if revision.file_path:
        pdf_path = Path(revision.file_path)
        if pdf_path.exists():
            files_to_zip.append((pdf_path, f"{safe_doc}_{safe_rev}.pdf"))
    for item in attachments:
        p = Path(item.file_path)
        if p.exists():
            files_to_zip.append((p, item.file_name))
    if not files_to_zip:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="К этой ревизии не приложены файлы")
    temp_file = tempfile.NamedTemporaryFile(prefix=f"rev_{revision.id}_", suffix=".zip", delete=False)
    temp_path = Path(temp_file.name)
    temp_file.close()
    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        seen_names: set[str] = set()
        for src, arcname in files_to_zip:
            # избегаем коллизий имён в архиве
            name = arcname
            i = 1
            while name in seen_names:
                stem = Path(arcname).stem
                suffix = Path(arcname).suffix
                name = f"{stem}_{i}{suffix}"
                i += 1
            seen_names.add(name)
            archive.write(src, arcname=name)
    return FileResponse(
        path=str(temp_path),
        media_type="application/zip",
        filename=f"{safe_doc}_{safe_rev}_files.zip",
    )


@router.get("/revisions/{revision_id}/file")
def get_revision_file(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not revision.file_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PDF file is not attached")

    path = Path(revision.file_path).resolve()
    allowed_roots = {UPLOAD_ROOT.resolve(), *LEGACY_UPLOAD_ROOTS}
    if not any(root in path.parents for root in allowed_roots):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Invalid file path")
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
    if path.suffix.lower() != ".pdf":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only PDF preview supported")

    return FileResponse(path=path, media_type="application/pdf", filename=path.name)


def _comment_attachment_read(item: CommentAttachment, uploader: User | None) -> CommentAttachmentRead:
    return CommentAttachmentRead(
        id=item.id,
        comment_id=item.comment_id,
        uploaded_by_id=item.uploaded_by_id,
        uploaded_by_name=uploader.full_name if uploader else None,
        uploaded_by_email=uploader.email if uploader else None,
        file_name=item.file_name,
        created_at=item.created_at,
    )


@router.post("/comments/{comment_id}/attachments", response_model=CommentAttachmentRead, status_code=status.HTTP_201_CREATED)
def upload_comment_attachment(
    comment_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_raise_comments")),
):
    """Прикрепить файл к замечанию. Только автор замечания (или админ);
    любой формат. Признак наличия отражается в карточке и выгрузках."""
    comment = db.query(Comment).filter(Comment.id == comment_id).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    if current_user.role.value != "admin" and comment.author_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only the comment author can attach files")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is not None and _is_completed_document(db, revision.document_id):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); uploads are locked")
    original_name = file.filename or "attachment.bin"
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).name)
    destination_dir = UPLOAD_ROOT / "comment_attachments" / str(comment.id)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / f"{uuid4().hex}_{safe_name}"
    destination.write_bytes(file.file.read())
    item = CommentAttachment(
        comment_id=comment.id,
        uploaded_by_id=current_user.id,
        file_name=safe_name,
        file_path=str(destination),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _comment_attachment_read(item, current_user)


@router.get("/remarks/{remark_number}", response_model=RemarkCardRead)
def get_remark_card(
    remark_number: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Карточка замечания по уникальному номеру (IMP-RMK-000123): где и когда
    возникло, в каком CRS ушло, что ответил подрядчик, что решил LR, перенос на
    следующие ревизии, файлы. Подрядчику видны только опубликованные замечания."""
    comment = (
        db.query(Comment)
        .filter(Comment.remark_number == remark_number.strip().upper())
        .first()
    )
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Замечание с таким номером не найдено")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None or not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Замечание с таким номером не найдено")
    if current_user.company_type == CompanyType.contractor and not comment.is_published_to_contractor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Замечание с таким номером не найдено")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first() if document else None
    if document is None or mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Документ замечания не найден")

    author = db.query(User).filter(User.id == comment.author_id).first()
    name = lambda u: (u.full_name or u.email) if u else None  # noqa: E731

    events: list[RemarkHistoryEvent] = [
        RemarkHistoryEvent(
            at=comment.created_at,
            kind="CREATED",
            title=f"Замечание создано (ревизия {revision.revision_code}"
                  + (f", лист {comment.page}" if comment.page else "") + ")",
            actor=name(author),
            detail=_clean_remark_text(comment.text),
        )
    ]
    if comment.crs_sent_at or comment.is_published_to_contractor:
        events.append(
            RemarkHistoryEvent(
                at=comment.crs_sent_at or comment.created_at,
                kind="CRS_SENT",
                title=f"Передано подрядчику{f' (CRS {comment.crs_number})' if comment.crs_number else ''}",
            )
        )
    # Ответы подрядчика и решения LR — дочерние комментарии.
    children = (
        db.query(Comment, User)
        .outerjoin(User, User.id == Comment.author_id)
        .filter(Comment.parent_id == comment.id)
        .order_by(Comment.id.asc())
        .all()
    )
    for child, child_author in children:
        raw = child.text or ""
        if raw.startswith("[LR_REJECT]") or raw.startswith("[LR_WITHDRAW]"):
            kind, title = "LR_DECISION", "Замечание снято/отклонено LR"
        elif raw.startswith("[LR_UPDATE]"):
            kind, title = "LR_DECISION", "Замечание скорректировано LR"
        else:
            kind = "CONTRACTOR_REPLY"
            label = {"A": "принято (A)", "I": "не согласен (I)"}.get(
                child.contractor_status.value if child.contractor_status else "", ""
            )
            title = f"Ответ подрядчика{f': {label}' if label else ''}"
        events.append(
            RemarkHistoryEvent(
                at=child.created_at,
                kind=kind,
                title=title,
                actor=name(child_author),
                detail=re.sub(r"^\[[A-Z_]+\]\s*", "", raw).strip() or None,
            )
        )
    # Итоговое решение по замечанию.
    if comment.status == CommentStatus.RESOLVED and comment.resolved_at:
        events.append(RemarkHistoryEvent(at=comment.resolved_at, kind="LR_DECISION", title="Замечание закрыто (устранено)"))
    elif comment.status == CommentStatus.REJECTED and comment.resolved_at:
        events.append(RemarkHistoryEvent(at=comment.resolved_at, kind="LR_DECISION", title="Замечание отклонено"))
    # Перенос на следующие ревизии (carry-over).
    carry_rows = (
        db.query(CarryOverDecision, User, Revision)
        .outerjoin(User, User.id == CarryOverDecision.decided_by_id)
        .outerjoin(Revision, Revision.id == CarryOverDecision.target_revision_id)
        .filter(CarryOverDecision.source_comment_id == comment.id)
        .order_by(CarryOverDecision.id.asc())
        .all()
    )
    carry_label = {
        "OPEN": "не устранено",
        "CLOSED": "устранено",
        "R_OPEN": "рекомендация R: не устранено",
        "R_CLOSED": "рекомендация R: устранено",
    }
    for decision, decider, target_rev in carry_rows:
        events.append(
            RemarkHistoryEvent(
                at=decision.decided_at,
                kind="CARRY_OVER",
                title=f"Проверка на ревизии {target_rev.revision_code if target_rev else '—'}: "
                      f"{carry_label.get(decision.status, decision.status)}",
                actor=name(decider),
            )
        )
    events.sort(key=lambda e: e.at)

    att_rows = (
        db.query(CommentAttachment, User)
        .outerjoin(User, User.id == CommentAttachment.uploaded_by_id)
        .filter(CommentAttachment.comment_id == comment.id)
        .order_by(CommentAttachment.id.asc())
        .all()
    )
    return RemarkCardRead(
        id=comment.id,
        remark_number=comment.remark_number,
        project_code=mdr.project_code,
        document_num=document.document_num,
        document_title=document.title,
        revision_id=revision.id,
        revision_code=revision.revision_code,
        issue_purpose=revision.issue_purpose,
        page=comment.page,
        review_code=comment.review_code,
        status=comment.status,
        contractor_status=comment.contractor_status,
        text=_clean_remark_text(comment.text),
        author_name=name(author),
        author_email=author.email if author else None,
        created_at=comment.created_at,
        crs_number=comment.crs_number,
        crs_sent_at=comment.crs_sent_at,
        is_published_to_contractor=comment.is_published_to_contractor,
        attachments=[_comment_attachment_read(item, uploader) for item, uploader in att_rows],
        history=events,
    )


@router.get("/comments/{comment_id}/attachments", response_model=list[CommentAttachmentRead])
def list_comment_attachments(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    comment = db.query(Comment).filter(Comment.id == comment_id).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None or not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    items = (
        db.query(CommentAttachment)
        .filter(CommentAttachment.comment_id == comment.id)
        .order_by(CommentAttachment.created_at.desc(), CommentAttachment.id.desc())
        .all()
    )
    user_ids = {item.uploaded_by_id for item in items}
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {user.id: user for user in users}
    return [_comment_attachment_read(item, user_map.get(item.uploaded_by_id)) for item in items]


@router.get("/comments/attachments/{attachment_id}/file")
def get_comment_attachment_file(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    item = db.query(CommentAttachment).filter(CommentAttachment.id == attachment_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    comment = db.query(Comment).filter(Comment.id == item.comment_id).first()
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first() if comment else None
    if revision is None or not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    path = Path(item.file_path).resolve()
    allowed_roots = {UPLOAD_ROOT.resolve(), *LEGACY_UPLOAD_ROOTS}
    if not any(root in path.parents for root in allowed_roots):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Invalid file path")
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
    return FileResponse(path=path, filename=item.file_name)


def _clean_remark_text(text: str | None) -> str:
    raw = (text or "").strip()
    return re.sub(r"^\[(REMARK|QUESTION)\]\s*", "", raw, flags=re.IGNORECASE)


# Ширина, при которой PDF рендерится в аннотаторе (RevisionPdfAnnotator: <Page
# width={780}>). Координаты области замечания хранятся в пикселях этого рендера
# от левого-верхнего угла; пересчёт в точки страницы: k = W_pt / 780.
_ANNOTATOR_RENDER_WIDTH = 780.0


def _register_unicode_font() -> str:
    """Зарегистрировать TTF с кириллицей для сводки. Возвращает имя шрифта
    (или 'Helvetica', если ни один TTF не найден — тогда кириллица не отрисуется,
    но рамки/номера на страницах работают в любом случае)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    name = "TdoUni"
    if name in pdfmetrics.getRegisteredFontNames():
        return name
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Docker (fonts-dejavu-core)
        "/Library/Fonts/Arial Unicode.ttf",                  # macOS
        "/System/Library/Fonts/Supplemental/Arial.ttf",       # macOS
    ]
    for p in candidates:
        if Path(p).exists():
            try:
                pdfmetrics.registerFont(TTFont(name, p))
                return name
            except Exception:  # noqa: BLE001
                continue
    return "Helvetica"


def _build_annotated_pdf(original_bytes: bytes, remarks: list[dict]) -> bytes:
    """Наложить замечания на PDF: рамка+номер для замечаний с областью,
    маркер-«звёздочка» с номером в углу листа для текстовых, и страница-сводка
    в конце. Возвращает готовый PDF-байты."""
    from io import BytesIO
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    uni = _register_unicode_font()

    reader = PdfReader(BytesIO(original_bytes))
    num_pages = len(reader.pages)

    # Раскладываем замечания по листам (1-based). Точечные (без области)
    # копим отдельно, чтобы расставить звёздочками в углу.
    by_page_area: dict[int, list[dict]] = {}
    by_page_point: dict[int, list[dict]] = {}
    for r in remarks:
        page = r.get("page") or 1
        if r.get("area_x") is not None and r.get("area_w"):
            by_page_area.setdefault(page, []).append(r)
        else:
            by_page_point.setdefault(page, []).append(r)

    def _annotation_contents(r: dict) -> str:
        head = f"#{r['number']} · {r.get('review_code') or '—'} · {r.get('author') or ''}".strip()
        if r.get("has_file"):
            head += " · [есть файл]"
        return f"{head}\n{_clean_remark_text(r.get('text'))}"

    # Настоящие PDF-аннотации (всплывающая заметка): при наведении в любом
    # просмотрщике видно текст замечания. Собираем и добавляем после того,
    # как все страницы попали в writer (индексы страниц совпадают).
    annots: list[tuple[int, tuple[float, float, float, float], str]] = []

    def _wrap(c_obj, text: str, font: str, size: float, max_w: float) -> list[str]:
        """Перенос текста по ширине панели (по словам, длинные слова режем)."""
        words = " ".join((text or "").split()).split(" ")
        lines: list[str] = []
        cur = ""
        for word in words:
            probe = f"{cur} {word}".strip()
            if c_obj.stringWidth(probe, font, size) <= max_w:
                cur = probe
                continue
            if cur:
                lines.append(cur)
            # слово длиннее строки — режем посимвольно
            while c_obj.stringWidth(word, font, size) > max_w and len(word) > 1:
                cut = len(word)
                while cut > 1 and c_obj.stringWidth(word[:cut], font, size) > max_w:
                    cut -= 1
                lines.append(word[:cut])
                word = word[cut:]
            cur = word
        if cur:
            lines.append(cur)
        return lines

    writer = PdfWriter()
    for idx in range(num_pages):
        base_page = reader.pages[idx]
        page_no = idx + 1
        mb = base_page.mediabox
        w = float(mb.width)
        h = float(mb.height)
        k = w / _ANNOTATOR_RENDER_WIDTH  # пиксели рендера → точки страницы

        page_area = by_page_area.get(page_no, [])
        page_point = by_page_point.get(page_no, [])
        # Поле справа со списком замечаний листа: сам чертёж ничем не
        # перекрывается (подписи поверх листа накладывались друг на друга).
        panel_w = max(210.0, w * 0.32) if (page_area or page_point) else 0.0
        total_w = w + panel_w

        overlay_buf = BytesIO()
        c = canvas.Canvas(overlay_buf, pagesize=(total_w, h))
        c.setStrokeColorRGB(0.85, 0.15, 0.15)
        c.setLineWidth(1.5)

        for r in page_area:
            x = float(r["area_x"]) * k
            y_top = float(r["area_y"]) * k
            rw = float(r["area_w"]) * k
            rh = float(r["area_h"]) * k
            y_bottom = h - (y_top + rh)
            c.setStrokeColorRGB(0.85, 0.15, 0.15)
            c.rect(x, y_bottom, rw, rh, stroke=1, fill=0)
            # Номер в кружке у левого-верхнего угла рамки.
            c.setFillColorRGB(0.85, 0.15, 0.15)
            c.circle(x, h - y_top, 8, stroke=0, fill=1)
            c.setFillColorRGB(1, 1, 1)
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(x, h - y_top - 3, str(r["number"]))
            # Заметка-«комментарий» левее кружка (иконка не перекрывает номер).
            ax = max(2.0, x - 22.0)
            annots.append((idx, (ax, h - y_top - 8.0, ax + 16.0, h - y_top + 8.0), _annotation_contents(r)))

        # Точечные замечания — звёздочки в правом верхнем углу чертежа.
        for i, r in enumerate(page_point):
            cx = w - 24
            cy = h - 24 - i * 22
            c.setFillColorRGB(0.85, 0.15, 0.15)
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(cx, cy, "*")
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(cx, cy - 12, str(r["number"]))
            annots.append((idx, (cx - 8.0, cy - 8.0, cx + 8.0, cy + 8.0), _annotation_contents(r)))

        # Панель со списком замечаний этого листа.
        if panel_w:
            pad = 10.0
            col_x = w + pad
            col_w = panel_w - pad * 2
            c.setFillColorRGB(0.99, 0.99, 0.99)
            c.rect(w, 0, panel_w, h, stroke=0, fill=1)
            c.setStrokeColorRGB(0.75, 0.75, 0.75)
            c.setLineWidth(0.8)
            c.line(w, 0, w, h)
            c.setFillColorRGB(0.1, 0.1, 0.1)
            c.setFont(uni, 11)
            c.drawString(col_x, h - 26, f"Замечания — лист {page_no}")
            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.line(col_x, h - 32, col_x + col_w, h - 32)

            y = h - 48
            for r in [*page_area, *page_point]:
                head = f"#{r['number']}  {r.get('review_code') or '—'}  {r.get('author') or ''}".rstrip()
                if r.get("has_file"):
                    head += "  [файл]"
                body_lines = _wrap(c, _clean_remark_text(r.get("text")), uni, 7.5, col_w)
                need = 11 + len(body_lines) * 9.5 + 8
                if y - need < 20:
                    c.setFont(uni, 7.5)
                    c.setFillColorRGB(0.4, 0.4, 0.4)
                    c.drawString(col_x, 12, "…остальные замечания — на странице-сводке в конце")
                    break
                c.setFillColorRGB(0.85, 0.15, 0.15)
                c.setFont(uni, 8)
                c.drawString(col_x, y, head[:60])
                y -= 11
                c.setFillColorRGB(0.15, 0.15, 0.15)
                c.setFont(uni, 7.5)
                for line in body_lines:
                    c.drawString(col_x, y, line)
                    y -= 9.5
                y -= 8

        c.showPage()
        c.save()
        overlay_buf.seek(0)

        overlay_page = PdfReader(overlay_buf).pages[0]
        if panel_w:
            # Лист шире исходного: кладём оригинал слева, панель — справа.
            from pypdf import PageObject
            from pypdf import Transformation

            wide = PageObject.create_blank_page(width=total_w, height=h)
            left = float(mb.left)
            bottom = float(mb.bottom)
            if left or bottom:
                wide.merge_transformed_page(base_page, Transformation().translate(-left, -bottom))
            else:
                wide.merge_page(base_page)
            wide.merge_page(overlay_page)
            writer.add_page(wide)
        else:
            base_page.merge_page(overlay_page)
            writer.add_page(base_page)

    # Вставляем заметки (не роняем экспорт, если версия pypdf не поддержит).
    try:
        from pypdf.annotations import Text as PdfTextAnnotation

        for page_idx, rect, contents in annots:
            writer.add_annotation(
                page_number=page_idx,
                annotation=PdfTextAnnotation(rect=rect, text=contents, open=False),
            )
    except Exception:  # noqa: BLE001 — аннотации-бонус поверх рамок
        pass

    # Страница-сводка со всеми замечаниями (номер, лист, автор, код, файл, текст).
    summary_buf = BytesIO()
    sc = canvas.Canvas(summary_buf, pagesize=A4)
    sw, sh = A4
    y = sh - 50
    sc.setFont(uni, 14)
    sc.drawString(40, y, "Сводка замечаний")
    y -= 24
    for r in remarks:
        if y < 50:
            sc.showPage()
            y = sh - 50
        head = f"#{r['number']}  лист {r.get('page') or '-'}  {r.get('author') or ''}  {r.get('review_code') or ''}"
        if r.get("has_file"):
            head += "  [файл]"
        sc.setFont(uni, 10)
        sc.drawString(40, y, head[:120])
        y -= 13
        sc.setFont(uni, 9)
        # Текст переносим по ~110 символов на строку (латиница/кириллица грубо).
        text = _clean_remark_text(r.get("text"))
        for line_start in range(0, len(text), 110):
            if y < 50:
                sc.showPage()
                y = sh - 50
                sc.setFont(uni, 9)
            sc.drawString(52, y, text[line_start:line_start + 110])
            y -= 12
        y -= 6
    sc.showPage()
    sc.save()
    summary_buf.seek(0)
    for p in PdfReader(summary_buf).pages:
        writer.add_page(p)

    out = BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()


@router.get("/revisions/{revision_id}/annotated.pdf")
def get_revision_annotated_pdf(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """PDF ревизии с нанесёнными замечаниями (рамки/номера/звёздочки) и
    страницей-сводкой. Доступно обеим сторонам, у кого есть доступ к ревизии."""
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not revision.file_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PDF file is not attached")
    path = Path(revision.file_path).resolve()
    allowed_roots = {UPLOAD_ROOT.resolve(), *LEGACY_UPLOAD_ROOTS}
    if not any(root in path.parents for root in allowed_roots) or not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    rows = (
        db.query(Comment, User)
        .outerjoin(User, User.id == Comment.author_id)
        .filter(
            Comment.revision_id == revision.id,
            Comment.parent_id.is_(None),
            Comment.status != CommentStatus.REJECTED,
        )
        .order_by(Comment.page.is_(None), Comment.page.asc(), Comment.id.asc())
        .all()
    )
    att_counts = _comment_attachment_counts(db, [c for c, _ in rows])
    remarks: list[dict] = []
    for i, (comment, author) in enumerate(rows, start=1):
        remarks.append({
            "number": i,
            "page": comment.page,
            "area_x": comment.area_x,
            "area_y": comment.area_y,
            "area_w": comment.area_w,
            "area_h": comment.area_h,
            "text": comment.text,
            "review_code": comment.review_code.value if comment.review_code else "",
            "author": (author.full_name or author.email) if author else "",
            "has_file": bool(att_counts.get(comment.id)),
        })

    original_bytes = path.read_bytes()
    try:
        annotated = _build_annotated_pdf(original_bytes, remarks)
    except Exception as exc:  # noqa: BLE001 — не роняем весь запрос из-за кривого PDF
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Не удалось собрать PDF с замечаниями: {exc}")

    from io import BytesIO
    from fastapi.responses import StreamingResponse

    doc = db.query(Document).filter(Document.id == revision.document_id).first()
    fname = f"{doc.document_num if doc else 'revision'}_{revision.revision_code}_annotated.pdf"
    return StreamingResponse(
        BytesIO(annotated),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={re.sub(r'[^A-Za-z0-9._-]+', '_', fname)}"},
    )


@router.get("/documents/{document_id}/revisions", response_model=list[RevisionRead])
def list_revisions(
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    if current_user.role.value != "admin":
        mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
        if mdr is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
        has_project_access = (
            db.query(ProjectMember.id)
            .join(Project, Project.id == ProjectMember.project_id)
            .filter(Project.code == mdr.project_code, ProjectMember.user_id == current_user.id)
            .first()
            is not None
        )
        if not has_project_access:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to document revisions")

    revisions = (
        db.query(Revision)
        .filter(Revision.document_id == document_id)
        .order_by(Revision.id.desc())
        .all()
    )
    if current_user.role.value != "admin" and current_user.company_type == CompanyType.owner:
        revisions = [item for item in revisions if item.status in OWNER_VISIBLE_REVISION_STATUSES]
    return revisions


@router.get("/revisions/{revision_id}", response_model=RevisionRead)
def get_revision(revision_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    rev = db.query(Revision).filter(Revision.id == revision_id).first()
    if not rev:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, rev):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    return rev


@router.post("/revisions/{revision_id}/review-code", response_model=RevisionRead)
def set_revision_review_code(
    revision_id: int,
    payload: RevisionReviewCodeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if payload.review_code != ReviewCode.AP:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only AP is supported")
    if revision.review_code == ReviewCode.AP:
        return revision

    document = db.query(Document).filter(Document.id == revision.document_id).first()
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first() if document else None
    project = db.query(Project).filter(Project.code == mdr.project_code).first() if mdr else None
    if document is None or mdr is None or project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document context not found")

    if not _is_lr_for_document(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        doc_type=None,
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only LR/Admin can set AP")

    published_parent_comments = (
        db.query(Comment)
        .filter(
            Comment.revision_id == revision.id,
            Comment.parent_id.is_(None),
            Comment.is_published_to_contractor.is_(True),
            Comment.status != CommentStatus.REJECTED,
        )
        .all()
    )
    if any(item.status in {CommentStatus.OPEN, CommentStatus.IN_PROGRESS} for item in published_parent_comments):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Resolve/reject all remarks before AP")

    latest_revision = (
        db.query(Revision)
        .filter(Revision.document_id == revision.document_id)
        .order_by(Revision.id.desc())
        .first()
    )
    if latest_revision is None or latest_revision.id != revision.id:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="AP can be set only for latest revision",
        )

    revision.review_code = ReviewCode.AP
    revision.reviewed_at = datetime.utcnow()
    review_events_service.record_event(
        db,
        revision=revision,
        document=document,
        mdr=mdr,
        actor=current_user,
        actor_role="ADMIN" if current_user.role.value == "admin" else "LR",
        event_type="AP_SET",
        note="Документ согласован (AP)",
    )
    if revision.status in {"UNDER_REVIEW", "OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I"}:
        _set_revision_status(revision, "CONTRACTOR_REPLY_A")
    receiver_code = _project_reference_value(
        db,
        project_id=project.id,
        code="TRM_RECEIVER_COMPANY_CODE",
    ) or "IVA"
    sender_code = (mdr.originator_code or "").strip().upper() or "CTR"
    auto_crs_number = _next_crs_number(
        db,
        document_id=document.id,
        project_code=mdr.project_code,
        sender_company_code=sender_code,
        receiver_company_code=receiver_code,
    )
    db.add(
        Notification(
            user_id=revision.author_id or document.created_by_id,
            event_type="OWNER_COMMENTS_PUBLISHED",
            message=(
                f"CRS по документу {document.document_num}, ревизия {revision.revision_code}: "
                f"статус AP (замечаний нет). CRS: {auto_crs_number}"
            ),
            project_code=mdr.project_code,
            document_num=document.document_num,
            revision_id=revision.id,
        )
    )
    _archive_document_notifications(
        db,
        project_code=mdr.project_code,
        document_num=document.document_num,
        revision_id=revision.id,
        event_types=[
            "TDO_SENT_TO_OWNER",
            "OWNER_COMMENT_CREATED",
            "CARRY_OVER_DECISION",
            "NEW_COMMENT",
            "COMMENT_RESPONSE",
            "OWNER_COMMENT_PUBLISHED",
            "OWNER_COMMENTS_PUBLISHED",
            "REVISION_UPLOADED_FOR_TDO",
            "NEW_REVISION_FOR_TDO",
        ],
    )
    db.add(revision)
    db.commit()
    db.refresh(revision)
    return revision


def _revision_context(db: Session, revision: Revision):
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    return document, mdr, project


def _assigned_owner_reviewers(db: Session, *, project_id: int, discipline_code: str | None):
    """Матричные ревьюеры (R/LR) уровня 1 по дисциплине документа."""
    return (
        db.query(ReviewMatrixMember)
        .filter(
            ReviewMatrixMember.project_id == project_id,
            ReviewMatrixMember.discipline_code == discipline_code,
            ReviewMatrixMember.level == 1,
            ReviewMatrixMember.state.in_(["LR", "R"]),
        )
        .all()
    )


@router.post("/revisions/{revision_id}/no-comments", response_model=RevisionReviewerSummary)
def mark_revision_no_comments(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """R/LR отмечает «Рассмотрено, без замечаний» (NC).

    LR видит, что его ревьюер посмотрел. После NC комментирование этого
    ревьюера закрывается (если включена настройка review_nc_locks_commenting)."""
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document, mdr, project = _revision_context(db, revision)
    discipline = _matrix_discipline(db, mdr)

    role = _owner_matrix_role_for_document(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=discipline,
        doc_type=_matrix_doc_type(db, mdr),
    )
    if current_user.role.value != "admin" and role not in {"LR", "R"}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нет назначения по матрице для этого документа")

    has_comments = (
        db.query(Comment.id)
        .filter(Comment.revision_id == revision.id, Comment.author_id == current_user.id)
        .first()
        is not None
    )
    if has_comments:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="У вас уже есть замечания по ревизии — «нет замечаний» недоступно",
        )

    state = review_events_service.get_or_create_reviewer_state(db, revision_id=revision.id, user_id=current_user.id)
    state.no_comments = True
    state.decided_at = datetime.utcnow()
    db.add(state)

    review_events_service.record_event(
        db,
        revision=revision,
        document=document,
        mdr=mdr,
        actor=current_user,
        actor_role="LR" if role == "LR" else "R",
        event_type="R_NO_COMMENTS",
        note="Рассмотрено, без замечаний",
    )

    # Уведомляем LR, что ревьюер посмотрел.
    actor_name = current_user.full_name or current_user.email
    for member in _assigned_owner_reviewers(db, project_id=project.id, discipline_code=discipline):
        if member.state == "LR" and member.user_id != current_user.id:
            db.add(
                Notification(
                    user_id=member.user_id,
                    event_type="R_NO_COMMENTS",
                    message=(
                        f"Ревьювер {actor_name} рассмотрел без замечаний: "
                        f"{document.document_num}, ревизия {revision.revision_code}."
                    ),
                    project_code=mdr.project_code,
                    document_num=document.document_num,
                    revision_id=revision.id,
                )
            )
    _mark_notifications_read(
        db,
        user_id=current_user.id,
        revision_id=revision.id,
        event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "NEW_COMMENT", "CARRY_OVER_DECISION", "REVIEW_DEADLINE_SOON"],
    )
    db.commit()
    return _build_reviewer_summary(db, revision, project, discipline, current_user)


def _build_reviewer_summary(db, revision, project, discipline, current_user) -> RevisionReviewerSummary:
    members = _assigned_owner_reviewers(db, project_id=project.id, discipline_code=discipline)
    states = {
        s.user_id: s
        for s in db.query(RevisionReviewerState).filter(RevisionReviewerState.revision_id == revision.id).all()
    }
    # Только ЖИВЫЕ замечания: отклонённые (REJECTED) не должны держать
    # ревьювера в статусе «есть замечания». Иначе LR, который сам отклонил
    # своё замечание и согласовал ревизию, продолжал висеть красным.
    commented_ids = {
        row[0]
        for row in db.query(Comment.author_id)
        .filter(Comment.revision_id == revision.id, Comment.status != CommentStatus.REJECTED)
        .distinct()
        .all()
    }
    user_ids = {m.user_id for m in members}
    users = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    reviewers: list[ReviewerStateRead] = []
    r_members = [m for m in members if m.state == "R"]
    for member in members:
        user = users.get(member.user_id)
        if user is None:
            continue
        st = states.get(member.user_id)
        reviewers.append(
            ReviewerStateRead(
                user_id=member.user_id,
                full_name=user.full_name or user.email,
                email=user.email,
                role=member.state,
                no_comments=bool(st and st.no_comments),
                has_comments=member.user_id in commented_ids,
                decided_at=st.decided_at if st else None,
            )
        )
    all_r_no_comments = bool(r_members) and all(
        (states.get(m.user_id) and states[m.user_id].no_comments) and (m.user_id not in commented_ids)
        for m in r_members
    )
    return RevisionReviewerSummary(
        revision_id=revision.id,
        reviewers=sorted(reviewers, key=lambda r: (r.role != "LR", r.full_name)),
        all_reviewers_no_comments=all_r_no_comments,
        approved=(revision.review_code.value if revision.review_code else None) == "AP",
        nc_locks_commenting=review_events_service.nc_locks_commenting(db),
        my_locked=review_events_service.reviewer_commenting_locked(db, revision_id=revision.id, user_id=current_user.id),
    )


@router.get("/revisions/{revision_id}/reviewer-states", response_model=RevisionReviewerSummary)
def get_revision_reviewer_states(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    _document, mdr, project = _revision_context(db, revision)
    return _build_reviewer_summary(db, revision, project, _matrix_discipline(db, mdr), current_user)


@router.get("/revisions/{revision_id}/events", response_model=list[ReviewEventRead])
def list_revision_events(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    events = (
        db.query(ReviewEvent)
        .filter(ReviewEvent.revision_id == revision.id)
        .order_by(ReviewEvent.created_at.asc(), ReviewEvent.id.asc())
        .all()
    )
    user_ids = {e.actor_id for e in events if e.actor_id} | {e.target_user_id for e in events if e.target_user_id}
    users = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    result: list[ReviewEventRead] = []
    for e in events:
        actor = users.get(e.actor_id) if e.actor_id else None
        target = users.get(e.target_user_id) if e.target_user_id else None
        result.append(
            ReviewEventRead(
                id=e.id,
                revision_id=e.revision_id,
                project_code=e.project_code,
                document_num=e.document_num,
                discipline_code=e.discipline_code,
                revision_code=e.revision_code,
                actor_id=e.actor_id,
                actor_name=(actor.full_name or actor.email) if actor else None,
                actor_role=e.actor_role,
                event_type=e.event_type,
                target_user_id=e.target_user_id,
                target_name=(target.full_name or target.email) if target else None,
                deadline=e.deadline,
                note=e.note,
                created_at=e.created_at,
            )
        )
    return result


_ACTION_EVENTS_BY_ROLE = {
    "R": {"R_NO_COMMENTS", "R_COMMENTED"},
    "LR": {"LR_COMMENTED", "LR_SENT_TO_CONTRACTOR", "AP_SET"},
}
_ACTION_LABELS = {
    "R_NO_COMMENTS": "Рассмотрено без замечаний",
    "R_COMMENTED": "Оставлено замечание",
    "LR_COMMENTED": "Оставлено замечание",
    "LR_SENT_TO_CONTRACTOR": "Замечания переданы подрядчику",
    "AP_SET": "Согласовано (AP)",
}


def _build_review_report_rows(
    db: Session,
    *,
    project_code: str | None,
    date_from: date | None,
    date_to: date | None,
) -> list[ReviewReportRow]:
    """Детализация по документам: каждое назначение на рассмотрение (кому,
    дедлайн) + закрывающее действие ревьюера + статус в срок/просрочено."""
    q = db.query(ReviewEvent).filter(ReviewEvent.event_type == "SENT_TO_OWNER")
    if project_code:
        q = q.filter(ReviewEvent.project_code == project_code)
    if date_from:
        q = q.filter(ReviewEvent.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(ReviewEvent.created_at <= datetime.combine(date_to, datetime.max.time()))
    assignments = q.order_by(ReviewEvent.created_at.asc()).all()

    # Роли ревьюеров по (project, discipline) из матрицы — чтобы знать R/LR.
    matrix = db.query(ReviewMatrixMember).filter(ReviewMatrixMember.level == 1).all()
    role_by = {(m.project_id, m.discipline_code, m.user_id): m.state for m in matrix}
    project_id_by_code = {p.code: p.id for p in db.query(Project).all()}

    # Закрывающие действия: (revision_id, actor_id) -> самое раннее действие после назначения.
    actions = (
        db.query(ReviewEvent)
        .filter(ReviewEvent.event_type.in_(["R_NO_COMMENTS", "R_COMMENTED", "LR_COMMENTED", "LR_SENT_TO_CONTRACTOR", "AP_SET"]))
        .order_by(ReviewEvent.created_at.asc())
        .all()
    )
    actions_by_key: dict[tuple[int, int], list[ReviewEvent]] = defaultdict(list)
    for a in actions:
        if a.actor_id is not None:
            actions_by_key[(a.revision_id, a.actor_id)].append(a)

    user_ids = {a.target_user_id for a in assignments if a.target_user_id}
    users = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    today = date.today()
    rows: list[ReviewReportRow] = []
    for asg in assignments:
        reviewer = users.get(asg.target_user_id) if asg.target_user_id else None
        if reviewer is None:
            continue
        pid = project_id_by_code.get(asg.project_code)
        role = role_by.get((pid, asg.discipline_code, asg.target_user_id), "R")
        allowed = _ACTION_EVENTS_BY_ROLE.get(role, _ACTION_EVENTS_BY_ROLE["R"])
        closing = next(
            (
                a
                for a in actions_by_key.get((asg.revision_id, asg.target_user_id), [])
                if a.event_type in allowed and a.created_at >= asg.created_at
            ),
            None,
        )
        deadline = asg.deadline
        if closing is not None:
            acted_at = closing.created_at
            late = bool(deadline and acted_at.date() > deadline)
            status_code = "DONE_LATE" if late else "DONE_ON_TIME"
            days_over = (acted_at.date() - deadline).days if late and deadline else None
            action_label = _ACTION_LABELS.get(closing.event_type, closing.event_type)
        else:
            acted_at = None
            overdue = bool(deadline and today > deadline)
            status_code = "OVERDUE" if overdue else "OPEN_ON_TIME"
            days_over = (today - deadline).days if overdue and deadline else None
            action_label = "Не рассмотрено"
        rows.append(
            ReviewReportRow(
                document_num=asg.document_num,
                revision_code=asg.revision_code,
                discipline_code=asg.discipline_code,
                reviewer_name=reviewer.full_name or reviewer.email,
                reviewer_role=role,
                assigned_at=asg.created_at,
                deadline=deadline,
                acted_at=acted_at,
                action_label=action_label,
                status=status_code,
                days_overdue=days_over,
            )
        )
    return rows


@router.get("/reports/review-actions", response_model=list[ReviewReportRow])
def review_actions_report(
    project_code: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin" and not get_effective_permissions(current_user).get("can_view_reporting"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Отчёт доступен только администратору")
    return _build_review_report_rows(db, project_code=project_code, date_from=date_from, date_to=date_to)


@router.get("/reports/review-actions.xlsx")
def review_actions_report_xlsx(
    project_code: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin" and not get_effective_permissions(current_user).get("can_view_reporting"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Отчёт доступен только администратору")
    from io import BytesIO
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse

    rows = _build_review_report_rows(db, project_code=project_code, date_from=date_from, date_to=date_to)
    status_label = {
        "DONE_ON_TIME": "В срок",
        "DONE_LATE": "Просрочено (сделано)",
        "OPEN_ON_TIME": "В работе",
        "OVERDUE": "Просрочено (не сделано)",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Действия R-LR"
    ws.append(["Документ", "Ревизия", "Дисциплина", "Ревьювер", "Роль", "Назначено", "Дедлайн", "Факт", "Действие", "Статус", "Дней просрочки"])
    for r in rows:
        ws.append([
            r.document_num,
            r.revision_code or "",
            r.discipline_code or "",
            r.reviewer_name,
            r.reviewer_role,
            r.assigned_at.strftime("%Y-%m-%d %H:%M") if r.assigned_at else "",
            r.deadline.strftime("%Y-%m-%d") if r.deadline else "",
            r.acted_at.strftime("%Y-%m-%d %H:%M") if r.acted_at else "",
            r.action_label,
            status_label.get(r.status, r.status),
            r.days_overdue if r.days_overdue is not None else "",
        ])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=review-actions.xlsx"},
    )


@router.get("/reports/comments-export.xlsx")
def export_comments_xlsx(
    project_code: str | None = Query(default=None),
    document_num: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Большая выгрузка всех замечаний и ответов: кто, ФИО, даты, статусы,
    код замечания, ревизия, документ, CRS. Доступна всем ролям — область
    ограничена проектами, где пользователь участник (админ — все проекты)."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from fastapi.responses import StreamingResponse

    # Область видимости: как в list_documents — только свои проекты.
    allowed_codes: set[str] | None = None
    if current_user.role.value != "admin":
        rows = (
            db.query(Project.code)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .all()
        )
        allowed_codes = {r[0] for r in rows}
        if not allowed_codes:
            allowed_codes = set()

    q = (
        db.query(Comment, Revision, Document, MDRRecord, User)
        .join(Revision, Revision.id == Comment.revision_id)
        .join(Document, Document.id == Revision.document_id)
        .join(MDRRecord, MDRRecord.id == Document.mdr_id)
        .outerjoin(User, User.id == Comment.author_id)
    )
    if project_code:
        q = q.filter(MDRRecord.project_code == project_code)
    if document_num:
        q = q.filter(Document.document_num == document_num)
    if allowed_codes is not None:
        q = q.filter(MDRRecord.project_code.in_(list(allowed_codes) or ["__none__"]))
    records = q.order_by(MDRRecord.project_code, Document.document_num, Revision.id, Comment.id).all()

    # ФИО автора родительского замечания (для строк-ответов).
    parent_ids = {c.parent_id for c, *_ in records if c.parent_id}
    parents = {c.id: c for c in db.query(Comment).filter(Comment.id.in_(parent_ids)).all()} if parent_ids else {}
    parent_author_ids = {p.author_id for p in parents.values()}
    parent_authors = (
        {u.id: u for u in db.query(User).filter(User.id.in_(parent_author_ids)).all()} if parent_author_ids else {}
    )

    status_label = {
        "OPEN": "Открыто",
        "IN_PROGRESS": "В работе",
        "RESOLVED": "Устранено",
        "REJECTED": "Отклонено",
    }
    contractor_label = {"A": "Согласен (A)", "I": "Не согласен (I)"}
    side_label = {"owner": "Заказчик", "contractor": "Подрядчик", "admin": "Администратор"}

    # Признак наличия файла у замечания (item 3) — по всем замечаниям выборки.
    att_counts = _comment_attachment_counts(db, [c for c, *_ in records])

    wb = Workbook()
    ws = wb.active
    ws.title = "Замечания"

    # Шапка-«форма» для выгрузки по конкретному документу: шифр, название,
    # текущая ревизия — чтобы экселькой можно было делиться как самостоятельным
    # документом. Для общей (проектной) выгрузки шапка не нужна.
    if document_num and records:
        _c, _rev, head_doc, head_mdr, _a = records[0]
        last_rev = (
            db.query(Revision)
            .filter(Revision.document_id == head_doc.id)
            .order_by(Revision.id.desc())
            .first()
        )
        header_pairs = [
            ("Проект", head_mdr.project_code),
            ("Шифр документа", head_doc.document_num),
            ("Название", head_doc.title or ""),
            ("Текущая ревизия", last_rev.revision_code if last_rev else ""),
        ]
        for label, value in header_pairs:
            ws.append([label, value])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([])  # пустая строка-разделитель

    header_row_idx = ws.max_row + 1
    headers = [
        "Проект", "Документ", "Ревизия", "Тип", "№ замечания", "ID", "Родит. ID",
        "Автор (ФИО)", "Сторона", "Текст", "Код замечания", "Статус",
        "Ответ подрядчика", "Файл", "В CRS", "№ CRS", "Опубл. подрядчику",
        "Дата создания", "Дата решения",
    ]
    ws.append(headers)
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True)

    def fmt(dt) -> str:
        return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

    for comment, revision, document, mdr, author in records:
        is_reply = comment.parent_id is not None
        author_name = (author.full_name or author.email) if author else "—"
        side = side_label.get(author.company_type.value, "—") if author and author.company_type else "—"
        ws.append([
            mdr.project_code,
            document.document_num,
            revision.revision_code,
            "Ответ" if is_reply else "Замечание",
            comment.remark_number or "",
            comment.id,
            comment.parent_id or "",
            author_name,
            side,
            comment.text or "",
            (comment.review_code.value if comment.review_code else ""),
            status_label.get(comment.status.value if comment.status else "", ""),
            contractor_label.get(comment.contractor_status.value if comment.contractor_status else "", ""),
            "Да" if att_counts.get(comment.id) else "",
            "Да" if comment.in_crs else "",
            comment.crs_number or "",
            "Да" if comment.is_published_to_contractor else "",
            fmt(comment.created_at),
            fmt(comment.resolved_at),
        ])

    widths = [10, 26, 8, 12, 16, 8, 10, 26, 14, 60, 14, 12, 18, 7, 8, 14, 16, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=i).column_letter].width = w

    safe_doc = re.sub(r"[^A-Za-z0-9._-]+", "_", document_num) if document_num else None
    filename = f"comments-{safe_doc}.xlsx" if safe_doc else "comments-export.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/revisions/{revision_id}/carry-decisions", response_model=list[CarryDecisionRead])
def list_carry_decisions(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if current_user.role.value != "admin" and current_user.company_type != CompanyType.owner:
        return []
    items = (
        db.query(CarryOverDecision)
        .filter(CarryOverDecision.target_revision_id == revision.id)
        .order_by(CarryOverDecision.id.desc())
        .all()
    )
    user_ids = {item.decided_by_id for item in items}
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {user.id: user for user in users}
    return [_carry_decision_read(item, user_map.get(item.decided_by_id)) for item in items]


@router.post("/revisions/{revision_id}/carry-decisions", response_model=CarryDecisionRead)
def upsert_carry_decision(
    revision_id: int,
    payload: CarryDecisionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if current_user.role.value != "admin" and current_user.company_type != CompanyType.owner:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only owner can confirm carry-over")

    source = (
        db.query(Comment)
        .join(Revision, Revision.id == Comment.revision_id)
        .filter(Comment.id == payload.source_comment_id, Comment.parent_id.is_(None))
        .first()
    )
    if source is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Source comment not found")
    source_revision = db.query(Revision).filter(Revision.id == source.revision_id).first()
    if source_revision is None or source_revision.document_id != revision.document_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Source comment is not from this document")
    if source_revision.id == revision.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Source comment must be from previous revision")
    if source.status != CommentStatus.RESOLVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only RESOLVED comments can be carried")

    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    matrix_role = _owner_matrix_role_for_document(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        doc_type=_matrix_doc_type(db, mdr),
    )
    if current_user.role.value != "admin" and matrix_role not in {"LR", "R"}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No matrix assignment for carry-over decision")

    effective_status = payload.status
    if matrix_role == "R":
        effective_status = "R_CLOSED" if payload.status == "CLOSED" else "R_OPEN"

    existing_final = (
        db.query(CarryOverDecision)
        .filter(
            CarryOverDecision.target_revision_id == revision.id,
            CarryOverDecision.source_comment_id == payload.source_comment_id,
            CarryOverDecision.status.in_(["OPEN", "CLOSED"]),
        )
        .first()
    )
    existing = None
    if matrix_role == "R":
        existing = (
            db.query(CarryOverDecision)
            .filter(
                CarryOverDecision.target_revision_id == revision.id,
                CarryOverDecision.source_comment_id == payload.source_comment_id,
                CarryOverDecision.decided_by_id == current_user.id,
                CarryOverDecision.status.in_(["R_OPEN", "R_CLOSED"]),
            )
            .first()
        )
        if existing is None:
            existing = CarryOverDecision(
                target_revision_id=revision.id,
                source_comment_id=payload.source_comment_id,
                status=effective_status,
                decided_by_id=current_user.id,
            )
        else:
            existing.status = effective_status
            existing.decided_at = datetime.utcnow()
    else:
        if existing_final is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Carry-over decision is already locked for this remark",
            )
        existing = CarryOverDecision(
            target_revision_id=revision.id,
            source_comment_id=payload.source_comment_id,
            status=effective_status,
            decided_by_id=current_user.id,
        )
        if effective_status == "CLOSED":
            source.carry_finalized = True
        elif effective_status == "OPEN":
            source.carry_finalized = False
        db.add(source)

    db.add(existing)
    recipients = (
        db.query(ReviewMatrixMember)
        .filter(
            ReviewMatrixMember.project_id == project.id,
            ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
            ReviewMatrixMember.level == 1,
            ReviewMatrixMember.state.in_(["LR", "R"]),
        )
        .all()
    )
    recipient_ids = {item.user_id for item in recipients if item.user_id != current_user.id}
    decision_label = {
        "OPEN": "не устранено",
        "CLOSED": "устранено",
        "R_OPEN": "автор считает не устранено",
        "R_CLOSED": "автор считает устранено",
    }.get(effective_status, effective_status)
    actor_name = current_user.full_name or current_user.email
    for user_id in recipient_ids:
        db.add(
            Notification(
                user_id=user_id,
                event_type="CARRY_OVER_DECISION",
                message=(
                    f"Carry-over по документу {document.document_num}, ревизия {revision.revision_code}: "
                    f"{decision_label}. Действие: {actor_name}."
                ),
                project_code=mdr.project_code,
                document_num=document.document_num,
                revision_id=revision.id,
            )
        )
    if matrix_role == "R":
        db.add(
            Notification(
                user_id=current_user.id,
                event_type="CARRY_OVER_DECISION",
                message=(
                    f"Записана рекомендация R по carry-over: {decision_label}. "
                    f"Документ {document.document_num}, ревизия {revision.revision_code}."
                ),
                project_code=mdr.project_code,
                document_num=document.document_num,
                revision_id=revision.id,
            )
        )
    db.commit()
    db.refresh(existing)
    return _carry_decision_read(existing, current_user)


@router.post("/revisions", response_model=RevisionRead, status_code=status.HTTP_201_CREATED)
def create_revision(
    payload: RevisionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.query(Document).filter(Document.id == payload.document_id).first()
    if not doc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    if _is_completed_document(db, doc.id):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); new revisions are locked")

    latest = (
        db.query(Revision)
        .filter(Revision.document_id == payload.document_id)
        .order_by(Revision.id.desc())
        .first()
    )
    latest_effective_review_code: ReviewCode | None = None
    if latest is not None:
        if latest.review_code is not None:
            latest_effective_review_code = latest.review_code
        else:
            # Fallback to calculated "worst" code from published parent remarks.
            parent_codes = (
                db.query(Comment.review_code)
                .filter(
                    Comment.revision_id == latest.id,
                    Comment.parent_id.is_(None),
                    Comment.is_published_to_contractor.is_(True),
                    Comment.status != CommentStatus.REJECTED,
                    Comment.review_code.isnot(None),
                )
                .all()
            )
            code_values: set[str] = set()
            for item in parent_codes:
                raw = item[0]
                if raw is None:
                    continue
                normalized = str(raw.value if hasattr(raw, "value") else raw).upper().strip()
                if normalized.startswith("REVIEWCODE."):
                    normalized = normalized.split(".", 1)[1]
                code_values.add(normalized)
            if "RJ" in code_values:
                latest_effective_review_code = ReviewCode.RJ
            elif "CO" in code_values:
                latest_effective_review_code = ReviewCode.CO
            elif "AN" in code_values:
                latest_effective_review_code = ReviewCode.AN
            elif "AP" in code_values:
                latest_effective_review_code = ReviewCode.AP
    allow_same_code_after_rj = (
        latest is not None
        and latest_effective_review_code == ReviewCode.RJ
        and payload.revision_code == latest.revision_code
        and payload.issue_purpose.upper() == latest.issue_purpose.upper()
    )
    duplicate = (
        db.query(Revision.id)
        .filter(Revision.document_id == payload.document_id, Revision.revision_code == payload.revision_code)
        .first()
    )
    if duplicate is not None and not allow_same_code_after_rj:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Revision code already exists for document")

    if latest is not None and latest_effective_review_code is not None:
        prev_purpose = (latest.issue_purpose or "").upper()
        next_purpose = (payload.issue_purpose or "").upper()
        if latest_effective_review_code in {ReviewCode.AN, ReviewCode.CO} and prev_purpose != next_purpose:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="For AN/CO review code, next revision must keep same issue purpose",
            )
        if latest_effective_review_code == ReviewCode.AP and prev_purpose == next_purpose:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="For AP review code, next revision must change issue purpose",
            )
        if latest_effective_review_code == ReviewCode.RJ and not allow_same_code_after_rj:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="For RJ review code, document must be reissued with same revision code and issue purpose",
            )
    if latest is not None and latest.status in {
        "REVISION_CREATED",
        "UPLOADED_WAITING_TDO",
        "UNDER_REVIEW",
        "SUBMITTED",
        "OWNER_COMMENTS_SENT",
        "CONTRACTOR_REPLY_I",
    }:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Cannot create new revision while previous revision is in progress: {latest.status}",
        )

    mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="MDR link not found for document")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Project not found for document")

    existing_count = db.query(Revision.id).filter(Revision.document_id == payload.document_id).count()
    is_initial_revision = existing_count == 0
    payload_data = payload.model_dump()
    payload_data["author_id"] = payload.author_id or current_user.id
    payload_data["status"] = "REVISION_CREATED"
    sla_days = _sla_days_for_revision(
        db,
        project_id=project.id,
        category=mdr.category,
        issue_purpose=payload.issue_purpose,
        is_initial_revision=is_initial_revision,
    )
    payload_data["review_deadline"] = (datetime.utcnow() + timedelta(days=sla_days)).date()
    rev = Revision(**payload_data)
    db.add(rev)
    db.flush()

    lead_members = (
        db.query(ProjectMember)
        .filter(
            ProjectMember.project_id == project.id,
            ProjectMember.member_role == ProjectMemberRole.contractor_tdo_lead,
        )
        .all()
    )
    lead_ids = {item.user_id for item in lead_members}
    if lead_ids:
        receiver_ids = lead_ids
    else:
        matrix_level_1 = (
        db.query(ReviewMatrixMember)
        .filter(
            ReviewMatrixMember.project_id == project.id,
            ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
            ReviewMatrixMember.level == 1,
        )
        .all()
        )
        if matrix_level_1:
            receiver_ids = {item.user_id for item in matrix_level_1}
        else:
            receiver_ids = _matrix_fallback_receivers(db)

    for receiver_id in receiver_ids:
        event_type = "NEW_REVISION_FOR_TDO" if lead_ids else "NEW_REVISION"
        previous_review_code = latest.review_code.value if latest and latest.review_code is not None else "—"
        author_label = current_user.full_name or current_user.email
        db.add(
            Notification(
                user_id=receiver_id,
                event_type=event_type,
                message=(
                    f"Выпущен документ {doc.document_num}, ревизия {rev.revision_code}. "
                    f"Дисциплина: {mdr.discipline_code}, тип: {mdr.doc_type}. "
                    f"Автор: {author_label}. Статус: {previous_review_code}"
                ),
                project_code=mdr.project_code,
                document_num=doc.document_num,
                revision_id=rev.id,
            )
        )

    db.commit()
    db.refresh(rev)
    return rev


@router.post("/revisions/{revision_id}/tdo-decision", response_model=RevisionRead)
def make_tdo_decision(
    revision_id: int,
    payload: RevisionTdoDecision,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    doc = db.query(Document).filter(Document.id == revision.document_id).first()
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if not _is_project_tdo_lead(db, project.id, current_user.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only contractor TDO lead can process queue")

    note = (payload.note or "").strip()
    if payload.action == "SEND_TO_OWNER":
        # Идемпотентность отправки. Раньше повторный клик по «В TRM» (двойной
        # клик, повторный вызов) на уже отправленной ревизии заново плодил
        # уведомления TDO_SENT_TO_OWNER и записи журнала SENT_TO_OWNER, а также
        # ПЕРЕВЫДАВАЛ trm_number через _next_trm_number — отсюда «задвоение»
        # уведомлений и скачки нумерации ТРМ. Отправить можно только ревизию,
        # реально ждущую отправки (PDF загружен, статус UPLOADED_WAITING_TDO).
        if revision.status == "UNDER_REVIEW":
            db.refresh(revision)
            return revision  # уже отправлена — ничего не делаем
        if revision.status != "UPLOADED_WAITING_TDO":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ревизию нельзя отправить на рассмотрение: ожидается загрузка PDF",
            )
        receiver_code = _project_reference_value(
            db,
            project_id=project.id,
            code="TRM_RECEIVER_COMPANY_CODE",
        ) or "IVA"
        sender_code = (current_user.company_code or "").strip().upper() or (mdr.originator_code or "").strip().upper() or "CTR"
        revision.trm_number = _next_trm_number(
            db,
            project_code=mdr.project_code,
            sender_company_code=sender_code,
            receiver_company_code=receiver_code,
        )
        _set_revision_status(revision, "UNDER_REVIEW")
        revision.reviewed_at = datetime.utcnow()
        owner_review_days = (
            _setting_days(db, "review_sla_owner_dcc_incoming_days", 1)
            + _setting_days(db, "review_sla_owner_specialist_review_days", 7)
            + _setting_days(db, "review_sla_owner_lr_approval_days", 1)
        )
        revision.review_deadline = (datetime.utcnow() + timedelta(days=owner_review_days)).date()
        matrix_level_1 = (
            db.query(ReviewMatrixMember)
            .filter(
                ReviewMatrixMember.project_id == project.id,
                ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
                ReviewMatrixMember.level == 1,
            )
            .all()
        )
        if matrix_level_1:
            receiver_ids = {item.user_id for item in matrix_level_1}
        else:
            receiver_ids = _matrix_fallback_receivers(db)
        for receiver_id in receiver_ids:
            db.add(
                Notification(
                    user_id=receiver_id,
                    event_type="TDO_SENT_TO_OWNER",
                    message=(
                        f"Ревизия {revision.revision_code} по документу {doc.document_num} "
                        f"направлена на рассмотрение заказчика."
                    ),
                    project_code=mdr.project_code,
                    document_num=doc.document_num,
                    revision_id=revision.id,
                )
            )
            # Журнал: кому улетело на рассмотрение и с каким дедлайном.
            review_events_service.record_event(
                db,
                revision=revision,
                document=doc,
                mdr=mdr,
                actor=current_user,
                actor_role="TDO",
                event_type="SENT_TO_OWNER",
                target_user_id=receiver_id,
                deadline=revision.review_deadline,
            )
        # Новая отправка на рассмотрение обнуляет прошлые отметки «нет замечаний»
        # (ревизия пересматривается заново).
        db.query(RevisionReviewerState).filter(
            RevisionReviewerState.revision_id == revision.id
        ).delete(synchronize_session=False)
        db.query(Notification).filter(
            Notification.user_id == current_user.id,
            Notification.revision_id == revision.id,
            Notification.event_type.in_(["REVISION_UPLOADED_FOR_TDO", "NEW_REVISION_FOR_TDO"]),
            Notification.is_read.is_(False),
        ).update({Notification.is_read: True}, synchronize_session=False)
    else:
        _set_revision_status(revision, "CANCELLED_BY_TDO")
        revision.trm_number = None
        revision.reviewed_at = None
        cancel_message = (
            f"Загрузка ревизии {revision.revision_code} по документу {doc.document_num} отклонена руководителем ТДО. "
            "Скорректируйте ревизию и загрузите PDF заново."
        )
        if note:
            cancel_message += f" Комментарий ТДО: {note}"
        receiver_id = revision.author_id or doc.created_by_id
        db.add(
            Notification(
                user_id=receiver_id,
                event_type="TDO_CANCELLED_REVISION",
                message=cancel_message,
                project_code=mdr.project_code,
                document_num=doc.document_num,
                revision_id=revision.id,
            )
        )
        db.query(Notification).filter(
            Notification.revision_id == revision.id,
            Notification.event_type.in_(["REVISION_UPLOADED_FOR_TDO", "NEW_REVISION_FOR_TDO"]),
            Notification.is_read.is_(False),
        ).update({Notification.is_read: True}, synchronize_session=False)

    db.add(revision)
    db.commit()
    db.refresh(revision)
    return revision


@router.post("/revisions/tdo-decision/bulk", response_model=list[RevisionRead])
def make_tdo_bulk_decision(
    payload: RevisionTdoBulkDecision,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision_ids = [revision_id for revision_id in payload.revision_ids if isinstance(revision_id, int)]
    if not revision_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No revisions selected")
    seen: set[int] = set()
    ordered_ids: list[int] = []
    for revision_id in revision_ids:
        if revision_id in seen:
            continue
        seen.add(revision_id)
        ordered_ids.append(revision_id)

    revisions = db.query(Revision).filter(Revision.id.in_(ordered_ids)).all()
    revisions_by_id = {revision.id: revision for revision in revisions}
    missing_ids = [revision_id for revision_id in ordered_ids if revision_id not in revisions_by_id]
    if missing_ids:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Revision not found: {missing_ids[0]}")

    docs_by_id: dict[int, Document] = {}
    mdr_by_id: dict[int, MDRRecord] = {}
    projects_by_code: dict[str, Project] = {}
    for revision_id in ordered_ids:
        revision = revisions_by_id[revision_id]
        doc = db.query(Document).filter(Document.id == revision.document_id).first()
        if doc is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
        mdr = db.query(MDRRecord).filter(MDRRecord.id == doc.mdr_id).first()
        if mdr is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
        project = projects_by_code.get(mdr.project_code)
        if project is None:
            project = db.query(Project).filter(Project.code == mdr.project_code).first()
            if project is None:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
            projects_by_code[mdr.project_code] = project
        if not _is_project_tdo_lead(db, project.id, current_user.id):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only contractor TDO lead can process queue")
        docs_by_id[revision.id] = doc
        mdr_by_id[revision.id] = mdr

    note = (payload.note or "").strip()
    project_codes = {mdr.project_code.upper() for mdr in mdr_by_id.values()}
    if payload.action == "SEND_TO_OWNER" and len(project_codes) != 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Select revisions from one project to form a single TRM",
        )
    shared_trm_number: str | None = None
    shared_project_code: str | None = None
    shared_receiver_code: str | None = None
    shared_sender_code: str | None = None
    for revision_id in ordered_ids:
        revision = revisions_by_id[revision_id]
        doc = docs_by_id[revision.id]
        mdr = mdr_by_id[revision.id]
        project = projects_by_code[mdr.project_code]

        if payload.action == "SEND_TO_OWNER":
            # Идемпотентность (см. одиночный tdo-decision): уже отправленную
            # ревизию пропускаем — не плодим уведомления/журнал и не
            # перевыдаём общий TRM-номер. Невалидное состояние (нет PDF) — 409.
            if revision.status == "UNDER_REVIEW":
                continue
            if revision.status != "UPLOADED_WAITING_TDO":
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Ревизию нельзя отправить на рассмотрение: ожидается загрузка PDF",
                )
            if shared_trm_number is None:
                receiver_code = _project_reference_value(
                    db,
                    project_id=project.id,
                    code="TRM_RECEIVER_COMPANY_CODE",
                ) or "IVA"
                sender_code = (current_user.company_code or "").strip().upper() or (mdr.originator_code or "").strip().upper() or "CTR"
                shared_project_code = mdr.project_code
                shared_receiver_code = receiver_code
                shared_sender_code = sender_code
                shared_trm_number = _next_trm_number(
                    db,
                    project_code=shared_project_code,
                    sender_company_code=shared_sender_code,
                    receiver_company_code=shared_receiver_code,
                )
            revision.trm_number = shared_trm_number
            _set_revision_status(revision, "UNDER_REVIEW")
            revision.reviewed_at = datetime.utcnow()
            owner_review_days = (
                _setting_days(db, "review_sla_owner_dcc_incoming_days", 1)
                + _setting_days(db, "review_sla_owner_specialist_review_days", 7)
                + _setting_days(db, "review_sla_owner_lr_approval_days", 1)
            )
            revision.review_deadline = (datetime.utcnow() + timedelta(days=owner_review_days)).date()
            matrix_level_1 = (
                db.query(ReviewMatrixMember)
                .filter(
                    ReviewMatrixMember.project_id == project.id,
                    ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
                    ReviewMatrixMember.level == 1,
                )
                .all()
            )
            if matrix_level_1:
                receiver_ids = {item.user_id for item in matrix_level_1}
            else:
                receiver_ids = _matrix_fallback_receivers(db)
            for receiver_id in receiver_ids:
                db.add(
                    Notification(
                        user_id=receiver_id,
                        event_type="TDO_SENT_TO_OWNER",
                        message=(
                            f"Ревизия {revision.revision_code} по документу {doc.document_num} "
                            f"направлена на рассмотрение заказчика."
                        ),
                        project_code=mdr.project_code,
                        document_num=doc.document_num,
                        revision_id=revision.id,
                    )
                )
            db.query(Notification).filter(
                Notification.user_id == current_user.id,
                Notification.revision_id == revision.id,
                Notification.event_type.in_(["REVISION_UPLOADED_FOR_TDO", "NEW_REVISION_FOR_TDO"]),
                Notification.is_read.is_(False),
            ).update({Notification.is_read: True}, synchronize_session=False)
        else:
            _set_revision_status(revision, "CANCELLED_BY_TDO")
            revision.trm_number = None
            revision.reviewed_at = None
            cancel_message = (
                f"Загрузка ревизии {revision.revision_code} по документу {doc.document_num} отклонена руководителем ТДО. "
                "Скорректируйте ревизию и загрузите PDF заново."
            )
            if note:
                cancel_message += f" Комментарий ТДО: {note}"
            receiver_id = revision.author_id or doc.created_by_id
            db.add(
                Notification(
                    user_id=receiver_id,
                    event_type="TDO_CANCELLED_REVISION",
                    message=cancel_message,
                    project_code=mdr.project_code,
                    document_num=doc.document_num,
                    revision_id=revision.id,
                )
            )
            db.query(Notification).filter(
                Notification.revision_id == revision.id,
                Notification.event_type.in_(["REVISION_UPLOADED_FOR_TDO", "NEW_REVISION_FOR_TDO"]),
                Notification.is_read.is_(False),
            ).update({Notification.is_read: True}, synchronize_session=False)
        db.add(revision)

    db.commit()
    result: list[Revision] = []
    for revision_id in ordered_ids:
        revision = revisions_by_id[revision_id]
        db.refresh(revision)
        result.append(revision)
    return result


@router.get("/revisions/{revision_id}/comments", response_model=list[CommentRead])
def list_comments(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rev = db.query(Revision).filter(Revision.id == revision_id).first()
    if not rev:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, rev):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")

    query = db.query(Comment, User).outerjoin(User, User.id == Comment.author_id).filter(Comment.revision_id == revision_id)
    if current_user.company_type == CompanyType.owner:
        query = query.filter(Comment.parent_id.is_(None))
    elif current_user.company_type == CompanyType.contractor:
        query = query.filter(
            Comment.parent_id.is_(None),
            Comment.is_published_to_contractor.is_(True),
            Comment.status.in_([CommentStatus.OPEN, CommentStatus.IN_PROGRESS, CommentStatus.RESOLVED]),
        )
    else:
        query = query.filter(Comment.parent_id.is_(None))
    rows = query.order_by(Comment.id.asc()).all()
    parent_comments = [comment for comment, _author in rows]
    latest_contractor_responses = _latest_contractor_responses(db, parent_comments)
    att_counts = _comment_attachment_counts(db, parent_comments)
    return [
        _comment_read(
            comment,
            author,
            contractor_response_text=latest_contractor_responses.get(comment.id, (None, None))[0],
            contractor_response_at=latest_contractor_responses.get(comment.id, (None, None))[1],
            attachment_count=att_counts.get(comment.id, 0),
        )
        for comment, author in rows
    ]


@router.get("/comments/crs-queue", response_model=list[CsrQueueItem])
def list_crs_queue(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    query = (
        db.query(Comment, Revision, Document)
        .join(Revision, Revision.id == Comment.revision_id)
        .join(Document, Document.id == Revision.document_id)
        .filter(Comment.parent_id.is_(None), Comment.in_crs.is_(True), Comment.status != CommentStatus.REJECTED)
        .order_by(Comment.created_at.desc())
    )
    if current_user.role.value != "admin":
        allowed_codes = (
            db.query(Project.code)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .filter(ProjectMember.user_id == current_user.id)
            .all()
        )
        allowed_project_codes = [row[0] for row in allowed_codes]
        if not allowed_project_codes:
            return []
        query = query.join(MDRRecord, MDRRecord.id == Document.mdr_id).filter(MDRRecord.project_code.in_(allowed_project_codes))
    rows = query.all()
    result: list[CsrQueueItem] = []
    for comment, revision, document in rows:
        if current_user.role.value != "admin":
            mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
            if mdr is None:
                continue
            project = db.query(Project).filter(Project.code == mdr.project_code).first()
            if project is None:
                continue
            if not _is_lr_for_document(
                db,
                current_user=current_user,
                project_id=project.id,
                discipline_code=_matrix_discipline(db, mdr),
                doc_type=_matrix_doc_type(db, mdr),
            ):
                continue
        result.append(
            CsrQueueItem(
                comment_id=comment.id,
                trm_number=revision.trm_number,
                crs_number=comment.crs_number,
                document_num=document.document_num,
                revision_id=revision.id,
                revision_code=revision.revision_code,
                comment_text=comment.text,
                review_code=comment.review_code,
                comment_status=comment.status,
                in_crs=comment.in_crs,
                crs_sent_at=comment.crs_sent_at,
            )
        )
    return result


@router.post("/comments/{comment_id}/add-to-crs", response_model=CommentRead)
def add_comment_to_crs(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    comment = db.query(Comment).filter(Comment.id == comment_id, Comment.parent_id.is_(None)).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    _ensure_revision_not_locked_by_ap(revision)
    document, mdr, project = _ensure_lr_can_publish_for_revision(db, current_user=current_user, revision=revision)
    if not _can_manage_owner_remark(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        comment_author_id=comment.author_id,
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No permissions to manage this remark")
    if comment.is_published_to_contractor:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Remark already published to contractor")
    if comment.in_crs:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Remark already added to CRS")
    if comment.contractor_status is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Remark already has contractor response and cannot be added to CRS")
    if comment.status == CommentStatus.REJECTED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Rejected remark cannot be added to CRS")
    sender_code = (current_user.company_code or "IVA").strip().upper() or "IVA"
    target_author = db.query(User).filter(User.id == (revision.author_id or document.created_by_id)).first()
    receiver_code = (target_author.company_code if target_author else None) or "SHP"
    receiver_code = receiver_code.strip().upper() or "SHP"
    shared_unsent = (
        db.query(Comment.crs_number)
        .join(Revision, Revision.id == Comment.revision_id)
        .filter(
            Revision.document_id == document.id,
            Comment.parent_id.is_(None),
            Comment.in_crs.is_(True),
            Comment.is_published_to_contractor.is_(False),
            Comment.crs_number.isnot(None),
        )
        .order_by(Comment.id.desc())
        .first()
    )
    if shared_unsent and shared_unsent[0]:
        comment.crs_number = shared_unsent[0]
    else:
        comment.crs_number = _next_crs_number(
            db,
            document_id=document.id,
            project_code=mdr.project_code,
            sender_company_code=sender_code,
            receiver_company_code=receiver_code,
        )
    comment.in_crs = True
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment


@router.post("/comments/crs-send", response_model=PublishCommentsResult)
def send_crs_comments(
    payload: CsrSendPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    selected_ids = [int(value) for value in payload.comment_ids]
    comments_query = db.query(Comment).filter(Comment.parent_id.is_(None), Comment.in_crs.is_(True), Comment.is_published_to_contractor.is_(False))
    if selected_ids:
        comments_query = comments_query.filter(Comment.id.in_(selected_ids))
    comments = comments_query.all()
    if not comments:
        return PublishCommentsResult(revision_id=0, published_count=0)

    touched_trm_numbers: set[str] = set()
    comment_context: list[tuple[Comment, Revision, Document, MDRRecord]] = []
    unresolved_by_revision: dict[tuple[str, str], int] = defaultdict(int)
    for comment in comments:
        revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
        if revision is None:
            continue
        document, mdr, _project = _ensure_lr_can_publish_for_revision(db, current_user=current_user, revision=revision)
        comment_context.append((comment, revision, document, mdr))
        if revision.trm_number:
            touched_trm_numbers.add(revision.trm_number)

    if touched_trm_numbers:
        pending_rows = (
            db.query(Comment, Revision, Document)
            .join(Revision, Revision.id == Comment.revision_id)
            .join(Document, Document.id == Revision.document_id)
            .filter(
                Comment.parent_id.is_(None),
                Revision.trm_number.in_(list(touched_trm_numbers)),
                Comment.in_crs.is_(False),
                Comment.status != CommentStatus.REJECTED,
            )
            .all()
        )
        for _pending_comment, pending_revision, pending_document in pending_rows:
            unresolved_by_revision[(pending_document.document_num, pending_revision.revision_code)] += 1
    if unresolved_by_revision:
        details = "; ".join(
            f"{document_num} / rev {revision_code} - {count} неотработанных"
            for (document_num, revision_code), count in sorted(unresolved_by_revision.items())
        )
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"CRS нельзя отправить: есть неотработанные замечания по TRM ({details}). "
            "Отработанные замечания: только отклоненные LR или добавленные в CRS.",
        )

    crs_number_by_document: dict[int, str] = {}
    revisions_touched: set[int] = set()
    for comment, revision, document, mdr in comment_context:
        author = db.query(User).filter(User.id == (revision.author_id or document.created_by_id)).first()
        sender_code = (current_user.company_code or "IVA").strip().upper() or "IVA"
        receiver_code = (author.company_code if author else None) or "SHP"
        receiver_code = receiver_code.strip().upper() or "SHP"
        if document.id not in crs_number_by_document:
            crs_number_by_document[document.id] = _next_crs_number(
                db,
                document_id=document.id,
                project_code=mdr.project_code,
                sender_company_code=sender_code,
                receiver_company_code=receiver_code,
            )
        if comment.crs_number:
            crs_number_by_document[document.id] = comment.crs_number
        comment.is_published_to_contractor = True
        comment.crs_sent_at = datetime.utcnow()
        if not comment.crs_number:
            comment.crs_number = crs_number_by_document.get(document.id)
        db.add(comment)
        revisions_touched.add(comment.revision_id)

    for revision_id in revisions_touched:
        revision = db.query(Revision).filter(Revision.id == revision_id).first()
        if revision is None:
            continue
        document = db.query(Document).filter(Document.id == revision.document_id).first()
        if document is None:
            continue
        mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
        if mdr is None:
            continue
        project = db.query(Project).filter(Project.code == mdr.project_code).first()
        if project is None:
            continue

        _set_revision_status(revision, "OWNER_COMMENTS_SENT")
        revision.review_deadline = (
            datetime.utcnow() + timedelta(days=_setting_days(db, "review_sla_contractor_consideration_days", 0.5))
        ).date()
        db.add(revision)
        target_author_id = revision.author_id or document.created_by_id
        _upsert_crs_notification(
            db,
            user_id=target_author_id,
            revision=revision,
            document=document,
            mdr=mdr,
            crs_number=crs_number_by_document.get(document.id),
        )

        all_members = db.query(ProjectMember).filter(ProjectMember.project_id == project.id).all()
        member_ids = {member.user_id for member in all_members}
        matrix_users = (
            db.query(ReviewMatrixMember.user_id)
            .filter(ReviewMatrixMember.project_id == project.id, ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr))
            .all()
        )
        for row in matrix_users:
            member_ids.add(row[0])
        for member_id in member_ids:
            _mark_notifications_read(
                db,
                user_id=member_id,
                revision_id=revision.id,
                event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT", "COMMENT_RESPONSE", "REVIEW_DEADLINE_SOON", "R_NO_COMMENTS"],
            )
            (
                db.query(Notification)
                .filter(
                    Notification.user_id == member_id,
                    Notification.project_code == mdr.project_code,
                    Notification.document_num == document.document_num,
                    Notification.is_read.is_(False),
                    Notification.event_type.in_(
                        [
                            "TDO_SENT_TO_OWNER",
                            "OWNER_COMMENT_CREATED",
                            "CARRY_OVER_DECISION",
                            "NEW_COMMENT",
                            "COMMENT_RESPONSE",
                            "OWNER_COMMENT_PUBLISHED",
                            "OWNER_COMMENTS_PUBLISHED",
                            "REVISION_UPLOADED_FOR_TDO",
                            "NEW_REVISION_FOR_TDO",
                            "REVIEW_DEADLINE_SOON",
                            "R_NO_COMMENTS",
                        ]
                    ),
                )
                .update({Notification.is_read: True}, synchronize_session=False)
            )
        _archive_document_notifications(
            db,
            project_code=mdr.project_code,
            document_num=document.document_num,
            revision_id=revision.id,
            event_types=[
                "TDO_SENT_TO_OWNER",
                "OWNER_COMMENT_CREATED",
                "CARRY_OVER_DECISION",
                "NEW_COMMENT",
                "COMMENT_RESPONSE",
                "OWNER_COMMENT_PUBLISHED",
                "OWNER_COMMENTS_PUBLISHED",
                "REVISION_UPLOADED_FOR_TDO",
                "NEW_REVISION_FOR_TDO",
                "REVIEW_DEADLINE_SOON",
                "R_NO_COMMENTS",
            ],
        )

    db.commit()
    first_revision_id = comments[0].revision_id if comments else 0
    return PublishCommentsResult(revision_id=first_revision_id, published_count=len(comments))


@router.post("/comments", response_model=CommentRead, status_code=status.HTTP_201_CREATED)
def create_comment(
    payload: CommentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_raise_comments")),
):
    rev = db.query(Revision).filter(Revision.id == payload.revision_id).first()
    if not rev:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, rev):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if _is_completed_document(db, rev.document_id):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); commenting is locked")
    _ensure_revision_not_locked_by_ap(rev)

    document = db.query(Document).filter(Document.id == rev.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    if current_user.company_type == CompanyType.owner and current_user.role.value != "admin":
        if rev.status in {"OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I", "CONTRACTOR_REPLY_A"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Comments are locked for owner side after LR sends remarks to contractor",
            )
        matrix_match = (
            db.query(ReviewMatrixMember.id)
            .filter(
                ReviewMatrixMember.project_id == project.id,
                ReviewMatrixMember.user_id == current_user.id,
                ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
                ReviewMatrixMember.doc_type == mdr.doc_type,
                ReviewMatrixMember.level == 1,
                ReviewMatrixMember.state.in_(["LR", "R"]),
            )
            .first()
        )
        if matrix_match is None:
            # Fallback for legacy data where MDR doc_type may not match matrix doc_type.
            matrix_match = (
                db.query(ReviewMatrixMember.id)
                .filter(
                    ReviewMatrixMember.project_id == project.id,
                    ReviewMatrixMember.user_id == current_user.id,
                    ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
                    ReviewMatrixMember.level == 1,
                    ReviewMatrixMember.state.in_(["LR", "R"]),
                )
                .first()
            )
        if matrix_match is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No matrix assignment for this document")
        if payload.review_code is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="review_code is required for owner comment")
        # После «нет замечаний» (NC) комментирование этого R закрыто — открыть
        # может только админ, сняв глобальную настройку review_nc_locks_commenting.
        if review_events_service.reviewer_commenting_locked(db, revision_id=rev.id, user_id=current_user.id):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Комментирование закрыто: вы отметили «нет замечаний». Открыть может только администратор.",
            )

    comment = Comment(**payload.model_dump(), author_id=current_user.id, is_published_to_contractor=False)
    # Уникальный номер присваиваем только «настоящим» замечаниям (родительским).
    # Ответы-реплики (parent_id) своего номера не получают — они часть истории
    # родительского замечания.
    if comment.parent_id is None:
        comment.remark_number = _next_remark_number(db, project_code=mdr.project_code)
    db.add(comment)
    db.flush()

    if current_user.company_type == CompanyType.owner and current_user.role.value != "admin":
        owner_role = _owner_matrix_role_for_document(
            db,
            current_user=current_user,
            project_id=project.id,
            discipline_code=_matrix_discipline(db, mdr),
            doc_type=_matrix_doc_type(db, mdr),
        )
        review_events_service.record_event(
            db,
            revision=rev,
            document=document,
            mdr=mdr,
            actor=current_user,
            actor_role="LR" if owner_role == "LR" else "R",
            event_type="LR_COMMENTED" if owner_role == "LR" else "R_COMMENTED",
            note=f"Замечание ({payload.review_code.value})",
        )

    # Замечание заказчика (R/LR) на этапе рассмотрения — ВНУТРЕННЕЕ. Подрядчик
    # (создатель документа: рук. ТДО / разработчик) НЕ должен получать его до
    # отправки CRS. Уведомляем только LR (консолидирует и отправит CRS); при
    # отправке CRS подрядчик получит OWNER_COMMENTS_PUBLISHED. Раньше сюда
    # ошибочно попадал document.created_by_id → замечания «падали» подрядчику
    # сразу.
    recipients: set[int] = set()
    lr_rows = (
        db.query(ReviewMatrixMember)
        .filter(
            ReviewMatrixMember.project_id == project.id,
            ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
            ReviewMatrixMember.level == 1,
            ReviewMatrixMember.state == "LR",
        )
        .all()
    )
    for row in lr_rows:
        if row.user_id != current_user.id:
            recipients.add(row.user_id)

    for receiver_id in recipients:
        db.add(
            Notification(
                user_id=receiver_id,
                event_type="OWNER_COMMENT_CREATED",
                message=(
                    f"Новое замечание по ревизии {rev.revision_code}. "
                    f"Автор: {current_user.full_name or current_user.email}. "
                    f"Статус: {payload.review_code.value if payload.review_code is not None else '—'}"
                ),
                project_code=mdr.project_code,
                document_num=document.document_num,
                revision_id=rev.id,
            )
        )
    # Actor completed their own "review this revision" task by creating remark.
    _mark_notifications_read(
        db,
        user_id=current_user.id,
        revision_id=rev.id,
        event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT", "REVIEW_DEADLINE_SOON"],
    )

    db.commit()
    db.refresh(comment)
    return comment


@router.post("/comments/{comment_id}/response", response_model=CommentRead)
def respond_comment(
    comment_id: int,
    payload: CommentResponse,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_respond_comments")),
):
    parent = db.query(Comment).filter(Comment.id == comment_id).first()
    if not parent:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    if parent.author_id == current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot respond to own comment")
    if current_user.company_type == CompanyType.contractor and not parent.is_published_to_contractor:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Comment is not yet published to contractor")
    revision = db.query(Revision).filter(Revision.id == parent.revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if _is_completed_document(db, revision.document_id):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Document is completed (AFD+AP); commenting is locked")
    _ensure_revision_not_locked_by_ap(revision)
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")

    response = Comment(
        revision_id=parent.revision_id,
        parent_id=parent.id,
        author_id=current_user.id,
        text=payload.text,
        status=payload.status,
    )
    parent.status = payload.status
    if payload.status == CommentStatus.RESOLVED:
        parent.resolved_at = datetime.utcnow()
    elif payload.status == CommentStatus.IN_PROGRESS:
        parent.resolved_at = None
    if payload.backlog_status:
        parent.backlog_status = payload.backlog_status
    if current_user.company_type == CompanyType.contractor and payload.contractor_status is not None:
        # Allowed sequence:
        # 1) first contractor answer: I or A
        # 2) optional LR final confirmation after I (single iteration)
        # 3) contractor can send only A as final acceptance
        if parent.contractor_status is None:
            parent.contractor_status = payload.contractor_status
        elif (
            parent.contractor_status == ContractorCommentStatus.I
            and parent.backlog_status == "LR_FINAL_CONFIRM"
            and payload.contractor_status == ContractorCommentStatus.A
        ):
            parent.contractor_status = ContractorCommentStatus.A
        else:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Contractor status already finalized for this remark")

        if parent.contractor_status == ContractorCommentStatus.A and parent.backlog_status == "LR_FINAL_CONFIRM":
            parent.backlog_status = "IN_NEXT_REVISION"

    db.add(response)
    db.add(parent)
    if current_user.company_type == CompanyType.contractor:
        _recompute_revision_contractor_status(db, revision)
        document = db.query(Document).filter(Document.id == revision.document_id).first()
        mdr = db.query(MDRRecord).join(Document, Document.mdr_id == MDRRecord.id).filter(Document.id == revision.document_id).first()
        project = db.query(Project).filter(Project.code == (mdr.project_code if mdr else "")).first() if mdr else None
        if project and document:
            if revision.status == "CONTRACTOR_REPLY_I":
                revision.review_deadline = (
                    datetime.utcnow() + timedelta(days=_setting_days(db, "review_sla_contractor_consideration_days", 0.5))
                ).date()
            elif revision.status == "CONTRACTOR_REPLY_A":
                published = (
                    db.query(Comment)
                    .filter(
                        Comment.revision_id == revision.id,
                        Comment.parent_id.is_(None),
                        Comment.is_published_to_contractor.is_(True),
                    )
                    .all()
                )
                if any(item.review_code == ReviewCode.AN for item in published):
                    key = "review_sla_contractor_an_issue_days"
                    default_days = 5
                elif any(item.review_code in {ReviewCode.CO, ReviewCode.RJ} for item in published):
                    key = "review_sla_contractor_co_rj_issue_days"
                    default_days = 8
                else:
                    key = "review_sla_contractor_ap_issue_days"
                    default_days = 2
                revision.review_deadline = (datetime.utcnow() + timedelta(days=_setting_days(db, key, default_days))).date()
        db.add(revision)
        _mark_notifications_read(
            db,
            user_id=current_user.id,
            revision_id=revision.id,
            event_types=["OWNER_COMMENTS_PUBLISHED", "OWNER_COMMENT_PUBLISHED", "NEW_COMMENT"],
        )
        if mdr and document:
            _archive_document_notifications(
                db,
                project_code=mdr.project_code,
                document_num=document.document_num,
                revision_id=revision.id,
                event_types=["OWNER_COMMENTS_PUBLISHED", "OWNER_COMMENT_PUBLISHED"],
            )

    document = db.query(Document).filter(Document.id == revision.document_id).first()
    mdr = db.query(MDRRecord).filter(MDRRecord.id == (document.mdr_id if document else -1)).first() if document else None
    project = db.query(Project).filter(Project.code == mdr.project_code).first() if mdr else None
    contractor_status = payload.contractor_status.value if payload.contractor_status is not None else None
    if current_user.company_type != CompanyType.contractor or contractor_status == "I":
        receiver_ids: set[int] = {parent.author_id}
        if project is not None and mdr is not None:
            lr_rows = (
                db.query(ReviewMatrixMember)
                .filter(
                    ReviewMatrixMember.project_id == project.id,
                    ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
                    ReviewMatrixMember.level == 1,
                    ReviewMatrixMember.state == "LR",
                )
                .all()
            )
            for row in lr_rows:
                receiver_ids.add(row.user_id)
        receiver_ids.discard(current_user.id)
        for receiver_id in receiver_ids:
            db.add(
                Notification(
                    user_id=receiver_id,
                    event_type="COMMENT_RESPONSE",
                    message=(
                        f"Подрядчик ответил по замечанию #{parent.id}: статус {contractor_status or '—'}"
                        if current_user.company_type == CompanyType.contractor
                        else f"Response received for comment #{parent.id}"
                    ),
                    project_code=mdr.project_code if mdr else None,
                    document_num=document.document_num if document else None,
                    revision_id=revision.id,
                )
            )

    db.commit()
    db.refresh(response)
    return response


@router.post("/comments/{comment_id}/publish", response_model=CommentRead)
def publish_comment_to_contractor(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    comment = db.query(Comment).filter(Comment.id == comment_id).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    if comment.is_published_to_contractor:
        return comment
    comment.is_published_to_contractor = True
    db.add(comment)
    target_author_id = revision.author_id or document.created_by_id
    # Не плодим уведомление на каждое замечание — одна сводка на ревизию.
    _upsert_crs_notification(db, user_id=target_author_id, revision=revision, document=document, mdr=mdr)
    db.commit()
    db.refresh(comment)
    return comment


@router.post("/comments/{comment_id}/owner-decision", response_model=CommentRead)
def owner_comment_decision(
    comment_id: int,
    payload: CommentOwnerDecision,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    comment = db.query(Comment).filter(Comment.id == comment_id, Comment.parent_id.is_(None)).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    _ensure_revision_not_locked_by_ap(revision)
    if payload.action in {"PUBLISH", "WITHDRAW", "UPDATE", "FINAL_CONFIRM", "REJECT", "REOPEN"} and not _can_manage_owner_remark(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        comment_author_id=comment.author_id,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No permissions to manage this remark",
        )

    # При REJECT исключение: если подрядчик ответил «I» (не согласен),
    # LR имеет право согласиться с ним и снять замечание. Это нормальный
    # рабочий сценарий обсуждения, не нарушение линейности.
    contractor_status_is_i = (
        comment.contractor_status == ContractorCommentStatus.I
    )
    if (
        comment.is_published_to_contractor
        and payload.action in {"UPDATE", "WITHDRAW", "REJECT", "PUBLISH", "REOPEN"}
        and not (payload.action == "REJECT" and contractor_status_is_i)
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Remark is already sent to contractor and cannot be changed by LR/R",
        )
    if (
        comment.contractor_status is not None
        and payload.action in {"UPDATE", "REJECT", "WITHDRAW", "REOPEN", "PUBLISH"}
        and not (payload.action == "REJECT" and contractor_status_is_i)
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Remark already has contractor response and cannot be changed",
        )

    if payload.action == "PUBLISH":
        if comment.is_published_to_contractor and comment.contractor_status == ContractorCommentStatus.A:
            db.refresh(comment)
            return comment
        comment.is_published_to_contractor = True
        comment.contractor_status = ContractorCommentStatus.A
        target_author_id = revision.author_id or document.created_by_id
        _upsert_crs_notification(db, user_id=target_author_id, revision=revision, document=document, mdr=mdr)
        _mark_notifications_read(
            db,
            user_id=current_user.id,
            revision_id=revision.id,
            event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT"],
        )
    elif payload.action == "REJECT":
        # in_crs блокирует REJECT, ПОКА подрядчик не ответил. Когда
        # контрактор ответил «I» (не согласен), LR имеет право
        # согласиться с ним и снять замечание — это нормальный
        # сценарий обсуждения, даже если замечание было в CRS-пакете.
        if comment.in_crs and not contractor_status_is_i:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Remark already added to CRS and cannot be rejected")
        if comment.status == CommentStatus.REJECTED and comment.backlog_status == "REJECTED":
            db.refresh(comment)
            return comment
        comment.status = CommentStatus.REJECTED
        comment.backlog_status = "REJECTED"
        comment.resolved_at = datetime.utcnow()
        # Журнал: кто и какое замечание отклонил (чтобы история не выглядела
        # так, будто замечание всё ещё висит нерешённым).
        review_events_service.record_event(
            db,
            revision=revision,
            document=document,
            mdr=mdr,
            actor=current_user,
            actor_role="LR",
            event_type="COMMENT_REJECTED",
            target_user_id=comment.author_id,
            note=(comment.text or "")[:200],
        )
        if payload.note:
            db.add(
                Comment(
                    revision_id=comment.revision_id,
                    parent_id=comment.id,
                    author_id=current_user.id,
                    text=f"[LR_REJECT] {payload.note}",
                    status=CommentStatus.REJECTED,
                )
            )
    elif payload.action == "WITHDRAW":
        if comment.in_crs:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Remark already added to CRS and cannot be withdrawn")
        if comment.status == CommentStatus.REJECTED and comment.backlog_status == "REJECTED":
            db.refresh(comment)
            return comment
        comment.status = CommentStatus.REJECTED
        comment.backlog_status = "REJECTED"
        comment.resolved_at = datetime.utcnow()
        review_events_service.record_event(
            db,
            revision=revision,
            document=document,
            mdr=mdr,
            actor=current_user,
            actor_role="LR",
            event_type="COMMENT_REJECTED",
            target_user_id=comment.author_id,
            note=(comment.text or "")[:200],
        )
        if payload.note:
            db.add(
                Comment(
                    revision_id=comment.revision_id,
                    parent_id=comment.id,
                    author_id=current_user.id,
                    text=f"[LR_WITHDRAW] {payload.note}",
                    status=CommentStatus.REJECTED,
                )
            )
    elif payload.action == "UPDATE":
        if payload.text is not None and payload.text.strip():
            comment.text = payload.text.strip()
        if payload.review_code is not None:
            comment.review_code = payload.review_code
        # After LR/R correction, table should reflect who edited and when.
        comment.author_id = current_user.id
        comment.created_at = datetime.utcnow()
        if payload.note:
            db.add(
                Comment(
                    revision_id=comment.revision_id,
                    parent_id=comment.id,
                    author_id=current_user.id,
                    text=f"[LR_UPDATE] {payload.note}",
                    status=comment.status,
                )
            )
    elif payload.action == "REOPEN":
        if comment.status != CommentStatus.REJECTED and comment.backlog_status != "REJECTED":
            db.refresh(comment)
            return comment
        comment.status = CommentStatus.OPEN
        comment.backlog_status = None
        comment.resolved_at = None
        if payload.note:
            db.add(
                Comment(
                    revision_id=comment.revision_id,
                    parent_id=comment.id,
                    author_id=current_user.id,
                    text=f"[LR_REOPEN] {payload.note}",
                    status=CommentStatus.OPEN,
                )
            )
    elif payload.action == "FINAL_CONFIRM":
        if comment.contractor_status != ContractorCommentStatus.I:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Final confirmation is available only after contractor status I")
        if comment.backlog_status == "LR_FINAL_CONFIRM":
            db.refresh(comment)
            return comment
        note = (payload.note or "").strip()
        if len(note) < 3:
            note = "LR финально подтвердил замечание. Исполнение обязательно."
        comment.backlog_status = "LR_FINAL_CONFIRM"
        comment.status = CommentStatus.RESOLVED
        comment.resolved_at = datetime.utcnow()
        comment.contractor_status = ContractorCommentStatus.A
        comment.is_published_to_contractor = True
        db.add(
            Comment(
                revision_id=comment.revision_id,
                parent_id=comment.id,
                author_id=current_user.id,
                text=f"[LR_FINAL_CONFIRM] {note}",
                status=CommentStatus.RESOLVED,
            )
        )
    prev_status = revision.status
    _recompute_revision_contractor_status(db, revision)
    db.add(revision)
    if prev_status == "CONTRACTOR_REPLY_I" and revision.status != "CONTRACTOR_REPLY_I":
        _archive_document_notifications(
            db,
            project_code=mdr.project_code,
            document_num=document.document_num,
            revision_id=revision.id,
            event_types=["COMMENT_RESPONSE", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT"],
        )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment


@router.delete("/comments/{comment_id}")
def delete_owner_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    comment = db.query(Comment).filter(Comment.id == comment_id, Comment.parent_id.is_(None)).first()
    if comment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comment not found")
    revision = db.query(Revision).filter(Revision.id == comment.revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    _ensure_revision_not_locked_by_ap(revision)
    if not _can_manage_owner_remark(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        comment_author_id=comment.author_id,
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No permissions to manage this remark")
    if comment.is_published_to_contractor or comment.in_crs or comment.contractor_status is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete remark after CRS/publication/contractor response",
        )

    child_rows = db.query(Comment).filter(Comment.parent_id == comment.id).all()
    for child in child_rows:
        db.delete(child)
    db.delete(comment)
    _recompute_revision_contractor_status(db, revision)
    db.add(revision)
    db.commit()
    return {"ok": True}


@router.post("/revisions/{revision_id}/comments/publish-all", response_model=PublishCommentsResult)
def publish_revision_comments_to_contractor(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions("can_publish_comments")),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")
    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if not _is_lr_for_document(
        db,
        current_user=current_user,
        project_id=project.id,
        discipline_code=_matrix_discipline(db, mdr),
        doc_type=_matrix_doc_type(db, mdr),
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Только LR для дисциплины документа может передать замечания подрядчику",
        )

    comments = (
        db.query(Comment)
        .filter(Comment.revision_id == revision_id, Comment.parent_id.is_(None), Comment.is_published_to_contractor.is_(False))
        .all()
    )
    for comment in comments:
        comment.is_published_to_contractor = True
        db.add(comment)

    if comments:
        _set_revision_status(revision, "OWNER_COMMENTS_SENT")
        db.add(revision)
        review_events_service.record_event(
            db,
            revision=revision,
            document=document,
            mdr=mdr,
            actor=current_user,
            actor_role="ADMIN" if current_user.role.value == "admin" else "LR",
            event_type="LR_SENT_TO_CONTRACTOR",
            note=f"Передано замечаний: {len(comments)}",
        )
        target_author_id = revision.author_id or document.created_by_id
        db.add(
            Notification(
                user_id=target_author_id,
                event_type="OWNER_COMMENTS_PUBLISHED",
                message=f"Замечания заказчика переданы подрядчику: {document.document_num}, ревизия {revision.revision_code}",
                project_code=mdr.project_code,
                document_num=document.document_num,
                revision_id=revision.id,
            )
        )
        _mark_notifications_read(
            db,
            user_id=current_user.id,
            revision_id=revision.id,
            event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT"],
        )
        _archive_document_notifications(
            db,
            project_code=mdr.project_code,
            document_num=document.document_num,
            revision_id=revision.id,
            event_types=["TDO_SENT_TO_OWNER", "OWNER_COMMENT_CREATED", "CARRY_OVER_DECISION", "NEW_COMMENT", "COMMENT_RESPONSE"],
        )
    db.commit()
    return PublishCommentsResult(revision_id=revision_id, published_count=len(comments))


@router.get("/revisions/{revision_id}/card", response_model=RevisionCardRead)
def get_revision_card(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    revision = db.query(Revision).filter(Revision.id == revision_id).first()
    if revision is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    if not _owner_can_access_revision(db, current_user, revision):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Revision not found")
    document = db.query(Document).filter(Document.id == revision.document_id).first()
    if document is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Document not found")
    mdr = db.query(MDRRecord).filter(MDRRecord.id == document.mdr_id).first()
    if mdr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MDR not found")

    project = db.query(Project).filter(Project.code == mdr.project_code).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    matrix_role: str | None = None
    lr_reviewer_name: str | None = None
    developer_name: str | None = None
    can_owner_raise_comments = True
    if current_user.company_type == CompanyType.owner and current_user.role.value != "admin":
        can_owner_raise_comments = False
        matrix_row = (
            db.query(ReviewMatrixMember)
            .filter(
                ReviewMatrixMember.project_id == project.id,
                ReviewMatrixMember.user_id == current_user.id,
                ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
            )
            .order_by(ReviewMatrixMember.level.asc(), ReviewMatrixMember.id.asc())
            .first()
        )
        if matrix_row is not None:
            can_owner_raise_comments = True
            matrix_role = matrix_row.state

    revisions = (
        db.query(Revision)
        .filter(Revision.document_id == document.id)
        .order_by(Revision.id.asc())
        .all()
    )
    developer_user = db.query(User).filter(User.id == (revision.author_id or document.created_by_id)).first()
    developer_name = developer_user.full_name if developer_user and developer_user.full_name else (developer_user.email if developer_user else None)
    lr_rows = (
        db.query(ReviewMatrixMember, User)
        .join(User, User.id == ReviewMatrixMember.user_id)
        .filter(
            ReviewMatrixMember.project_id == project.id,
            ReviewMatrixMember.discipline_code == _matrix_discipline(db, mdr),
            ReviewMatrixMember.level == 1,
            ReviewMatrixMember.state == "LR",
        )
        .order_by(ReviewMatrixMember.id.asc())
        .all()
    )
    if lr_rows:
        unique_lr_names: list[str] = []
        for _matrix, lr_user in lr_rows:
            label = (lr_user.full_name or lr_user.email or "").strip()
            if label and label not in unique_lr_names:
                unique_lr_names.append(label)
        if unique_lr_names:
            lr_reviewer_name = ", ".join(unique_lr_names)
    history: list[RevisionCommentThreadRead] = []
    for item in revisions:
        comments_query = (
            db.query(Comment, User)
            .outerjoin(User, User.id == Comment.author_id)
            .filter(Comment.revision_id == item.id, Comment.parent_id.is_(None))
        )
        if current_user.company_type == CompanyType.contractor:
            comments_query = comments_query.filter(
                Comment.is_published_to_contractor.is_(True),
                Comment.status.in_([CommentStatus.OPEN, CommentStatus.IN_PROGRESS, CommentStatus.RESOLVED]),
            )
        comments_rows = comments_query.order_by(Comment.created_at.asc(), Comment.id.asc()).all()
        parent_comments = [comment for comment, _author in comments_rows]
        latest_contractor_responses = _latest_contractor_responses(db, parent_comments)
        att_counts = _comment_attachment_counts(db, parent_comments)
        history.append(
            RevisionCommentThreadRead(
                revision_id=item.id,
                revision_code=item.revision_code,
                status=item.status,
                created_at=item.created_at,
                comments=[
                    _comment_read(
                        comment,
                        author,
                        contractor_response_text=latest_contractor_responses.get(comment.id, (None, None))[0],
                        contractor_response_at=latest_contractor_responses.get(comment.id, (None, None))[1],
                        attachment_count=att_counts.get(comment.id, 0),
                    )
                    for comment, author in comments_rows
                ],
            )
        )

    first_upload_at = None
    for item in revisions:
        if item.file_path and (first_upload_at is None or item.created_at < first_upload_at):
            first_upload_at = item.created_at

    latest_issue_at = revisions[-1].created_at if revisions else None
    planned_issue_date = mdr.planned_dev_start + timedelta(days=14) if mdr.planned_dev_start else None
    def _is_sent_to_owner(rev: Revision | None) -> bool:
        if rev is None:
            return False
        return bool(rev.reviewed_at) or (
            bool(rev.trm_number)
            and rev.status
            in {"UNDER_REVIEW", "OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_I", "CONTRACTOR_REPLY_A", "SUBMITTED"}
        )

    rev_a = next(
        (item for item in revisions if (item.revision_code or "").upper() == "A" and (item.issue_purpose or "").upper() == "IFR"),
        None,
    )
    rev_b = next(
        (item for item in revisions if (item.revision_code or "").upper() == "B" and (item.issue_purpose or "").upper() == "IFR"),
        None,
    )
    rev_00 = next(
        (item for item in revisions if (item.revision_code or "") == "00" and (item.issue_purpose or "").upper() == "AFD"),
        None,
    )

    actual_progress = 0.0
    if _is_sent_to_owner(rev_a):
        actual_progress = max(actual_progress, 70.0)
    if rev_a is not None and rev_a.status in {"CONTRACTOR_REPLY_A", "SUBMITTED"}:
        actual_progress = max(actual_progress, 75.0)
    if _is_sent_to_owner(rev_b):
        actual_progress = max(actual_progress, 80.0)
    if rev_b is not None and rev_b.status in {"OWNER_COMMENTS_SENT", "CONTRACTOR_REPLY_A", "SUBMITTED"}:
        actual_progress = max(actual_progress, 85.0)
    if _is_sent_to_owner(rev_00):
        actual_progress = max(actual_progress, 90.0)
    if any((item.status == "SUBMITTED") for item in revisions):
        actual_progress = max(actual_progress, 100.0)
    # Документ считается полностью завершённым, когда есть AFD-ревизия с AP.
    # В этом случае независимо от статуса ставим 100% — карточка отражает
    # фактическое закрытие документа.
    if any(
        (item.review_code == ReviewCode.AP and (item.issue_purpose or "").upper() == "AFD")
        for item in revisions
    ):
        actual_progress = max(actual_progress, 100.0)

    return RevisionCardRead(
        revision_id=revision.id,
        project_code=mdr.project_code,
        document_num=document.document_num,
        document_title=document.title,
        discipline_code=mdr.discipline_code,
        doc_type=mdr.doc_type,
        category=mdr.category,
        current_revision_code=revision.revision_code,
        current_status=revision.status,
        planned_dev_start=mdr.planned_dev_start,
        planned_issue_date=planned_issue_date,
        actual_first_upload_date=first_upload_at,
        actual_latest_issue_date=latest_issue_at,
        actual_progress_percent=actual_progress,
        can_current_user_raise_comments=can_owner_raise_comments,
        current_user_matrix_role=matrix_role,
        is_observer=(mdr.project_code in _observer_project_codes(db, current_user)),
        lr_reviewer_name=lr_reviewer_name,
        developer_name=developer_name,
        revisions=[RevisionRead.model_validate(item, from_attributes=True) for item in revisions],
        history=history,
    )
