"""Внутренний роутер модуля Vendors (VQM) — /api/v1/mr/...

Доступен ТОЛЬКО аутентифицированным пользователям заказчика (JWT).
Гостевой портал подрядчика живёт в отдельном роутере (PR-3) и сюда
доступа не имеет.

PR-2: CRUD MR + теги + документы заказчика. Приглашения, гостевой портал,
Q&A и отчёт — в следующих PR.
"""

from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import get_db
from app.deps import get_current_user
from app.models import (
    MaterialRequisition,
    MrOwnerFile,
    MrOwnerItem,
    MrStatus,
    MrTag,
    MrVendorItem,
    MrQuestion,
    MrQuestionVisibility,
    Project,
    User,
    UserRole,
    VendorInvitation,
    VendorQuote,
)
from app.schemas import (
    MrCreate,
    MrFeedLink,
    MrOwnerItemCreate,
    MrOwnerItemRead,
    MrOwnerItemUpdate,
    MrRead,
    MrTagCreate,
    MrTagRead,
    MrTagUpdate,
    MrUpdate,
    MrQuestionRead,
    MrQuestionReply,
    MrVendorItemCreate,
    MrVendorItemRead,
    MrVendorItemUpdate,
    QuestionAnswerCreate,
    QuestionVisibilitySet,
    ReqImportResult,
    VendorInvitationCreate,
    VendorInvitationCreated,
    VendorInvitationRead,
    VendorReport,
    VendorReportCell,
    VendorReportRow,
    VendorReportVendor,
)
from app.services import vendor_email, vendor_invitations
from app.services.vendor_security import (
    can_manage_mr,
    ensure_can_manage_mr,
    write_audit,
)

router = APIRouter()
settings = get_settings()

MR_UPLOAD_ROOT = Path(settings.vendor_uploads_root)
MAX_MR_DOC_BYTES = 50 * 1024 * 1024  # 50 MB на документ заказчика


def _get_mr_or_404(db: Session, mr_id: int) -> MaterialRequisition:
    mr = db.query(MaterialRequisition).filter(MaterialRequisition.id == mr_id).first()
    if mr is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="MR not found")
    return mr


def _mr_to_read(db: Session, mr: MaterialRequisition) -> MrRead:
    lr_name = None
    if mr.lr_user_id is not None:
        lr = db.query(User).filter(User.id == mr.lr_user_id).first()
        lr_name = lr.full_name if lr else None
    tags_count = db.query(MrTag).filter(MrTag.mr_id == mr.id).count()
    owner_items = db.query(MrOwnerItem).filter(MrOwnerItem.mr_id == mr.id).all()
    owner_count = len(owner_items)
    filled_item_ids = {
        f.owner_item_id
        for f in db.query(MrOwnerFile.owner_item_id).filter(MrOwnerFile.mr_id == mr.id).all()
    }
    owner_filled = sum(1 for it in owner_items if it.id in filled_item_ids)
    vendor_count = db.query(MrVendorItem).filter(MrVendorItem.mr_id == mr.id).count()
    inv_count = db.query(VendorInvitation).filter(VendorInvitation.mr_id == mr.id).count()
    data = MrRead.model_validate(mr, from_attributes=True)
    data.lr_user_name = lr_name
    data.tags_count = tags_count
    data.owner_items_count = owner_count
    data.owner_items_filled = owner_filled
    data.vendor_items_count = vendor_count
    data.invitations_count = inv_count
    return data


# --------------------------------------------------------------------- MR list
@router.get("/mr", response_model=list[MrRead])
def list_mr(
    project_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Список MR, которыми текущий пользователь может управлять.
    Подрядчики (company_type=contractor) сюда не ходят — у них только
    гостевой портал по ссылке."""
    query = db.query(MaterialRequisition)
    if project_id is not None:
        query = query.filter(MaterialRequisition.project_id == project_id)
    items = query.order_by(MaterialRequisition.id.desc()).all()
    visible = [mr for mr in items if can_manage_mr(db, user=current_user, mr=mr)]
    return [_mr_to_read(db, mr) for mr in visible]


@router.get("/mr/{mr_id}", response_model=MrRead)
def get_mr(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    return _mr_to_read(db, mr)


@router.post("/mr", response_model=MrRead, status_code=status.HTTP_201_CREATED)
def create_mr(
    payload: MrCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Создавать MR может admin или пользователь заказчика с правом управления
    # ревью-матрицей/публикации. Минимальный гейт: не подрядчик.
    if current_user.role != UserRole.admin and current_user.company_type and current_user.company_type.value == "contractor":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Contractors cannot create MR")
    project = db.query(Project).filter(Project.id == payload.project_id).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    existing = db.query(MaterialRequisition).filter(MaterialRequisition.code == payload.code).first()
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="MR code already exists")
    mr = MaterialRequisition(
        project_id=payload.project_id,
        code=payload.code.strip(),
        title=payload.title.strip(),
        description=payload.description,
        lr_user_id=payload.lr_user_id,
        deadline_at=payload.deadline_at,
        currency=(payload.currency or "RUB").strip().upper(),
        status=MrStatus.DRAFT,
        created_by_id=current_user.id,
    )
    db.add(mr)
    db.commit()
    db.refresh(mr)
    return _mr_to_read(db, mr)


@router.patch("/mr/{mr_id}", response_model=MrRead)
def update_mr(
    mr_id: int,
    payload: MrUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    if payload.title is not None:
        mr.title = payload.title.strip()
    if payload.equipment_type is not None:
        mr.equipment_type = payload.equipment_type.strip() or None
    if payload.description is not None:
        mr.description = payload.description
    if payload.lr_user_id is not None:
        mr.lr_user_id = payload.lr_user_id
    if payload.deadline_at is not None:
        mr.deadline_at = payload.deadline_at
    if payload.currency is not None:
        mr.currency = payload.currency.strip().upper()
    if payload.status is not None:
        mr.status = payload.status
    db.commit()
    db.refresh(mr)
    return _mr_to_read(db, mr)


@router.delete("/mr/{mr_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_mr(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    # Удалять можно только черновик без приглашений (не было контакта с
    # подрядчиками). Это защищает историю тендера.
    has_invitations = db.query(VendorInvitation).filter(VendorInvitation.mr_id == mr.id).first() is not None
    if mr.status != MrStatus.DRAFT or has_invitations:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Only a DRAFT MR without invitations can be deleted",
        )
    db.query(MrTag).filter(MrTag.mr_id == mr.id).delete()
    db.query(MrOwnerFile).filter(MrOwnerFile.mr_id == mr.id).delete()
    db.query(MrOwnerItem).filter(MrOwnerItem.mr_id == mr.id).delete()
    db.query(MrVendorItem).filter(MrVendorItem.mr_id == mr.id).delete()
    db.delete(mr)
    db.commit()
    return None


# --------------------------------------------------------------------- Tags
@router.get("/mr/{mr_id}/tags", response_model=list[MrTagRead])
def list_tags(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    items = (
        db.query(MrTag)
        .filter(MrTag.mr_id == mr.id)
        .order_by(MrTag.order_index.asc(), MrTag.id.asc())
        .all()
    )
    return [MrTagRead.model_validate(item, from_attributes=True) for item in items]


@router.post("/mr/{mr_id}/tags", response_model=MrTagRead, status_code=status.HTTP_201_CREATED)
def create_tag(
    mr_id: int,
    payload: MrTagCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    dup = (
        db.query(MrTag)
        .filter(MrTag.mr_id == mr.id, MrTag.tag_code == payload.tag_code.strip())
        .first()
    )
    if dup is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Tag code already exists in this MR")
    tag = MrTag(
        mr_id=mr.id,
        sr_no=payload.sr_no,
        item_no=payload.item_no,
        tag_code=payload.tag_code.strip(),
        name=payload.name.strip(),
        quantity=payload.quantity,
        unit=payload.unit,
        note=payload.note,
        order_index=payload.order_index,
    )
    db.add(tag)
    db.commit()
    db.refresh(tag)
    return MrTagRead.model_validate(tag, from_attributes=True)


@router.patch("/mr/{mr_id}/tags/{tag_id}", response_model=MrTagRead)
def update_tag(
    mr_id: int,
    tag_id: int,
    payload: MrTagUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    tag = db.query(MrTag).filter(MrTag.id == tag_id, MrTag.mr_id == mr.id).first()
    if tag is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tag not found")
    if payload.tag_code is not None:
        tag.tag_code = payload.tag_code.strip()
    if payload.name is not None:
        tag.name = payload.name.strip()
    if payload.quantity is not None:
        tag.quantity = payload.quantity
    if payload.unit is not None:
        tag.unit = payload.unit
    if payload.note is not None:
        tag.note = payload.note
    if payload.order_index is not None:
        tag.order_index = payload.order_index
    db.commit()
    db.refresh(tag)
    return MrTagRead.model_validate(tag, from_attributes=True)


@router.delete("/mr/{mr_id}/tags/{tag_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_tag(
    mr_id: int,
    tag_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    tag = db.query(MrTag).filter(MrTag.id == tag_id, MrTag.mr_id == mr.id).first()
    if tag is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tag not found")
    db.delete(tag)
    db.commit()
    return None


# ------------------------------------------------- Owner checklist (attachments)
def _owner_item_read(db: Session, item: MrOwnerItem) -> MrOwnerItemRead:
    files = db.query(MrOwnerFile).filter(MrOwnerFile.owner_item_id == item.id).order_by(MrOwnerFile.id.asc()).all()
    data = MrOwnerItemRead.model_validate(item, from_attributes=True)
    from app.schemas import MrOwnerFileRead

    data.files = [MrOwnerFileRead.model_validate(f, from_attributes=True) for f in files]
    return data


@router.get("/mr/{mr_id}/owner-items", response_model=list[MrOwnerItemRead])
def list_owner_items(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    items = (
        db.query(MrOwnerItem)
        .filter(MrOwnerItem.mr_id == mr.id)
        .order_by(MrOwnerItem.order_index.asc(), MrOwnerItem.id.asc())
        .all()
    )
    return [_owner_item_read(db, it) for it in items]


@router.post("/mr/{mr_id}/owner-items", response_model=MrOwnerItemRead, status_code=status.HTTP_201_CREATED)
def create_owner_item(
    mr_id: int,
    payload: MrOwnerItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = MrOwnerItem(
        mr_id=mr.id,
        att_no=payload.att_no,
        category=payload.category,
        title=payload.title.strip(),
        doc_number=payload.doc_number,
        rev=payload.rev,
        is_required=payload.is_required,
        allow_questions=payload.allow_questions,
        is_group=payload.is_group,
        order_index=payload.order_index,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _owner_item_read(db, item)


@router.patch("/mr/{mr_id}/owner-items/{item_id}", response_model=MrOwnerItemRead)
def update_owner_item(
    mr_id: int,
    item_id: int,
    payload: MrOwnerItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = db.query(MrOwnerItem).filter(MrOwnerItem.id == item_id, MrOwnerItem.mr_id == mr.id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Owner item not found")
    for field in ("att_no", "category", "title", "doc_number", "rev", "is_required", "allow_questions", "order_index"):
        value = getattr(payload, field)
        if value is not None:
            setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return _owner_item_read(db, item)


@router.delete("/mr/{mr_id}/owner-items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_owner_item(
    mr_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = db.query(MrOwnerItem).filter(MrOwnerItem.id == item_id, MrOwnerItem.mr_id == mr.id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Owner item not found")
    for f in db.query(MrOwnerFile).filter(MrOwnerFile.owner_item_id == item.id).all():
        try:
            Path(f.file_path).unlink(missing_ok=True)
        except OSError:
            pass
        db.delete(f)
    db.delete(item)
    db.commit()
    return None


@router.post("/mr/{mr_id}/owner-items/{item_id}/files", response_model=MrOwnerItemRead, status_code=status.HTTP_201_CREATED)
def upload_owner_file(
    mr_id: int,
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = db.query(MrOwnerItem).filter(MrOwnerItem.id == item_id, MrOwnerItem.mr_id == mr.id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Owner item not found")
    raw = file.file.read()
    if len(raw) > MAX_MR_DOC_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large")
    original_name = file.filename or "document"
    safe_name = f"{uuid4().hex}_{Path(original_name).name.replace(' ', '_')}"
    destination_dir = MR_UPLOAD_ROOT / "mr_owner" / str(mr.id) / str(item.id)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / safe_name
    destination.write_bytes(raw)
    db.add(
        MrOwnerFile(
            owner_item_id=item.id,
            mr_id=mr.id,
            file_path=str(destination),
            file_name=original_name,
            mime=file.content_type,
            size_bytes=len(raw),
            uploaded_by_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(item)
    return _owner_item_read(db, item)


@router.delete("/mr/{mr_id}/owner-files/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_owner_file(
    mr_id: int,
    file_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    f = db.query(MrOwnerFile).filter(MrOwnerFile.id == file_id, MrOwnerFile.mr_id == mr.id).first()
    if f is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
    try:
        Path(f.file_path).unlink(missing_ok=True)
    except OSError:
        pass
    db.delete(f)
    db.commit()
    return None


# ------------------------------------------------- Vendor checklist (template)
@router.get("/mr/{mr_id}/vendor-items", response_model=list[MrVendorItemRead])
def list_vendor_items(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    items = (
        db.query(MrVendorItem)
        .filter(MrVendorItem.mr_id == mr.id)
        .order_by(MrVendorItem.order_index.asc(), MrVendorItem.id.asc())
        .all()
    )
    return [MrVendorItemRead.model_validate(it, from_attributes=True) for it in items]


@router.post("/mr/{mr_id}/vendor-items", response_model=MrVendorItemRead, status_code=status.HTTP_201_CREATED)
def create_vendor_item(
    mr_id: int,
    payload: MrVendorItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = MrVendorItem(
        mr_id=mr.id,
        section=payload.section,
        category=payload.category,
        code=payload.code,
        title=payload.title.strip(),
        purpose=payload.purpose,
        with_bid=payload.with_bid,
        is_required=payload.is_required,
        allow_questions=payload.allow_questions,
        order_index=payload.order_index,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return MrVendorItemRead.model_validate(item, from_attributes=True)


@router.patch("/mr/{mr_id}/vendor-items/{item_id}", response_model=MrVendorItemRead)
def update_vendor_item(
    mr_id: int,
    item_id: int,
    payload: MrVendorItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = db.query(MrVendorItem).filter(MrVendorItem.id == item_id, MrVendorItem.mr_id == mr.id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vendor item not found")
    for field in ("section", "category", "code", "title", "purpose", "with_bid", "is_required", "allow_questions", "order_index"):
        value = getattr(payload, field)
        if value is not None:
            setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return MrVendorItemRead.model_validate(item, from_attributes=True)


@router.delete("/mr/{mr_id}/vendor-items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_vendor_item(
    mr_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    item = db.query(MrVendorItem).filter(MrVendorItem.id == item_id, MrVendorItem.mr_id == mr.id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vendor item not found")
    db.delete(item)
    db.commit()
    return None


# --------------------------------------------------------------------- Invitations
@router.get("/mr/{mr_id}/invitations", response_model=list[VendorInvitationRead])
def list_invitations(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    items = (
        db.query(VendorInvitation)
        .filter(VendorInvitation.mr_id == mr.id)
        .order_by(VendorInvitation.id.asc())
        .all()
    )
    return [VendorInvitationRead.model_validate(i, from_attributes=True) for i in items]


@router.post(
    "/mr/{mr_id}/invitations",
    response_model=VendorInvitationCreated,
    status_code=status.HTTP_201_CREATED,
)
def create_invitation(
    mr_id: int,
    payload: VendorInvitationCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    # Приглашать подрядчиков можно только когда приём заявок открыт —
    # иначе подрядчик получит ссылку, но не сможет войти (портал блокирует
    # не-OPEN статусы). Сначала откройте приём, затем приглашайте.
    if mr.status != MrStatus.OPEN:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Open the MR for bids before inviting vendors",
        )
    # Бизнес-ограничение: не больше N подрядчиков на MR.
    active_count = (
        db.query(VendorInvitation)
        .filter(VendorInvitation.mr_id == mr.id, VendorInvitation.revoked_at.is_(None))
        .count()
    )
    if active_count >= settings.vendor_max_invitations_per_mr:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Maximum {settings.vendor_max_invitations_per_mr} vendors per MR",
        )
    token = vendor_invitations.generate_invitation_token()
    inv = VendorInvitation(
        mr_id=mr.id,
        vendor_company_name=payload.vendor_company_name.strip(),
        vendor_contact_email=payload.vendor_contact_email.strip().lower(),
        token_hash=vendor_invitations.hash_secret(token),
        expires_at=payload.expires_at or mr.deadline_at,
        created_by_id=current_user.id,
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)

    # Ссылку строим от Origin админского запроса — чтобы в локальной сети она
    # вела на IP сайта, а не на localhost (если PUBLIC_BASE_URL не задан явно).
    origin = request.headers.get("origin") or request.headers.get("referer")
    link = vendor_invitations.build_invitation_link(inv.id, token, origin=origin)
    # Письмо-приглашение со ссылкой (код придёт отдельно при входе).
    email_sent = False
    email_note: str | None = None
    try:
        email_sent = vendor_email.send_invitation_link(
            to=inv.vendor_contact_email,
            company_name=inv.vendor_company_name,
            mr_code=mr.code,
            link=link,
        )
        if not email_sent:
            email_note = "SMTP не настроен — письмо не отправлено. Передайте ссылку вручную."
    except Exception as exc:  # noqa: BLE001 — письмо не должно ронять создание
        email_note = f"Ошибка отправки письма: {str(exc)[:150]}"
    write_audit(
        db,
        action="invitation_created",
        mr_id=mr.id,
        invitation_id=inv.id,
        actor_user_id=current_user.id,
        request=request,
        summary=f"vendor={inv.vendor_company_name}",
    )

    base = VendorInvitationRead.model_validate(inv, from_attributes=True)
    return VendorInvitationCreated(
        **base.model_dump(),
        invitation_link=link,
        token=token,
        email_sent=email_sent,
        email_note=email_note,
    )


@router.post("/mr/{mr_id}/invitations/{invitation_id}/revoke", response_model=VendorInvitationRead)
def revoke_invitation(
    mr_id: int,
    invitation_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import datetime as _dt

    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    inv = (
        db.query(VendorInvitation)
        .filter(VendorInvitation.id == invitation_id, VendorInvitation.mr_id == mr.id)
        .first()
    )
    if inv is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invitation not found")
    inv.revoked_at = _dt.utcnow()
    db.commit()
    db.refresh(inv)
    write_audit(
        db,
        action="invitation_revoked",
        mr_id=mr.id,
        invitation_id=inv.id,
        actor_user_id=current_user.id,
        request=request,
    )
    return VendorInvitationRead.model_validate(inv, from_attributes=True)


# --------------------------------------------------------------------- REQ import
@router.post("/mr/import", response_model=ReqImportResult, status_code=status.HTTP_201_CREATED)
def import_req(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Импорт REQ (.docx) → создаёт MR со структурой (теги, чек-лист
    заказчика, чек-лист подрядчика). Файлы и приглашения добавляются
    отдельно. Всё созданное редактируемо (корректировки)."""
    from app.services.req_importer import parse_req

    if current_user.role != UserRole.admin and current_user.company_type and current_user.company_type.value == "contractor":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Contractors cannot import REQ")
    project = db.query(Project).filter(Project.id == project_id).first()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

    raw = file.file.read()
    tmp = MR_UPLOAD_ROOT / "req_import" / (uuid4().hex + ".docx")
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(raw)
    try:
        parsed = parse_req(str(tmp), file.filename or "req.docx")
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass

    code = parsed.get("req_number") or f"MR-{uuid4().hex[:8].upper()}"
    # Уникальность кода MR
    if db.query(MaterialRequisition).filter(MaterialRequisition.code == code).first():
        code = f"{code}-{uuid4().hex[:4]}"

    mr = MaterialRequisition(
        project_id=project_id,
        code=code,
        title=parsed.get("title") or "Requisition",
        equipment_type=parsed.get("equipment_type"),
        req_number=parsed.get("req_number"),
        discipline_code=parsed.get("discipline_code"),
        status=MrStatus.DRAFT,
        created_by_id=current_user.id,
    )
    db.add(mr)
    db.flush()

    # Коды тегов должны быть уникальны в пределах MR (UNIQUE-констрейнт).
    # В некоторых REQ позиции дают одинаковый tag_code (обрезанное имя) —
    # дедуплицируем, добавляя суффикс #N.
    seen_codes: dict[str, int] = {}
    for i, t in enumerate(parsed.get("tags", [])):
        base_code = (t.get("tag_code") or t.get("name") or "TAG")[:110].strip() or "TAG"
        if base_code in seen_codes:
            seen_codes[base_code] += 1
            code = f"{base_code}#{seen_codes[base_code]}"
        else:
            seen_codes[base_code] = 1
            code = base_code
        db.add(
            MrTag(
                mr_id=mr.id, order_index=i, sr_no=t.get("sr_no"), item_no=t.get("item_no"),
                tag_code=code[:120], name=(t.get("name") or "")[:255],
                quantity=t.get("quantity"), unit=t.get("unit"), note=t.get("note"),
            )
        )
    from app.models import MrOwnerItemCategory as _OIC, MrVendorItemSection as _VIS

    for i, o in enumerate(parsed.get("owner_items", [])):
        try:
            cat = _OIC(o.get("category", "OTHER"))
        except ValueError:
            cat = _OIC.OTHER
        db.add(
            MrOwnerItem(
                mr_id=mr.id, order_index=i, att_no=o.get("att_no"), category=cat,
                title=(o.get("title") or "")[:500], doc_number=o.get("doc_number"), rev=o.get("rev"),
                is_required=bool(o.get("is_required", True)), allow_questions=bool(o.get("allow_questions", False)),
                is_group=bool(o.get("is_group", False)),
            )
        )
    for v in parsed.get("vendor_items", []):
        try:
            sec = _VIS(v.get("section", "RFD"))
        except ValueError:
            sec = _VIS.RFD
        db.add(
            MrVendorItem(
                mr_id=mr.id, order_index=v.get("order_index", 0), section=sec, category=v.get("category"),
                code=v.get("code"), title=(v.get("title") or "")[:500], purpose=v.get("purpose"),
                with_bid=bool(v.get("with_bid", False)), is_required=bool(v.get("is_required", True)),
                allow_questions=bool(v.get("allow_questions", False)),
            )
        )
    db.commit()
    return ReqImportResult(
        mr_id=mr.id, code=code,
        tags_created=len(parsed.get("tags", [])),
        owner_items_created=len(parsed.get("owner_items", [])),
        vendor_items_created=len(parsed.get("vendor_items", [])),
    )


# --------------------------------------------------------------------- Report
def _build_report(db: Session, mr: MaterialRequisition) -> VendorReport:
    """Сводка теги × подрядчики × цены. Только активные приглашения."""
    invitations = (
        db.query(VendorInvitation)
        .filter(VendorInvitation.mr_id == mr.id, VendorInvitation.revoked_at.is_(None))
        .order_by(VendorInvitation.id.asc())
        .all()
    )
    tags = (
        db.query(MrTag)
        .filter(MrTag.mr_id == mr.id)
        .order_by(MrTag.order_index.asc(), MrTag.id.asc())
        .all()
    )
    # quotes[invitation_id][tag_id] = VendorQuote
    inv_ids = [i.id for i in invitations]
    quotes = (
        db.query(VendorQuote)
        .filter(VendorQuote.invitation_id.in_(inv_ids))
        .all()
        if inv_ids
        else []
    )
    qmap: dict[tuple[int, int], VendorQuote] = {(q.invitation_id, q.tag_id): q for q in quotes}

    # Итог по подрядчику = сумма его цен по всем тегам (где есть цена).
    totals: dict[int, float | None] = {}
    for inv in invitations:
        vals = [qmap[(inv.id, t.id)].price for t in tags if qmap.get((inv.id, t.id)) and qmap[(inv.id, t.id)].price is not None]
        totals[inv.id] = round(sum(vals), 2) if vals else None

    vendors = [
        VendorReportVendor(
            invitation_id=inv.id,
            company_name=inv.vendor_company_name,
            submitted=inv.email_verified_at is not None,
            total_price=totals.get(inv.id),
        )
        for inv in invitations
    ]

    rows: list[VendorReportRow] = []
    for t in tags:
        cells = []
        prices_with_inv = []
        for inv in invitations:
            q = qmap.get((inv.id, t.id))
            price = q.price if q else None
            note = q.note if q else None
            cells.append(VendorReportCell(invitation_id=inv.id, price=price, note=note))
            if price is not None:
                prices_with_inv.append((price, inv.id))
        min_inv = min(prices_with_inv)[1] if prices_with_inv else None
        rows.append(
            VendorReportRow(
                tag_id=t.id, sr_no=t.sr_no, item_no=t.item_no, name=t.name,
                quantity=t.quantity, unit=t.unit, cells=cells, min_invitation_id=min_inv,
            )
        )
    return VendorReport(mr_id=mr.id, code=mr.code, currency=mr.currency, vendors=vendors, rows=rows)


@router.get("/mr/{mr_id}/report", response_model=VendorReport)
def get_report(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    return _build_report(db, mr)


@router.get("/mr/{mr_id}/report.xlsx")
def export_report_xlsx(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    report = _build_report(db, mr)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение"
    bold = Font(bold=True)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    header = ["Sr", "Item No", "Наименование", "Кол-во", "Ед."] + [v.company_name for v in report.vendors]
    ws.append(header)
    for c in ws[1]:
        c.font = bold

    for row in report.rows:
        line = [row.sr_no or "", row.item_no or "", row.name, row.quantity, row.unit or ""]
        line += [next((c.price for c in row.cells if c.invitation_id == v.invitation_id), None) for v in report.vendors]
        ws.append(line)
        # подсветка минимальной цены
        if row.min_invitation_id is not None:
            col = 5 + [v.invitation_id for v in report.vendors].index(row.min_invitation_id) + 1
            ws.cell(row=ws.max_row, column=col).fill = green

    total_line = ["", "", "ИТОГО", "", ""] + [v.total_price for v in report.vendors]
    ws.append(total_line)
    for c in ws[ws.max_row]:
        c.font = bold

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"report_{mr.code}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# --------------------------------------------------------------------- Q&A (owner side)
def _q_author_label(db: Session, q: MrQuestion) -> str:
    if q.invitation_id is None:
        return "Заказчик"
    inv = db.query(VendorInvitation).filter(VendorInvitation.id == q.invitation_id).first()
    return inv.vendor_company_name if inv else "Поставщик"


def _owner_question_read(db: Session, q: MrQuestion) -> MrQuestionRead:
    replies = (
        db.query(MrQuestion).filter(MrQuestion.parent_id == q.id).order_by(MrQuestion.id.asc()).all()
    )
    return MrQuestionRead(
        id=q.id,
        body=q.body,
        author_label=_q_author_label(db, q),
        is_public=q.visibility == MrQuestionVisibility.PUBLIC,
        mr_owner_item_id=q.mr_owner_item_id,
        mr_vendor_item_id=q.mr_vendor_item_id,
        created_at=q.created_at,
        replies=[
            MrQuestionReply(
                id=r.id, body=r.body, is_owner=r.invitation_id is None,
                author_label=_q_author_label(db, r), created_at=r.created_at,
            )
            for r in replies
        ],
    )


@router.get("/mr/{mr_id}/questions", response_model=list[MrQuestionRead])
def list_questions(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """LR/admin видит ВСЕ вопросы подрядчиков по MR (с именами компаний)."""
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    questions = (
        db.query(MrQuestion)
        .filter(MrQuestion.mr_id == mr.id, MrQuestion.parent_id.is_(None))
        .order_by(MrQuestion.id.desc())
        .all()
    )
    return [_owner_question_read(db, q) for q in questions]


@router.post("/mr/{mr_id}/questions/{question_id}/answer", response_model=MrQuestionRead)
def answer_question(
    mr_id: int,
    question_id: int,
    payload: QuestionAnswerCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """LR/admin отвечает на вопрос подрядчика. Уведомление автору."""
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    q = db.query(MrQuestion).filter(MrQuestion.id == question_id, MrQuestion.mr_id == mr.id, MrQuestion.parent_id.is_(None)).first()
    if q is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")
    body = (payload.body or "").strip()
    if not body:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty answer")
    reply = MrQuestion(
        mr_id=mr.id, invitation_id=None, parent_id=q.id, body=body,
        visibility=q.visibility, answered_by_id=current_user.id, answered_at=__import__("datetime").datetime.utcnow(),
    )
    db.add(reply)
    db.commit()
    write_audit(db, action="owner_answered_question", mr_id=mr.id, actor_user_id=current_user.id, request=request, summary=f"q={q.id}")
    db.refresh(q)
    return _owner_question_read(db, q)


@router.post("/mr/{mr_id}/questions/{question_id}/visibility", response_model=MrQuestionRead)
def set_question_visibility(
    mr_id: int,
    question_id: int,
    payload: QuestionVisibilitySet,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Сделать вопрос+ответ публичным (виден всем подрядчикам MR) или приватным."""
    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    q = db.query(MrQuestion).filter(MrQuestion.id == question_id, MrQuestion.mr_id == mr.id, MrQuestion.parent_id.is_(None)).first()
    if q is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")
    vis = MrQuestionVisibility.PUBLIC if payload.public else MrQuestionVisibility.PRIVATE
    q.visibility = vis
    # Реплики наследуют видимость родителя.
    for r in db.query(MrQuestion).filter(MrQuestion.parent_id == q.id).all():
        r.visibility = vis
    db.commit()
    db.refresh(q)
    return _owner_question_read(db, q)


# ------------------------------------------------- Интеграция с FEED
@router.get("/mr/{mr_id}/feed-links", response_model=list[MrFeedLink])
def mr_feed_links(
    mr_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Сверка документов REQ (чек-лист заказчика) с базой FEED по номеру
    документа: подтянут ли документ в FEED и совпадает ли крайняя ревизия.
    rev_mismatch=True — в FEED ревизия новее/другая, надо проверить REQ."""
    from app.models import FeedDocument
    from app.services.feed_import import split_rev

    mr = _get_mr_or_404(db, mr_id)
    ensure_can_manage_mr(db, user=current_user, mr=mr)
    items = (
        db.query(MrOwnerItem)
        .filter(MrOwnerItem.mr_id == mr.id, MrOwnerItem.doc_number.isnot(None))
        .all()
    )
    links: list[MrFeedLink] = []
    for item in items:
        raw_no = (item.doc_number or "").strip().upper()
        # Attachment No.N — внутренняя нумерация REQ, в FEED не ищем.
        if not raw_no or raw_no.startswith("ATTACHMENT"):
            continue
        base_no, item_rev = split_rev(raw_no)
        item_rev = item_rev or (item.rev or None)
        feed_doc = (
            db.query(FeedDocument)
            .filter(FeedDocument.project_id == mr.project_id, FeedDocument.doc_number == base_no)
            .first()
        )
        mismatch = bool(
            feed_doc
            and feed_doc.latest_rev
            and item_rev
            and feed_doc.latest_rev != item_rev
        )
        feed_file = None
        if feed_doc:
            from app.models import FeedFile, FeedFileKind

            feed_file = (
                db.query(FeedFile)
                .filter(FeedFile.feed_document_id == feed_doc.id, FeedFile.kind == FeedFileKind.REVISION)
                .order_by(FeedFile.id.desc())
                .first()
            )
        links.append(
            MrFeedLink(
                owner_item_id=item.id,
                owner_doc_number=raw_no,
                owner_rev=item_rev,
                feed_document_id=feed_doc.id if feed_doc else None,
                feed_latest_rev=feed_doc.latest_rev if feed_doc else None,
                feed_file_id=feed_file.id if feed_file else None,
                feed_file_name=feed_file.file_name if feed_file else None,
                rev_mismatch=mismatch,
            )
        )
    return links
